from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse, HttpResponse
from django.db.models import F, Prefetch, Sum, Count
import json
from .models import projectMapping, settlementwithCoordinates, Access, nawecinfrastructure
from .forms import MappingForm, NAWECInfrastructureForm, settlementwithCoordinatesForm
from .filters import ProjectMappingFilter
from django.contrib import messages
from PIU_Financial_mgt.models import Project, Donor
from social_and_env.models import Settlement, Regions
from setup.models import YEAR, Districts
from django.views.decorators.http import require_http_methods
from django.contrib.auth.decorators import login_required
from django.core.cache import cache
from django.core.paginator import Paginator
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.units import inch
from io import BytesIO
from datetime import datetime


def working_map(request):
    """Working map implementation"""
    
    project_communities = projectMapping.objects.select_related(
        'region', 'district', 'settlement', 'profile_year', 'access'
    ).prefetch_related('project', 'donor').filter(
        Latitude__isnull=False,
        Longitude__isnull=False
    ).all()
    
    project_data = []
    for community in project_communities:
        try:
            if community.Latitude and community.Longitude:
                projects = community.project.all()
                project_names = [str(p.project) for p in projects]
                
                project_data.append({
                    'id': community.id,
                    'latitude': float(community.Latitude),
                    'longitude': float(community.Longitude),
                    'settlement': community.settlement.settlement_name if community.settlement else 'Unknown',
                    'region': community.region.region_name if community.region else 'Unknown', 
                    'district': community.district.district_name if community.district else 'Unknown',
                    'projects': project_names,
                    'total_households': community.Total_No_of_Households or 0,
                    'connected_households': community.no_of_connected_household or 0,
                    'access_type': community.access.access_type if community.access else 'Unknown',
                    'year': community.profile_year.profile_year if community.profile_year else 'Unknown',
                })
        except Exception as e:
            continue
    
    context = {
        'total_communities': project_communities.count(),
        'project_data_json': json.dumps(project_data),
    }
    
    return render(request, 'PIU_Mapping_project_Sites/working_map.html', context)

def simple_map(request):
    """Simplified map view with basic Leaflet implementation"""
    
    project_communities = projectMapping.objects.select_related(
        'region', 'district', 'settlement', 'profile_year', 'access'
    ).prefetch_related('project', 'donor').filter(
        Latitude__isnull=False,
        Longitude__isnull=False
    ).all()
    
    # Generate simple project data for JavaScript
    project_data = []
    for community in project_communities:
        try:
            if community.Latitude and community.Longitude:
                projects = community.project.all()
                project_names = [str(p.project) for p in projects]
                
                project_data.append({
                    'id': community.id,
                    'latitude': float(community.Latitude),
                    'longitude': float(community.Longitude),
                    'settlement': community.settlement.settlement_name if community.settlement else 'Unknown',
                    'region': community.region.region_name if community.region else 'Unknown', 
                    'district': community.district.district_name if community.district else 'Unknown',
                    'projects': project_names,
                    'total_households': community.Total_No_of_Households or 0,
                    'connected_households': community.no_of_connected_household or 0,
                })
        except Exception as e:
            print(f"Error processing community {community.id}: {str(e)}")
            continue
    
    context = {
        'total_communities': project_communities.count(),
        'project_data_json': json.dumps(project_data),
    }
    
    return render(request, 'PIU_Mapping_project_Sites/simple_map.html', context)

def index(request):
    """Main mapping view displaying authentic NAWEC PIU project coordinates"""
    
    project_communities = projectMapping.objects.select_related(
        'region', 'district', 'settlement', 'profile_year', 'access'
    ).prefetch_related('project', 'donor').filter(
        Latitude__isnull=False,
        Longitude__isnull=False
    ).all()
    
    # Debug: Print number of communities found
    print(f"DEBUG: Found {project_communities.count()} communities with coordinates")
    
    # Check for coordinate parameters to focus on specific location
    focus_lat = request.GET.get('lat')
    focus_lng = request.GET.get('lng')
    mapping_id = request.GET.get('mappingId')
    
    if focus_lat and focus_lng:
        try:
            center_lat = float(focus_lat)
            center_lng = float(focus_lng)
            zoom_level = 16  # Zoom in when focusing on specific location
        except (ValueError, TypeError):
            center_lat = 13.4667  # Default to Gambia center
            center_lng = -15.3100
            zoom_level = 9
    else:
        center_lat = 13.4667  # Default to Gambia center
        center_lng = -15.3100
        zoom_level = 9
    
    print(f"DEBUG: Map created successfully centered at [{center_lat}, {center_lng}]")

    # Generate project data for JavaScript map
    project_data = []
    
    for community in project_communities:
        try:
            if community.Latitude and community.Longitude:
                projects = community.project.all()
                project_names = [str(p.project) for p in projects]
                
                # Check if this is the focused location from URL parameters
                is_focused = False
                if focus_lat and focus_lng:
                    try:
                        focus_lat_f = float(focus_lat)
                        focus_lng_f = float(focus_lng)
                        # Check if coordinates match (within small tolerance)
                        if abs(float(community.Latitude) - focus_lat_f) < 0.0001 and abs(float(community.Longitude) - focus_lng_f) < 0.0001:
                            is_focused = True
                    except (ValueError, TypeError):
                        pass
                
                # Also check mapping ID match for precise highlighting
                if mapping_id and str(community.id) == str(mapping_id):
                    is_focused = True
                
                # Build project data object
                project_data.append({
                    'id': community.id,
                    'latitude': float(community.Latitude),
                    'longitude': float(community.Longitude),
                    'settlement': community.settlement.settlement_name if community.settlement else 'Unknown',
                    'region': community.region.region_name if community.region else 'Unknown',
                    'district': community.district.district_name if community.district else 'Unknown',
                    'projects': project_names,
                    'total_households': community.Total_No_of_Households or 0,
                    'connected_households': community.no_of_connected_household or 0,
                    'access_type': community.access.access_type if community.access else 'Unknown',
                    'profile_year': community.profile_year.profile_year if community.profile_year else 'Unknown',
                    'is_focused': is_focused
                })
        except Exception as e:
            print(f"DEBUG: Error processing community {community.id}: {str(e)}")
            continue
    
    print(f"DEBUG: Generated data for {len(project_data)} project locations")

    total_communities = project_communities.count()
    total_households = project_communities.aggregate(total=Sum('Total_No_of_Households'))['total'] or 0
    total_connected = project_communities.aggregate(total=Sum('no_of_connected_household'))['total'] or 0

    context = {
        'total_communities': total_communities,
        'total_households': total_households,
        'total_connected': total_connected,
        'markers_added': len(project_data),
        'overall_access_rate': (total_connected / total_households * 100) if total_households else 0,
        'project_data_json': json.dumps(project_data),
        'center_lat': center_lat,
        'center_lng': center_lng,
        'zoom_level': zoom_level
    }

    return render(request, 'PIU_Mapping_project_Sites/index.html', context)


def mapping_dashboard(request):
    """Mapping dashboard - route to main map view"""
    return index(request)


@login_required
def mapping_list(request):
    """List view of all project mappings with filtering and export functionality"""
    
    # Get all mappings with related fields
    mappings_queryset = projectMapping.objects.select_related(
        'region', 'district', 'settlement', 'profile_year', 'access'
    ).prefetch_related('project', 'donor').order_by('region__region_name', 'district__district_name', 'settlement__settlement_name')
    
    # Apply filters
    mapping_filter = ProjectMappingFilter(request.GET, queryset=mappings_queryset)
    filtered_mappings = mapping_filter.qs
    
    # Handle export requests
    export_format = request.GET.get('export')
    if export_format == 'excel':
        return export_mappings_excel(filtered_mappings)
    elif export_format == 'pdf':
        return export_mappings_pdf(filtered_mappings)
    
    # Add pagination
    paginator = Paginator(filtered_mappings, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Statistics for filtered data
    total_mappings = filtered_mappings.count()
    total_households = sum(mapping.Total_No_of_Households or 0 for mapping in filtered_mappings)
    total_connected = sum(mapping.no_of_connected_household or 0 for mapping in filtered_mappings)
    
    # Get all dropdown data directly from models
    from setup.models import Regions, Districts, YEAR, Access
    from social_and_env.models import Settlement
    from PIU_Financial_mgt.models import Project, Donor
    
    all_regions = Regions.objects.all().order_by('region_name')
    all_districts = Districts.objects.all().order_by('district_name')
    all_settlements = Settlement.objects.all().order_by('settlement_name')
    all_projects = Project.objects.all().order_by('project')
    all_donors = Donor.objects.all().order_by('name')
    all_years = YEAR.objects.all().order_by('profile_year')
    all_access_types = Access.objects.all().order_by('access_type')
    
    context = {
        'page_obj': page_obj,
        'mappings': page_obj.object_list,
        'filter': mapping_filter,
        'total_mappings': total_mappings,
        'total_households': total_households,
        'total_connected': total_connected,
        'connection_rate': round((total_connected / total_households * 100) if total_households > 0 else 0, 2),
        # Add dropdown data to context
        'all_regions': all_regions,
        'all_districts': all_districts,
        'all_settlements': all_settlements,
        'all_projects': all_projects,
        'all_donors': all_donors,
        'all_years': all_years,
        'all_access_types': all_access_types,
    }
    
    return render(request, 'PIU_Mapping_project_Sites/mapping_list.html', context)


def mapping_detail(request, pk):
    """Detail view of a specific project mapping"""
    mapping = get_object_or_404(projectMapping, pk=pk)
    
    context = {
        'mapping': mapping,
        'projects': mapping.project.all(),
        'donors': mapping.donor.all(),
    }
    
    return render(request, 'PIU_Mapping_project_Sites/mapping_detail.html', context)


def export_mappings_excel(mappings_queryset):
    """Export filtered project mappings to Excel"""
    
    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Project Mapping Export"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = [
        'Region', 'District', 'Settlement', 'Projects', 'Donors', 'Year',
        'Access Type', 'Total Households', 'Connected Households', 'Customer Connections',
        'Connection Rate (%)', 'Latitude', 'Longitude'
    ]
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Write data
    for row, mapping in enumerate(mappings_queryset, 2):
        projects = ', '.join([str(p.project) for p in mapping.project.all()])
        donors = ', '.join([str(d.name) for d in mapping.donor.all()])
        connection_rate = round((mapping.no_of_connected_household / mapping.Total_No_of_Households * 100) if mapping.Total_No_of_Households > 0 else 0, 2)
        
        row_data = [
            mapping.region.region_name if mapping.region else '',
            mapping.district.district_name if mapping.district else '',
            mapping.settlement.settlement_name if mapping.settlement else '',
            projects,
            donors,
            mapping.profile_year.profile_year if mapping.profile_year else '',
            mapping.access.access_type if mapping.access else '',
            mapping.Total_No_of_Households or 0,
            mapping.no_of_connected_household or 0,
            mapping.no_of_customer_connections or 0,
            connection_rate,
            mapping.Latitude or 0,
            mapping.Longitude or 0,
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = border
            if isinstance(value, (int, float)) and col >= 8:  # Numeric columns
                cell.alignment = Alignment(horizontal='right')
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="project_mappings_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    wb.save(response)
    return response


def export_mappings_pdf(mappings_queryset):
    """Export filtered project mappings to PDF"""
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), 
                          rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=18)
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#366092'),
        alignment=1,  # Center alignment
        spaceAfter=20
    )
    
    # Build PDF content
    elements = []
    
    # Title
    title = Paragraph("Project Mapping Export Report", title_style)
    elements.append(title)
    elements.append(Spacer(1, 12))
    
    # Export info
    export_info = Paragraph(f"<b>Export Date:</b> {datetime.now().strftime('%B %d, %Y at %I:%M %p')}<br/>"
                           f"<b>Total Records:</b> {mappings_queryset.count()}", styles['Normal'])
    elements.append(export_info)
    elements.append(Spacer(1, 20))
    
    # Table data
    table_data = [
        ['Region', 'District', 'Settlement', 'Projects', 'Access Type', 
         'Total HH', 'Connected HH', 'Connection %']
    ]
    
    for mapping in mappings_queryset:
        projects = ', '.join([str(p.project)[:20] + '...' if len(str(p.project)) > 20 else str(p.project) for p in mapping.project.all()])
        connection_rate = round((mapping.no_of_connected_household / mapping.Total_No_of_Households * 100) if mapping.Total_No_of_Households > 0 else 0, 1)
        
        table_data.append([
            mapping.region.region_name[:15] if mapping.region else '',
            mapping.district.district_name[:15] if mapping.district else '',
            mapping.settlement.settlement_name[:20] if mapping.settlement else '',
            projects[:25] + '...' if len(projects) > 25 else projects,
            mapping.access.access_type[:15] if mapping.access else '',
            str(mapping.Total_No_of_Households or 0),
            str(mapping.no_of_connected_household or 0),
            f"{connection_rate}%"
        ])
    
    # Create table
    table = Table(table_data, colWidths=[1*inch, 1*inch, 1.2*inch, 1.5*inch, 1*inch, 0.7*inch, 0.8*inch, 0.8*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    
    elements.append(table)
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="project_mappings_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
    
    return response


def add_mapping(request):
    """Add new mapping view"""
    if request.method == 'POST':
        form = MappingForm(request.POST)
        if form.is_valid():
            mapping = form.save(commit=False)
            mapping.loginUser = request.user
            mapping.save()
            form.save_m2m()
            messages.success(request, 'Mapping added successfully!')
            return redirect('PIU_Mapping_project_Sites:mapping_list')
    else:
        form = MappingForm()
    
    context = {
        'form': form,
        'title': 'Add New Mapping'
    }
    
    return render(request, 'PIU_Mapping_project_Sites/mapping_form.html', context)


def update_mapping(request, pk):
    """Update mapping view"""
    mapping = get_object_or_404(projectMapping, pk=pk)
    
    if request.method == 'POST':
        form = MappingForm(request.POST, instance=mapping)
        if form.is_valid():
            form.save()
            messages.success(request, 'Mapping updated successfully!')
            
            # Check if this was opened from a popup (from_popup parameter)
            if request.GET.get('from_popup'):
                return render(request, 'PIU_Mapping_project_Sites/popup_close.html', {
                    'message': 'Mapping updated successfully! Returning to map...'
                })
            
            return redirect('PIU_Mapping_project_Sites:mapping_list')
    else:
        form = MappingForm(instance=mapping)
    
    context = {
        'form': form,
        'mapping': mapping,
        'title': 'Update Mapping',
        'from_popup': request.GET.get('from_popup', False)
    }
    
    return render(request, 'PIU_Mapping_project_Sites/mapping_form.html', context)


def delete_mapping(request, pk):
    """Delete mapping view"""
    mapping = get_object_or_404(projectMapping, pk=pk)
    
    if request.method == 'POST':
        mapping.delete()
        messages.success(request, 'Mapping deleted successfully!')
        return redirect('PIU_Mapping_project_Sites:mapping_list')
    
    context = {
        'mapping': mapping,
        'title': 'Delete Mapping'
    }
    
    return render(request, 'PIU_Mapping_project_Sites/mapping_confirm_delete.html', context)


# AJAX endpoints for cascading dropdowns
def load_districts(request):
    """Load districts based on selected region"""
    region_id = request.GET.get('region_id')
    districts = Districts.objects.filter(region_code=region_id).order_by('district_name')
    return JsonResponse({'districts': [{'id': d.district_code, 'name': d.district_name} for d in districts]})


def load_settlement(request):
    """Load settlements based on selected district"""
    district_id = request.GET.get('district_id')
    settlements = Settlement.objects.filter(district_code=district_id).order_by('settlement_name')
    return JsonResponse({'settlements': [{'id': s.settlement_code, 'name': s.settlement_name} for s in settlements]})


# Placeholder functions for compatibility
def indexl(request):
    """Leaflet map placeholder - redirect to main map"""
    return redirect('PIU_Mapping_project_Sites:index')


def offline_map(request):
    """Offline map placeholder - redirect to main map"""
    return redirect('PIU_Mapping_project_Sites:index')


def togglemarker(request):
    """Toggle marker placeholder - redirect to main map"""
    return redirect('PIU_Mapping_project_Sites:index')


def mappingCreateView(request):
    """Mapping create placeholder - redirect to add mapping"""
    return redirect('PIU_Mapping_project_Sites:add_mapping')


def settlementswithcor(request):
    """Settlements with coordinates placeholder - redirect to main map"""
    return redirect('PIU_Mapping_project_Sites:index')