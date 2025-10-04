from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse
from datetime import datetime
from .models import ProjectProgress
from .forms import ProjectProgressForm, ProjectProgressFilterForm

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from io import BytesIO


@login_required
def project_progress_list(request):
    """List all project progress records with filtering"""
    progress_records = ProjectProgress.objects.all().select_related('project').order_by('-id')  # type: ignore
    
    # Apply filters
    filter_form = ProjectProgressFilterForm(request.GET)
    if filter_form.is_valid():
        if filter_form.cleaned_data.get('project'):
            progress_records = progress_records.filter(project=filter_form.cleaned_data['project'])
    
    context = {
        'progress_records': progress_records,
        'filter_form': filter_form,
    }
    return render(request, 'project_progress/list.html', context)


@login_required
def project_progress_create(request):
    """Create new project progress record"""
    if request.method == 'POST':
        form = ProjectProgressForm(request.POST)
        if form.is_valid():
            progress = form.save(commit=False)
            progress.loginuser = request.user
            progress.save()
            messages.success(request, 'Project progress record created successfully!')
            return redirect('project_progress:list')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = ProjectProgressForm()
    
    context = {
        'form': form,
        'is_edit': False,
    }
    return render(request, 'project_progress/form.html', context)


@login_required
def project_progress_update(request, pk):
    """Update existing project progress record"""
    progress = get_object_or_404(ProjectProgress, pk=pk)
    
    if request.method == 'POST':
        form = ProjectProgressForm(request.POST, instance=progress)
        if form.is_valid():
            form.save()
            messages.success(request, 'Project progress record updated successfully!')
            return redirect('project_progress:list')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = ProjectProgressForm(instance=progress)
    
    context = {
        'form': form,
        'is_edit': True,
        'progress': progress,
    }
    return render(request, 'project_progress/form.html', context)


@login_required
def project_progress_detail(request, pk):
    """View detailed information for a project progress record"""
    progress = get_object_or_404(ProjectProgress, pk=pk)
    
    context = {
        'progress': progress,
    }
    return render(request, 'project_progress/detail.html', context)


@login_required
def project_progress_delete(request, pk):
    """Delete project progress record"""
    progress = get_object_or_404(ProjectProgress, pk=pk)
    
    if request.method == 'POST':
        progress.delete()
        messages.success(request, 'Project progress record deleted successfully!')
        return redirect('project_progress:list')
    
    context = {
        'progress': progress,
    }
    return render(request, 'project_progress/delete_confirm.html', context)


# ============ Export Functions ============

@login_required
def export_progress_excel(request):
    """Export project progress to Excel"""
    # Get filtered queryset
    progress_records = ProjectProgress.objects.all().select_related('project').order_by('-id')  # type: ignore
    
    # Apply same filters as list view
    filter_form = ProjectProgressFilterForm(request.GET)
    if filter_form.is_valid():
        if filter_form.cleaned_data.get('project'):
            progress_records = progress_records.filter(project=filter_form.cleaned_data['project'])
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active  # type: ignore
    ws.title = "Project Progress"  # type: ignore
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="198754", end_color="198754", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    data_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    number_alignment = Alignment(horizontal="right", vertical="center")
    date_alignment = Alignment(horizontal="center", vertical="center")
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = [
        'Project', 'Total Funding', 'Disbursement', 'Disbursement Rate',
        'Physical Progress', 'Time Elapsed', 'Time Overrun',
        'Start Date', 'End Date', 'Created By', 'Date Created'
    ]
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)  # type: ignore
        cell.value = header  # type: ignore
        cell.font = header_font  # type: ignore
        cell.fill = header_fill  # type: ignore
        cell.alignment = header_alignment  # type: ignore
        cell.border = border  # type: ignore
    
    # Write data
    for row_num, progress in enumerate(progress_records, 2):
        data = [
            str(progress.project),
            float(progress.total_funding),
            float(progress.disbursement),
            progress.over_all_disbursement_rate,
            progress.over_all_physical_progress,
            progress.over_project_time_elapsed,
            progress.over_project_time_over_run,
            progress.start_date,
            progress.end_date,
            str(progress.loginuser.username) if progress.loginuser else '',
            progress.date_created,
        ]
        
        for col_num, value in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col_num)  # type: ignore
            cell.value = value  # type: ignore
            cell.border = border  # type: ignore
            
            # Apply alignment based on data type
            if col_num in [2, 3]:  # Funding and Disbursement
                cell.alignment = number_alignment  # type: ignore
                cell.number_format = '#,##0.00'  # type: ignore
            elif col_num in [8, 9, 11]:  # Dates
                cell.alignment = date_alignment  # type: ignore
                if isinstance(value, datetime):
                    cell.number_format = 'YYYY-MM-DD'  # type: ignore
            else:
                cell.alignment = data_alignment  # type: ignore
    
    # Auto-adjust column widths
    for column in ws.columns:  # type: ignore
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width  # type: ignore
    
    # Freeze header row
    ws.freeze_panes = 'A2'  # type: ignore
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    response['Content-Disposition'] = f'attachment; filename=Project_Progress_Report_{timestamp}.xlsx'
    
    wb.save(response)  # type: ignore
    return response


@login_required
def export_progress_pdf(request):
    """Export project progress to PDF"""
    # Get filtered queryset
    progress_records = ProjectProgress.objects.all().select_related('project').order_by('-id')  # type: ignore
    
    # Apply same filters as list view
    filter_form = ProjectProgressFilterForm(request.GET)
    if filter_form.is_valid():
        if filter_form.cleaned_data.get('project'):
            progress_records = progress_records.filter(project=filter_form.cleaned_data['project'])
    
    # Create PDF buffer
    buffer = BytesIO()
    
    # Create PDF document with A4 portrait
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=0.75*inch,
        leftMargin=0.75*inch,
        topMargin=1*inch,
        bottomMargin=0.75*inch
    )
    
    # Container for PDF elements
    elements = []
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#198754'),
        spaceAfter=30,
        alignment=1  # Center alignment
    )
    
    # Cell text style with word wrapping
    cell_style = ParagraphStyle(
        'CellStyle',
        parent=styles['Normal'],
        fontSize=7,
        leading=9,
        wordWrap='CJK',
        alignment=0
    )
    
    # Title
    title = Paragraph("Project Progress Tracking Report", title_style)
    elements.append(title)
    
    # Report info
    report_info = Paragraph(
        f"Generated on: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}<br/>"
        f"Total Records: {progress_records.count()}",
        styles['Normal']
    )
    elements.append(report_info)
    elements.append(Spacer(1, 0.3*inch))
    
    # Table data - Headers
    table_data = [[
        'Project', 'Total Funding', 'Disbursement', 'Disb. Rate',
        'Progress', 'Time Elapsed', 'Time Overrun', 'Start Date', 'End Date'
    ]]
    
    # Add data rows with Paragraph objects for text wrapping
    for progress in progress_records:
        table_data.append([  # type: ignore
            Paragraph(str(progress.project), cell_style),
            Paragraph(f"${float(progress.total_funding):,.2f}", cell_style),
            Paragraph(f"${float(progress.disbursement):,.2f}", cell_style),
            Paragraph(progress.over_all_disbursement_rate, cell_style),
            Paragraph(progress.over_all_physical_progress, cell_style),
            Paragraph(progress.over_project_time_elapsed, cell_style),
            Paragraph(progress.over_project_time_over_run, cell_style),
            Paragraph(progress.start_date.strftime('%Y-%m-%d') if progress.start_date else '', cell_style),
            Paragraph(progress.end_date.strftime('%Y-%m-%d') if progress.end_date else '', cell_style),
        ])
    
    # Create table
    table = Table(table_data, colWidths=[
        1.2*inch, 0.9*inch, 0.9*inch, 0.6*inch,
        0.6*inch, 0.9*inch, 0.9*inch, 0.8*inch, 0.8*inch
    ])
    
    # Table style
    table.setStyle(TableStyle([
        # Header row
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#198754')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('TOPPADDING', (0, 0), (-1, 0), 12),
        ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
        
        # Data rows
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('TOPPADDING', (0, 1), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ('LEFTPADDING', (0, 1), (-1, -1), 4),
        ('RIGHTPADDING', (0, 1), (-1, -1), 4),
        ('VALIGN', (0, 1), (-1, -1), 'TOP'),
        
        # Grid
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        
        # Alternating row colors
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
    ]))
    
    elements.append(table)
    
    # Build PDF
    doc.build(elements)
    
    # Get PDF from buffer
    pdf = buffer.getvalue()
    buffer.close()
    
    # Create response
    response = HttpResponse(content_type='application/pdf')
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    response['Content-Disposition'] = f'attachment; filename=Project_Progress_Report_{timestamp}.pdf'
    response.write(pdf)
    
    return response
