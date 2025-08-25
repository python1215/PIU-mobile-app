from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import csrf_exempt
from django.core.exceptions import ValidationError
from decimal import Decimal
import json
from .models import *
from .forms import *
# Using Django ORM exclusively for all database operations
# Import only available forms that exist

# Create your views here.
@login_required
def project_list(request):
    """Enhanced project list with filtering and statistics"""
    from django.db.models import Q, Sum, Count, Avg
    from setup.models import Donor
    
    # Get all projects - show all projects for comprehensive monitoring
    projects_qs = Project.objects.all().select_related('currency').prefetch_related('donors', 'contributors')
    
    # Filter parameters
    project_id = request.GET.get('projectID', '')
    project_name = request.GET.get('project', '')
    donor_filter = request.GET.get('donor', '')
    currency_filter = request.GET.get('currency', '')
    status_filter = request.GET.get('status', '')
    search_filter = request.GET.get('search', '')
    
    # Apply filters
    if project_id:
        projects_qs = projects_qs.filter(projectID__icontains=project_id)
    
    if project_name:
        projects_qs = projects_qs.filter(project__icontains=project_name)
    
    if donor_filter:
        projects_qs = projects_qs.filter(donors__id=donor_filter)
    
    if currency_filter:
        projects_qs = projects_qs.filter(currency__id=currency_filter)
    
    if search_filter:
        projects_qs = projects_qs.filter(
            Q(projectID__icontains=search_filter) |
            Q(project__icontains=search_filter) |
            Q(implementation_Status__icontains=search_filter)
        )
    
    # Calculate statistics
    total_projects = projects_qs.count()
    total_funding = projects_qs.aggregate(Sum('funding'))['funding__sum'] or 0
    unique_donors = projects_qs.values('donors').distinct().count()
    
    stats = {
        'total_projects': total_projects,
        'total_funding': total_funding,
        'unique_donors': unique_donors,
    }
    
    # Get filter options
    donors = Donor.objects.all()
    currencies = Currency.objects.all()
    
    is_filtered = bool(project_id or project_name or donor_filter or currency_filter or status_filter or search_filter)
    
    context = {
        'projects': projects_qs,
        'stats': stats,
        'is_filtered': is_filtered,
        'donors': donors,
        'currencies': currencies,
    }
    
    return render(request, 'PIU_Financial_mgt/projects/enhanced_project_list.html', context)

@login_required
def project_detail(request, project_id):
    """Detailed view of a single project with related data"""
    from django.db.models import Sum, Count
    
    project = get_object_or_404(Project, projectID=project_id)
    
    # Get related data - convert to lists for template evaluation
    components = list(Component.objects.filter(projectID=project).order_by('-date'))
    subcomponents = list(Subcomponent.objects.filter(projectID=project).order_by('-date'))
    recent_activities = list(Activities.objects.filter(projectID=project).order_by('-date')[:10])
    
    # Calculate statistics  
    components_count = len(components)
    subcomponents_count = len(subcomponents)
    activities_count = Activities.objects.filter(projectID=project).count()
    total_allocation = sum(c.allocation for c in components if c.allocation) or 0
    
    context = {
        'project': project,
        'components': components,
        'subcomponents': subcomponents,
        'recent_activities': recent_activities,
        'components_count': components_count,
        'subcomponents_count': subcomponents_count,
        'activities_count': activities_count,
        'total_allocation': total_allocation,
    }
    
    return render(request, 'PIU_Financial_mgt/projects/project_detail.html', context)

# Create your views here.
@login_required
def add_project(request):
    if request.method == 'POST':
        
        
        
        
        form = addProjectForm(request.POST)
        
        
        is_valid = form.is_valid()
        
        print("Form errors:", dict(form.errors))
        
        if form.is_valid():
            print("Form valid - saving project")
            project = form.save(commit=False)
            project.loginUser = request.user
            project.save()
            form.save_m2m()  # Save many-to-many relationships
            messages.success(request, 'Project created successfully!')
            print("Project saved successfully - redirecting")
            return redirect('PIU_Financial_mgt:projects')
        else:
            print("Form validation failed:")
            for field, errors in form.errors.items():
                print(f"Field {field}: {errors}")
    else:
        
        print("Method: GET")
        print("GET request - creating new form")
        form = addProjectForm()
    return render(request, 'PIU_Financial_mgt/projects/add-project.html', {'form': form})

@login_required
def add_project_test(request):
    if request.method == 'POST':
        form = addProjectForm(request.POST)
        if form.is_valid():
            project = form.save(commit=False)
            project.loginUser = request.user
            project.save()
            form.save_m2m()  # Save many-to-many relationships
            messages.success(request, 'Project created successfully!')
            return redirect('PIU_Financial_mgt:enhanced_project_dashboard')
    else:
        form = addProjectForm()
    return render(request, 'PIU_Financial_mgt/projects/add-project.html', {'form': form})

@login_required
def enhanced_project_dashboard(request, project_id=None):
    from django.db.models import Sum, Count, Avg
    from social_and_env.models import ESIA, GrievianceMonitoringLog, OHS_Monitoring, PAP, CommunityConsult_Engagement
    
    # Get specific project if project_id is provided
    selected_project = None
    if project_id:
        selected_project = get_object_or_404(Project, projectID=project_id)
        
    # Calculate dashboard statistics (project-specific or overall)
    if selected_project:
        # Project-specific statistics
        total_projects = 1
        total_components = Component.objects.filter(projectID=selected_project).count()
        total_subcomponents = Subcomponent.objects.filter(projectID=selected_project).count()
        total_activities = Activities.objects.filter(projectID=selected_project).count()
        total_funding = selected_project.funding or 0
        active_projects = 1 if not selected_project.closure_Date else 0
        total_disbursed = Component.objects.filter(projectID=selected_project).aggregate(Sum('allocation'))['allocation__sum'] or 0
        
        # Social and Environmental data for selected project
        esia_records = ESIA.objects.filter(project_name=selected_project)
        grievance_records = GrievianceMonitoringLog.objects.filter(projectID=selected_project)
        ohs_records = OHS_Monitoring.objects.filter(projectID=selected_project)
        pap_records = PAP.objects.filter(projectID=selected_project)
        community_records = CommunityConsult_Engagement.objects.filter(project_name=selected_project)
    else:
        # Overall statistics
        total_projects = Project.objects.count()
        total_components = Component.objects.count()
        total_subcomponents = Subcomponent.objects.count()
        total_activities = Activities.objects.count()
        total_funding = Project.objects.aggregate(Sum('funding'))['funding__sum'] or 0
        active_projects = Project.objects.filter(closure_Date__isnull=True).count()
        total_disbursed = Component.objects.aggregate(Sum('allocation'))['allocation__sum'] or 0
        
        # Social and Environmental data - Use Django ORM for SQLite compatibility
        esia_records = ESIA.objects.all()
        grievance_records = GrievianceMonitoringLog.objects.all()
        ohs_records = OHS_Monitoring.objects.all()
        pap_records = PAP.objects.all()
        community_records = CommunityConsult_Engagement.objects.all()
        print(f"Successfully loaded Social & Environmental data using Django ORM")
    
    # Recent data - Use Django ORM for SQLite compatibility
    try:
        recent_activities = list(Activities.objects.select_related('projectID', 'compID', 'subcompID').order_by('-date')[:5])
        print(f"Successfully loaded {len(recent_activities)} recent activities using Django ORM")
    except Exception as e:
        print(f"Error loading recent activities: {e}")
        recent_activities = []
    
    # Other recent data (project-specific or overall) - convert to lists for template evaluation
    if selected_project:
        recent_projects = [selected_project]
        recent_components = list(Component.objects.filter(projectID=selected_project).order_by('-date')[:5])
        recent_subcomponents = list(Subcomponent.objects.filter(projectID=selected_project).order_by('-date')[:5])
    else:
        recent_projects = list(Project.objects.order_by('-date')[:5])
        recent_components = list(Component.objects.order_by('-date')[:5])
        recent_subcomponents = list(Subcomponent.objects.order_by('-date')[:5])
    
    # Budget utilization percentage
    budget_utilization = (total_disbursed / total_funding * 100) if total_funding > 0 else 0
    
    context = {
        'selected_project': selected_project,
        'total_projects': total_projects,
        'total_components': total_components,
        'total_subcomponents': total_subcomponents,
        'total_activities': total_activities,
        'total_funding': total_funding,
        'total_disbursed': total_disbursed,
        'active_projects': active_projects,
        'projects_in_progress': active_projects,
        'budget_utilization': round(budget_utilization, 1),
        'recent_projects': recent_projects,
        'recent_components': recent_components,
        'recent_subcomponents': recent_subcomponents,
        'recent_activities': recent_activities,
        
        # Social and Environmental data
        'esia_records': esia_records,
        'grievance_records': grievance_records,
        'ohs_records': ohs_records,
        'pap_records': pap_records,
        'community_records': community_records,
        'esia_count': esia_records.count(),
        'grievance_count': grievance_records.count(),
        'ohs_count': ohs_records.count(),
        'pap_count': pap_records.count(),
        'community_count': community_records.count(),
    }
    
    return render(request, 'PIU_Financial_mgt/projects/enhanced_project_dashboard.html', context)

@login_required
def add_component(request):
    if request.method == 'POST':
        print("=== ADD COMPONENT VIEW ===")
        
        
        
        form = addComponentForm(request.POST)
        
        
        is_valid = form.is_valid()
        
        print("Form errors:", dict(form.errors))
        
        if form.is_valid():
            try:
                print("Form valid - saving component")
                component = form.save(commit=False)
                component.loginUser = request.user
                component.save()
                messages.success(request, 'Component created successfully!')
                print("Component saved successfully - redirecting")
                return redirect('PIU_Financial_mgt:enhanced_project_dashboard')
            except Exception as e:
                print(f"Error saving component: {e}")
                messages.error(request, f'Error saving component: {str(e)}')
                # Continue to re-render form with error
        else:
            print("Form validation failed:")
            for field, errors in form.errors.items():
                print(f"Field {field}: {errors}")
            # Add validation error message for user
            messages.error(request, 'Please correct the errors below and try again.')
    else:
        print("=== ADD COMPONENT VIEW ===")
        print("Method: GET")
        print("GET request - creating new form")
        form = addComponentForm()
    return render(request, 'PIU_Financial_mgt/components/add-component.html', {'form': form})

@login_required
def add_subcomponent(request):
    print("=== ADD SUBCOMPONENT VIEW ===")
    print(f"Method: {request.method}")
    print(f"POST data: {request.POST}")
    print(f"Files: {request.FILES}")
    
    if request.method == 'POST':
        # Check if this is a valid AJAX submission
        project_id = request.POST.get('projectID', '').strip()
        comp_id = request.POST.get('compID', '').strip()
        subcomponent_name = request.POST.get('subcomponent', '').strip()
        description = request.POST.get('subcomponent_Description', '').strip()
        allocation = request.POST.get('allocation', '').strip()
        currency = request.POST.get('currency', '').strip()
        
        print(f"Has actual form data: {subcomponent_name}")
        
        # Only process if we have actual data (not empty form submission)
        if not subcomponent_name or not description or not allocation or not currency:
            print("No form data detected, redirecting to GET")
            return redirect('PIU_Financial_mgt:add_subcomponent')
        
        form = addsubcomponentForm(request.POST)
        
        if form.is_valid():
            print("Form is valid - attempting to save")
            try:
                subcomponent = form.save(commit=False)
                subcomponent.loginUser = request.user
                subcomponent.save()
                print(f"Subcomponent saved successfully: {subcomponent.subcomponent}")
                print("Subcomponent saved successfully!")
                
                # Return JSON response for AJAX
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({'success': True, 'message': 'Subcomponent created successfully!'})
                
                messages.success(request, 'Subcomponent created successfully!')
                return redirect('PIU_Financial_mgt:subcomponents')
            except Exception as e:
                print(f"Error saving subcomponent: {e}")
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({'success': False, 'message': f'Error saving: {str(e)}'})
        else:
            print(f"Form validation errors: {form.errors}")
            # Return JSON error for AJAX
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'errors': dict(form.errors)})
    else:
        print("GET request - creating new form")
        form = addsubcomponentForm()
    return render(request, 'PIU_Financial_mgt/subcomponents/add_subcomponent.html', {'form': form})

@login_required
def add_subcomponent_isolated(request):
    """
    Isolated add subcomponent view using pure dropdowns and contenteditable fields
    """
    if request.method == 'POST':
        try:
            # Extract data from POST request
            project_id = request.POST.get('projectID', '').strip()
            component_id = request.POST.get('compID', '').strip()
            subcomponent_name = request.POST.get('subcomponent', '').strip()
            description = request.POST.get('subcomponent_Description', '').strip()
            currency_id = request.POST.get('currency', '').strip()
            allocation_str = request.POST.get('allocation', '').strip()
            
            # Validate required fields
            if not all([project_id, component_id, subcomponent_name, description, currency_id, allocation_str]):
                return JsonResponse({'success': False, 'message': 'All fields are required'})
            
            # Validate allocation is a number
            try:
                allocation = Decimal(allocation_str)
            except:
                return JsonResponse({'success': False, 'message': 'Invalid allocation amount'})
            
            # Get related objects
            try:
                project = Project.objects.get(projectID=project_id)
                component = Component.objects.get(compID=component_id)
                currency = Currency.objects.get(currency=currency_id)
            except:
                return JsonResponse({'success': False, 'message': 'Invalid project, component, or currency'})
            
            # Create subcomponent
            subcomponent = Subcomponent.objects.create(
                projectID=project,
                compID=component,
                subcomponent=subcomponent_name,
                subcomponent_Description=description,
                currency=currency,
                allocation=allocation,
                loginUser=request.user
            )
            
            return JsonResponse({'success': True, 'message': 'Subcomponent created successfully'})
            
        except Exception as e:
            return JsonResponse({'success': False, 'message': f'Error: {str(e)}'})
    
    # GET request - show form
    form = AddSubcomponentForm()
    context = {
        'form': form,
        'page_title': 'Add Subcomponent - Isolated',
    }
    return render(request, 'PIU_Financial_mgt/subcomponents/add_subcomponent.html', context)

@login_required
def load_project_components(request):
    print(f"=== LOAD PROJECT COMPONENTS VIEW ===")
    print(f"Method: {request.method}")
    print(f"GET params: {request.GET}")
    print(f"POST params: {request.POST}")
    print(f"Is HTMX request: {request.headers.get('HX-Request', 'No')}")
    
    project_id = request.GET.get("projectID")  # Changed from project_id to projectID
    print("Received projectID:", project_id)
    
    if project_id:
        # Filter components by the selected project using the correct field reference
        components = Component.objects.filter(projectID__projectID__icontains=project_id)
        print(f"Returning {components.count()} components for project {project_id}")
    else:
        # If no project selected, return empty queryset
        components = Component.objects.none()
        print("No project selected, returning empty components")
    
    # Show component details for debugging
    for comp in components:
        print(f"Component: {comp.project_components}, Project: {comp.projectID}")
    
    return render(request, "htmx/project_components_dropdown.html", {"components": components})

# Removed duplicate function - using the enhanced version below
    

def load_project_Activities(request):
    subcompID = request.GET.get("subcompID")
    print("Received subcompID:", subcompID)
    activities = Activities.objects.filter(subcompID=subcompID)
   
    print("activities:", activities)
    return render(request, "htmx/project_activities_dropdown.html", {"activities": activities})


@login_required
def addproject(request):
    if request.method == 'POST':
        form = addProjectForm(request.POST)
        if form.is_valid():
            project = form.save(commit=False)
            project.loginUser = request.user  # Set the logged-in user
            project.save()
            form.save_m2m()  # Save many-to-many relationships
            messages.success(request, 'Project created successfully!')
            return redirect('PIU_Financial_mgt:projects')  
    else:
        form = addProjectForm()
    return render(request, 'PIU_Financial_mgt/projects/add-project.html', {'form': form})


@login_required
def projects(request):
    # Enhanced project list with filtering and statistics
    from django.db.models import Q, Sum, Count, Avg
    from setup.models import Donor
    
    # Get all projects - show all projects for comprehensive monitoring
    projects_qs = Project.objects.all().select_related('currency').prefetch_related('donors', 'contributors')
    
    # Filter parameters
    project_id = request.GET.get('projectID', '')
    project_name = request.GET.get('project', '')
    donor_id = request.GET.get('donor', '')  # Fixed parameter name to match template
    currency_id = request.GET.get('currency', '')
    funding_min = request.GET.get('funding_min', '')
    funding_max = request.GET.get('funding_max', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    is_filtered = False
    
    # Apply filters
    if project_id:
        projects_qs = projects_qs.filter(projectID__icontains=project_id)
        is_filtered = True
    
    if project_name:
        projects_qs = projects_qs.filter(project__icontains=project_name)
        is_filtered = True
    
    if donor_id:
        projects_qs = projects_qs.filter(donors__donorID=donor_id)
        is_filtered = True
    
    if currency_id:
        projects_qs = projects_qs.filter(currency__id=currency_id)
        is_filtered = True
    
    if funding_min:
        try:
            projects_qs = projects_qs.filter(funding__gte=float(funding_min))
            is_filtered = True
        except ValueError:
            pass
    
    if funding_max:
        try:
            projects_qs = projects_qs.filter(funding__lte=float(funding_max))
            is_filtered = True
        except ValueError:
            pass
    
    if date_from:
        projects_qs = projects_qs.filter(date__gte=date_from)
        is_filtered = True
    
    if date_to:
        projects_qs = projects_qs.filter(date__lte=date_to)
        is_filtered = True
    
    # Calculate statistics
    total_projects = Project.objects.count()
    filtered_count = projects_qs.count()
    funding_stats = projects_qs.aggregate(
        total_funding=Sum('funding'),
        avg_funding=Avg('funding')
    )
    total_funding = funding_stats['total_funding'] or 0
    avg_funding = funding_stats['avg_funding'] or 0
    active_projects = projects_qs.filter(closure_Date__isnull=True).count()
    
    stats = {
        'total_projects': filtered_count,
        'total_funding': total_funding,
        'avg_funding': avg_funding,
        'active_projects': active_projects,
    }
    
    # Get filter options
    donors = Donor.objects.all()
    currencies = Currency.objects.all()
    
    context = {
        'projects': projects_qs.order_by('-date'),
        'stats': stats,
        'is_filtered': is_filtered,
        'donors': donors,
        'currencies': currencies,
    }
    
    return render(request, 'PIU_Financial_mgt/projects/enhanced_project_list.html', context)

@login_required
def addcomponent(request):
    if request.method == 'POST':
        form = addComponentForm(request.POST)
        if form.is_valid():
            component = form.save(commit=False)
            component.loginUser = request.user  # Set the logged-in user
            component.save()
            messages.success(request, 'Component created successfully!')
            return redirect('PIU_Financial_mgt:components')  
    else:
        form = addComponentForm()
    return render(request, 'PIU_Financial_mgt/components/add-component.html', {'form': form})

@login_required
def components(request):
    # Enhanced component list with filtering and statistics
    from django.db.models import Q, Sum, Count
    from PIU_Financial_mgt.models import Currency
    
    print(f"=== COMPONENTS VIEW ACCESSED ===")
    print(f"User: {request.user}")
    print(f"Request method: {request.method}")
    print(f"Request path: {request.path}")
    print(f"User authenticated: {request.user.is_authenticated}")
    print(f"===================================")
    
    # Get all components with detailed error handling
    try:
        components_qs = Component.objects.all().select_related('projectID', 'currency', 'loginUser')
        print(f"Successfully created component QuerySet")
    except Exception as e:
        print(f"Error creating component QuerySet: {e}")
        components_qs = Component.objects.all()
    
    # Filter parameters with enhanced debugging
    print(f"=== REQUEST PARAMETERS DEBUG ===")
    print(f"All GET parameters: {dict(request.GET)}")
    
    project_id = request.GET.get('project', '').strip()
    component_name = request.GET.get('component', '').strip()
    currency_id = request.GET.get('currency', '').strip()
    allocation_min = request.GET.get('allocation_min', '').strip()
    allocation_max = request.GET.get('allocation_max', '').strip()
    date_from = request.GET.get('date_from', '').strip()
    date_to = request.GET.get('date_to', '').strip()
    
    print(f"Processed parameters:")
    print(f"  project_id: '{project_id}'")
    print(f"  component_name: '{component_name}'")
    print(f"  currency_id: '{currency_id}'")
    print(f"  allocation_min: '{allocation_min}'")
    print(f"  allocation_max: '{allocation_max}'")
    print(f"  date_from: '{date_from}'")
    print(f"  date_to: '{date_to}'")
    print(f"===================================")
    
    is_filtered = False
    
    # Apply filters only if values are provided - Enhanced debugging for offline deployments
    print(f"=== FILTERING DEBUG ===")
    print(f"project_id filter: '{project_id}' (length: {len(project_id) if project_id else 0})")
    print(f"component_name filter: '{component_name}' (length: {len(component_name) if component_name else 0})")
    print(f"currency_id filter: '{currency_id}' (length: {len(currency_id) if currency_id else 0})")
    print(f"Initial components_qs count: {components_qs.count()}")
    
    if project_id and project_id.strip() != '':
        components_qs = components_qs.filter(projectID__projectID__icontains=project_id)
        is_filtered = True
        print(f"After project filter: {components_qs.count()} components")
    
    if component_name and component_name.strip() != '':
        components_qs = components_qs.filter(project_components__icontains=component_name)
        is_filtered = True
        print(f"After component name filter: {components_qs.count()} components")
    
    if currency_id and currency_id.strip() != '':
        components_qs = components_qs.filter(currency__id=currency_id)
        is_filtered = True
        print(f"After currency filter: {components_qs.count()} components")
    
    if allocation_min:
        try:
            components_qs = components_qs.filter(allocation__gte=float(allocation_min))
            is_filtered = True
        except ValueError:
            pass
    
    if allocation_max:
        try:
            components_qs = components_qs.filter(allocation__lte=float(allocation_max))
            is_filtered = True
        except ValueError:
            pass
    
    if date_from:
        components_qs = components_qs.filter(date__gte=date_from)
        is_filtered = True
    
    if date_to:
        components_qs = components_qs.filter(date__lte=date_to)
        is_filtered = True
    
    # Calculate statistics
    total_components = Component.objects.count()
    filtered_count = components_qs.count()
    total_allocation = components_qs.aggregate(Sum('allocation'))['allocation__sum'] or 0
    unique_projects = components_qs.values('projectID').distinct().count()
    
    stats = {
        'total_components': total_components,
        'filtered_count': filtered_count,
        'total_allocation': total_allocation,
        'unique_projects': unique_projects,
    }
    
    # Get filter options
    projects = Project.objects.all()
    currencies = Currency.objects.all()
    
    # Order components for display
    final_components = components_qs.order_by('-date')
    
    # Debug logging for troubleshooting (can be removed in production)
    print(f"=== COMPONENT LIST DEBUG ===")
    print(f"User: {request.user}")
    print(f"Total components in DB: {total_components}")
    print(f"Filtered components count: {filtered_count}")
    print(f"Is filtered: {is_filtered}")
    try:
        sample_data = list(final_components[:3].values('compID', 'project_components', 'projectID__projectID'))
        print(f"Sample components: {sample_data}")
    except Exception as e:
        print(f"Error getting sample data: {e}")
    print(f"============================")
    
    # Ensure we have a valid QuerySet for the template - Convert to list for offline deployment compatibility
    try:
        # Test if the QuerySet can be evaluated
        component_count = final_components.count()
        print(f"Successfully evaluated QuerySet: {component_count} components")
        
        # For offline deployments, ensure we get all components without filtering issues
        if component_count == 0 or (component_count < total_components and not is_filtered):
            print(f"WARNING: Filtered count ({component_count}) is less than total ({total_components}) without explicit filtering")
            print(f"Resetting to unfiltered query for offline deployment compatibility")
            final_components = Component.objects.all().select_related('projectID', 'currency', 'loginUser').order_by('-date')
            component_count = final_components.count()
            print(f"Unfiltered count: {component_count} components")
            # Update stats to reflect the reset
            stats['filtered_count'] = component_count
            is_filtered = False
        elif component_count < total_components and is_filtered:
            print(f"INFO: Filtering applied - showing {component_count} of {total_components} components")
        
        # Convert to list to ensure template compatibility in offline deployments
        components_list = list(final_components)
        print(f"Converted to list: {len(components_list)} components")
        
        # Additional debug for offline deployments
        if len(components_list) != component_count:
            print(f"MISMATCH: QuerySet count ({component_count}) != List length ({len(components_list)})")
        
    except Exception as e:
        print(f"QuerySet evaluation error: {e}")
        # Fallback to basic query if complex query fails
        final_components = Component.objects.all().select_related('projectID', 'currency', 'loginUser').order_by('-date')
        components_list = list(final_components)
        print(f"Fallback query result: {len(components_list)} components")
    
    context = {
        'components': components_list,  # Use the converted list for better template compatibility
        'stats': stats,
        'is_filtered': is_filtered,
        'projects': projects,
        'currencies': currencies,
    }
    
    return render(request, 'PIU_Financial_mgt/components/enhanced_component_list.html', context)

@login_required
def export_components_excel(request):
    """Export components to Excel with applied filters"""
    from django.http import HttpResponse
    from openpyxl import Workbook
    from PIU_Financial_mgt.models import Currency
    import datetime
    
    print(f"=== COMPONENTS EXCEL EXPORT DEBUG ===")
    print(f"Excel export accessed by user: {request.user}")
    print(f"GET parameters: {request.GET}")
    print(f"======================================")
    
    # Apply same filtering logic as components view
    components_qs = Component.objects.all().select_related('projectID', 'currency', 'loginUser')
    
    # Filter parameters (same as components view)
    project_id = request.GET.get('project', '').strip()
    component_name = request.GET.get('component', '').strip()
    currency_id = request.GET.get('currency', '').strip()
    allocation_min = request.GET.get('allocation_min', '').strip()
    allocation_max = request.GET.get('allocation_max', '').strip()
    date_from = request.GET.get('date_from', '').strip()
    date_to = request.GET.get('date_to', '').strip()
    
    print(f"Filters applied - project: '{project_id}', component: '{component_name}', currency: '{currency_id}'")
    
    # Apply filters
    if project_id and project_id != '':
        components_qs = components_qs.filter(projectID__projectID__icontains=project_id)
        print(f"Filtered by project ID: {project_id}")
    if component_name and component_name != '':
        components_qs = components_qs.filter(project_components__icontains=component_name)
        print(f"Filtered by component name: {component_name}")
    if currency_id and currency_id != '':
        components_qs = components_qs.filter(currency__id=currency_id)
        print(f"Filtered by currency ID: {currency_id}")
    if allocation_min:
        try:
            components_qs = components_qs.filter(allocation__gte=float(allocation_min))
            print(f"Filtered by min allocation: {allocation_min}")
        except ValueError:
            pass
    if allocation_max:
        try:
            components_qs = components_qs.filter(allocation__lte=float(allocation_max))
            print(f"Filtered by max allocation: {allocation_max}")
        except ValueError:
            pass
    if date_from:
        components_qs = components_qs.filter(date__gte=date_from)
        print(f"Filtered by date from: {date_from}")
    if date_to:
        components_qs = components_qs.filter(date__lte=date_to)
        print(f"Filtered by date to: {date_to}")
    
    print(f"Final component count for export: {components_qs.count()}")
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Components"
    
    # Headers
    headers = ['Component ID', 'Project Name', 'Component Name', 'Description', 
               'Currency', 'Allocation', 'Date Created', 'Created By']
    ws.append(headers)
    
    # Set column widths for better readability
    from openpyxl.utils import get_column_letter
    column_widths = {
        'A': 12,  # Component ID
        'B': 35,  # Project Name
        'C': 30,  # Component Name
        'D': 40,  # Description
        'E': 12,  # Currency
        'F': 15,  # Allocation
        'G': 18,  # Date
        'H': 15,  # Created By
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Add data rows
    for component in components_qs.order_by('-date'):
        row_data = [
            component.compID,
            component.projectID.project if component.projectID else '',
            component.project_components or '',
            component.component_description or '',
            component.currency.currency if component.currency else '',
            component.allocation or 0,
            component.date.strftime('%Y-%m-%d %H:%M') if component.date else '',
            component.loginUser.username if component.loginUser else '',
        ]
        ws.append(row_data)
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
    response['Content-Disposition'] = f'attachment; filename="components_export_{timestamp}.xlsx"'
    
    wb.save(response)
    return response

@login_required
def export_components_pdf(request):
    """Export components to PDF with A4 portrait layout"""
    from django.http import HttpResponse
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.units import mm
    from PIU_Financial_mgt.models import Currency
    import datetime
    
    print(f"=== COMPONENTS PDF EXPORT DEBUG ===")
    print(f"PDF export accessed by user: {request.user}")
    print(f"GET parameters: {request.GET}")
    print(f"====================================")
    
    # Apply same filtering logic as components view
    components_qs = Component.objects.all().select_related('projectID', 'currency', 'loginUser')
    
    # Filter parameters (same as components view)
    project_id = request.GET.get('project', '').strip()
    component_name = request.GET.get('component', '').strip()
    currency_id = request.GET.get('currency', '').strip()
    allocation_min = request.GET.get('allocation_min', '').strip()
    allocation_max = request.GET.get('allocation_max', '').strip()
    date_from = request.GET.get('date_from', '').strip()
    date_to = request.GET.get('date_to', '').strip()
    
    print(f"PDF filters applied - project: '{project_id}', component: '{component_name}', currency: '{currency_id}'")
    
    # Apply filters
    if project_id and project_id != '':
        components_qs = components_qs.filter(projectID__projectID__icontains=project_id)
        print(f"Filtered by project ID: {project_id}")
    if component_name and component_name != '':
        components_qs = components_qs.filter(project_components__icontains=component_name)
        print(f"Filtered by component name: {component_name}")
    if currency_id and currency_id != '':
        components_qs = components_qs.filter(currency__id=currency_id)
        print(f"Filtered by currency ID: {currency_id}")
    if allocation_min:
        try:
            components_qs = components_qs.filter(allocation__gte=float(allocation_min))
            print(f"Filtered by min allocation: {allocation_min}")
        except ValueError:
            pass
    if allocation_max:
        try:
            components_qs = components_qs.filter(allocation__lte=float(allocation_max))
            print(f"Filtered by max allocation: {allocation_max}")
        except ValueError:
            pass
    if date_from:
        components_qs = components_qs.filter(date__gte=date_from)
        print(f"Filtered by date from: {date_from}")
    if date_to:
        components_qs = components_qs.filter(date__lte=date_to)
        print(f"Filtered by date to: {date_to}")
    
    print(f"Final component count for PDF export: {components_qs.count()}")
    
    # Create PDF
    response = HttpResponse(content_type='application/pdf')
    timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
    response['Content-Disposition'] = f'attachment; filename="components_export_{timestamp}.pdf"'
    
    # Create document with A4 size and 20mm margins
    doc = SimpleDocTemplate(
        response,
        pagesize=A4,
        rightMargin=20*mm,
        leftMargin=20*mm,
        topMargin=20*mm,
        bottomMargin=20*mm
    )
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=20,
        textColor=colors.darkblue
    )
    title = Paragraph("Component Management Report", title_style)
    elements.append(title)
    
    # Export info
    export_info = f"Generated on: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}<br/>Total Components: {components_qs.count()}"
    if any([project_id, component_name, currency_id, allocation_min, allocation_max, date_from, date_to]):
        export_info += "<br/>Filters Applied: Yes"
    else:
        export_info += "<br/>Filters Applied: None (All Components)"
    
    info_para = Paragraph(export_info, styles['Normal'])
    elements.append(info_para)
    elements.append(Spacer(1, 20))
    
    # Prepare table data
    data = []
    
    # Headers
    headers = ['Comp ID', 'Project Name', 'Component Name', 'Currency', 'Allocation', 'Date']
    data.append(headers)
    
    # Cell style for text wrapping - improved for A4 portrait
    cell_style = ParagraphStyle(
        'CellStyle',
        parent=styles['Normal'],
        fontSize=7,  # Smaller font for better fit
        leading=8,   # Tighter line spacing
        wordWrap='CJK',  # Better word wrapping
        splitLongWords=True,  # Split long words if necessary
        leftIndent=1,
        rightIndent=1
    )
    
    # Add component data with text wrapping
    for component in components_qs.order_by('-date'):
        # Use project name instead of project ID and wrap text properly
        project_name = component.projectID.project if component.projectID else ''
        component_name = component.project_components or ''
        
        row = [
            str(component.compID),
            Paragraph(project_name, cell_style),  # Project name with text wrapping
            Paragraph(component_name, cell_style),  # Component name with text wrapping
            str(component.currency.currency if component.currency else ''),
            f"{component.allocation:,.2f}" if component.allocation else '0.00',
            component.date.strftime('%Y-%m-%d') if component.date else '',
        ]
        data.append(row)
    
    # Create table with optimized column widths for A4 portrait (170mm available width)
    # Adjusted widths to prevent overflow and accommodate text wrapping
    table = Table(data, colWidths=[
        12*mm,  # Comp ID (reduced)
        45*mm,  # Project Name (increased for long project names)
        55*mm,  # Component Name (reduced but still adequate)
        18*mm,  # Currency (reduced)
        22*mm,  # Allocation (reduced)
        18*mm,  # Date (reduced)
    ])
    
    # Apply table style
    table.setStyle(TableStyle([
        # Header style
        ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        
        # Data style
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        
        # Text wrapping and cell height
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        
        # Borders
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        
        # Alignment for specific columns
        ('ALIGN', (4, 1), (4, -1), 'RIGHT'),  # Allocation column right-aligned
        ('ALIGN', (5, 1), (5, -1), 'CENTER'), # Date column center-aligned
        
        # Padding - reduced for better fit
        ('TOPPADDING', (0, 1), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
        ('LEFTPADDING', (0, 0), (-1, -1), 3),
        ('RIGHTPADDING', (0, 0), (-1, -1), 3),
    ]))
    
    elements.append(table)
    
    # Build PDF
    doc.build(elements)
    return response

@login_required
def component_detail(request, component_id):
    """View details of a specific component"""
    component = get_object_or_404(Component, compID=component_id)
    
    # Get related subcomponents
    subcomponents = Subcomponent.objects.filter(compID=component)
    
    context = {
        'component': component,
        'subcomponents': subcomponents,
    }
    return render(request, 'PIU_Financial_mgt/components/component_detail.html', context)

@login_required
def edit_component(request, component_id):
    """Edit an existing component"""
    from django.contrib import messages
    component = get_object_or_404(Component, compID=component_id)
    
    if request.method == 'POST':
        print(f"=== EDIT COMPONENT DEBUG ===")
        print(f"POST data: {request.POST}")
        print(f"Component ID: {component_id}")
        print(f"Current component: {component.project_components}")
        
        # Create a mutable copy of POST data and add the projectID
        post_data = request.POST.copy()
        post_data['projectID'] = component.projectID.pk
        
        form = addComponentForm(post_data, instance=component)
        print(f"Form is valid: {form.is_valid()}")
        print(f"Form errors: {form.errors}")
        
        if form.is_valid():
            try:
                component = form.save(commit=False)
                component.loginUser = request.user
                component.save()
                print(f"Component saved successfully: {component.project_components}")
                messages.success(request, 'Component updated successfully!')
                return redirect('PIU_Financial_mgt:components')
            except Exception as e:
                print(f"Error saving component: {e}")
                messages.error(request, f'Error saving component: {str(e)}')
                # Continue to re-render form with error
        else:
            print(f"Form errors: {form.errors}")
            messages.error(request, f'Form validation failed: {form.errors}')
    else:
        form = addComponentForm(instance=component)
    
    context = {
        'form': form,
        'component': component,
        'is_edit': True,
    }
    return render(request, 'PIU_Financial_mgt/components/add-component.html', context)

@login_required
def delete_component(request, component_id):
    """Delete a component"""
    component = get_object_or_404(Component, compID=component_id)
    
    if request.method == 'POST':
        project_id = component.projectID.projectID
        component.delete()
        messages.success(request, 'Component deleted successfully!')
        return redirect('PIU_Financial_mgt:components')
    
    context = {
        'component': component,
        'delete_url': 'PIU_Financial_mgt:delete_component',
        'cancel_url': 'PIU_Financial_mgt:components',
    }
    return render(request, 'PIU_Financial_mgt/components/delete_component.html', context)

@login_required
def addsubcomponent(request):
    print(f"=== ADD SUBCOMPONENT VIEW ===")
    print(f"Method: {request.method}")
    
    if request.method == 'POST':
        print(f"POST data: {request.POST}")
        print(f"Files: {request.FILES}")
        
        # Check if this is an actual form submission with data vs just project selection
        has_submit_data = (request.POST.get('subcomponent', '').strip() and 
                          request.POST.get('subcomponent_Description', '').strip() and 
                          request.POST.get('allocation', '').strip() and 
                          request.POST.get('compID', '').strip())
        print(f"Has actual form data: {has_submit_data}")
        
        if not has_submit_data:
            # This is likely just a project selection change, redirect to GET to avoid POST loop
            print("No form data detected, redirecting to GET")
            return redirect('PIU_Financial_mgt:add_subcomponent')
        else:
            # This is a real form submission with data
            # Using standard addsubcomponentForm from forms.py
            form = addsubcomponentForm(request.POST)
            print(f"Form created: {form}")
            print(f"Form is valid: {form.is_valid()}")
            
            if not form.is_valid():
                print(f"Form errors: {form.errors}")
                print(f"Form non_field_errors: {form.non_field_errors()}")
                for field, errors in form.errors.items():
                    print(f"Field '{field}' errors: {errors}")
            
            if form.is_valid():
                print("Form is valid, saving...")
                subcomponent = form.save(commit=False)
                subcomponent.loginUser = request.user
                subcomponent.save()
                print(f"Subcomponent saved: {subcomponent}")
                messages.success(request, 'Subcomponent created successfully!')
                return redirect('PIU_Financial_mgt:subcomponents')
            else:
                print("Form validation failed, rendering with errors")
    else:
        print("GET request - creating new form")
        # Using standard forms
        form = addsubcomponentForm()
    
    # Get filter options for dropdowns
    projects = Project.objects.all()
    currencies = Currency.objects.all()
    
    context = {
        'form': form,
        'projects': projects,
        'currencies': currencies,
    }
    
    return render(request, 'PIU_Financial_mgt/subcomponents/add_subcomponent.html', context)

@login_required
def subcomponents(request):
    """Enhanced subcomponents list with filtering, pagination, and statistics"""
    from django.db.models import Sum, Count, Q
    from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
    from PIU_Financial_mgt.models import Currency
    
    print(f"=== SUBCOMPONENTS VIEW ACCESSED ===")
    print(f"User: {request.user}")
    print(f"User authenticated: {request.user.is_authenticated}")
    print(f"====================================")
    
    # Get all subcomponents
    try:
        subcomponents_qs = Subcomponent.objects.all().select_related('projectID', 'compID', 'currency', 'loginUser')
        print(f"Successfully created subcomponent QuerySet: {subcomponents_qs.count()} items")
    except Exception as e:
        print(f"Error creating subcomponent QuerySet: {e}")
        subcomponents_qs = Subcomponent.objects.all()
    
    # Filter parameters
    project_filter = request.GET.get('project', '')
    component_filter = request.GET.get('component', '')
    currency_filter = request.GET.get('currency', '')
    search_filter = request.GET.get('search', '')
    
    # Apply filters with correct field references
    is_filtered = False
    
    if project_filter:
        subcomponents_qs = subcomponents_qs.filter(projectID__projectID__icontains=project_filter)
        is_filtered = True
    
    if component_filter:
        subcomponents_qs = subcomponents_qs.filter(compID__project_components__icontains=component_filter)
        is_filtered = True
    
    if currency_filter:
        subcomponents_qs = subcomponents_qs.filter(currency__id=currency_filter)
        is_filtered = True
    
    if search_filter:
        subcomponents_qs = subcomponents_qs.filter(
            Q(subcomponent__icontains=search_filter) |
            Q(subcomponent_Description__icontains=search_filter) |
            Q(projectID__project__icontains=search_filter) |
            Q(compID__project_components__icontains=search_filter)
        )
        is_filtered = True
    
    # Calculate statistics
    overall_total = Subcomponent.objects.count()
    filtered_count = subcomponents_qs.count()
    total_allocation = subcomponents_qs.aggregate(Sum('allocation'))['allocation__sum'] or 0
    unique_projects = subcomponents_qs.values('projectID').distinct().count()
    
    # For project-specific filtering, calculate total for that project only
    if project_filter:
        project_total = Subcomponent.objects.filter(projectID__projectID__icontains=project_filter).count()
    else:
        project_total = overall_total
    
    stats = {
        'total_subcomponents': project_total,
        'overall_total': overall_total,
        'filtered_count': filtered_count,
        'total_allocation': total_allocation,
        'unique_projects': unique_projects,
    }
    
    # Order subcomponents
    ordered_subcomponents = subcomponents_qs.order_by('-date')
    
    # Debug subcomponent data
    print(f"=== SUBCOMPONENT DATA DEBUG ===")
    print(f"Total subcomponents before pagination: {ordered_subcomponents.count()}")
    if ordered_subcomponents.exists():
        sample = ordered_subcomponents.first()
        if sample and sample.projectID:
            print(f"Sample subcomponent: {sample.subcomponent} | Project: {sample.projectID.projectID}")
        else:
            print(f"Sample subcomponent: {sample.subcomponent if sample else 'None'} | Project: None")
    print(f"==============================")
    
    # Pagination - 10 records per page for better UX
    paginator = Paginator(ordered_subcomponents, 10)
    page = request.GET.get('page')
    
    try:
        subcomponents = paginator.page(page)
        print(f"Paginated subcomponents: {len(subcomponents)} items on page {subcomponents.number}")
    except PageNotAnInteger:
        # If page is not an integer, deliver first page
        subcomponents = paginator.page(1)
        print(f"PageNotAnInteger - showing first page with {len(subcomponents)} items")
    except EmptyPage:
        # If page is out of range, deliver last page
        subcomponents = paginator.page(paginator.num_pages)
        print(f"EmptyPage - showing last page with {len(subcomponents)} items")
    
    # Get filter options
    projects = Project.objects.all()
    components = Component.objects.all()
    currencies = Currency.objects.all()
    
    # is_filtered already set above
    
    context = {
        'subcomponents': subcomponents,  # Now a paginated object
        'stats': stats,
        'is_filtered': is_filtered,
        'projects': projects,
        'components': components,
        'currencies': currencies,
    }
    
    return render(request, 'PIU_Financial_mgt/subcomponents/enhanced_subcomponent_list.html', context)


@login_required
def export_subcomponents_excel(request):
    """Export subcomponents to Excel with professional formatting"""
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse
    from datetime import datetime
    from PIU_Financial_mgt.models import Currency
    
    print("=== SUBCOMPONENTS EXCEL EXPORT DEBUG ===")
    print(f"Excel export accessed by user: {request.user.username}")
    print(f"GET parameters: {request.GET}")
    print("=" * 45)
    
    # Get filtered subcomponents using same logic as list view
    subcomponents_qs = Subcomponent.objects.all().select_related('projectID', 'compID', 'currency', 'loginUser')
    
    # Apply same filters as main view
    project_filter = request.GET.get('project', '')
    component_filter = request.GET.get('component', '')
    currency_filter = request.GET.get('currency', '')
    search_filter = request.GET.get('search', '')
    
    print(f"Excel filters applied - project: '{project_filter}', component: '{component_filter}', currency: '{currency_filter}', search: '{search_filter}'")
    
    if project_filter:
        subcomponents_qs = subcomponents_qs.filter(projectID__projectID__icontains=project_filter)
        print(f"Filtered by project ID: {project_filter}")
    
    if component_filter:
        subcomponents_qs = subcomponents_qs.filter(compID__compID=component_filter)
        print(f"Filtered by component ID: {component_filter}")
    
    if currency_filter:
        subcomponents_qs = subcomponents_qs.filter(currency__id=currency_filter)
        print(f"Filtered by currency ID: {currency_filter}")
    
    if search_filter:
        subcomponents_qs = subcomponents_qs.filter(subcomponent__icontains=search_filter)
        print(f"Filtered by search: {search_filter}")
    
    print(f"Final subcomponent count for Excel export: {subcomponents_qs.count()}")
    
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Subcomponents"
    
    # Headers
    headers = ['Subcomponent ID', 'Project Name', 'Component Name', 'Subcomponent Name', 'Description', 
               'Currency', 'Allocation', 'Date Created', 'Created By']
    ws.append(headers)
    
    # Set column widths for better readability
    column_widths = {
        'A': 15,  # Subcomponent ID
        'B': 35,  # Project Name
        'C': 30,  # Component Name
        'D': 30,  # Subcomponent Name
        'E': 40,  # Description
        'F': 12,  # Currency
        'G': 15,  # Allocation
        'H': 18,  # Date
        'I': 15,  # Created By
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Add data rows
    for subcomponent in subcomponents_qs.order_by('-date'):
        row_data = [
            subcomponent.subcompID,
            subcomponent.projectID.project if subcomponent.projectID else '',
            subcomponent.compID.project_components if subcomponent.compID else '',
            subcomponent.subcomponent or '',
            subcomponent.subcomponent_Description or '',
            subcomponent.currency.currency if subcomponent.currency else '',
            subcomponent.allocation or 0,
            subcomponent.date.strftime('%Y-%m-%d %H:%M') if subcomponent.date else '',
            subcomponent.loginUser.username if subcomponent.loginUser else '',
        ]
        ws.append(row_data)
    
    # Style the header row
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Add borders to all cells
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border
            if cell.row > 1:  # Data rows
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    
    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'subcomponents_export_{timestamp}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    print(f"Excel export completed: {filename}")
    
    return response


@login_required
def export_subcomponents_pdf(request):
    """Export subcomponents to PDF with professional A4 formatting"""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from django.http import HttpResponse
    from datetime import datetime
    from PIU_Financial_mgt.models import Currency
    
    print("=== SUBCOMPONENTS PDF EXPORT DEBUG ===")
    print(f"PDF export accessed by user: {request.user.username}")
    print(f"GET parameters: {request.GET}")
    print("=" * 44)
    
    # Get filtered subcomponents using same logic as list view
    subcomponents_qs = Subcomponent.objects.all().select_related('projectID', 'compID', 'currency', 'loginUser')
    
    # Apply same filters as main view
    project_filter = request.GET.get('project', '')
    component_filter = request.GET.get('component', '')
    currency_filter = request.GET.get('currency', '')
    search_filter = request.GET.get('search', '')
    
    print(f"PDF filters applied - project: '{project_filter}', component: '{component_filter}', currency: '{currency_filter}', search: '{search_filter}'")
    
    if project_filter:
        subcomponents_qs = subcomponents_qs.filter(projectID__projectID__icontains=project_filter)
        print(f"Filtered by project ID: {project_filter}")
    
    if component_filter:
        subcomponents_qs = subcomponents_qs.filter(compID__compID=component_filter)
        print(f"Filtered by component ID: {component_filter}")
    
    if currency_filter:
        subcomponents_qs = subcomponents_qs.filter(currency__id=currency_filter)
        print(f"Filtered by currency ID: {currency_filter}")
    
    if search_filter:
        subcomponents_qs = subcomponents_qs.filter(subcomponent__icontains=search_filter)
        print(f"Filtered by search: {search_filter}")
    
    print(f"Final subcomponent count for PDF export: {subcomponents_qs.count()}")
    
    # Create response
    response = HttpResponse(content_type='application/pdf')
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'subcomponents_export_{timestamp}.pdf'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Create PDF document with A4 portrait page size
    doc = SimpleDocTemplate(response, pagesize=A4, rightMargin=20*mm, leftMargin=20*mm, 
                           topMargin=20*mm, bottomMargin=20*mm)
    
    # Get styles
    styles = getSampleStyleSheet()
    
    # Create title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=20,
        alignment=1,  # Center alignment
    )
    
    title = Paragraph("Project Subcomponents Report", title_style)
    
    # Prepare table data
    data = []
    
    # Headers
    headers = ['Sub ID', 'Project Name', 'Component', 'Subcomponent', 'Currency', 'Allocation', 'Date']
    data.append(headers)
    
    # Cell style for text wrapping - optimized for A4 portrait
    cell_style = ParagraphStyle(
        'CellStyle',
        parent=styles['Normal'],
        fontSize=7,  # Smaller font for better fit
        leading=8,   # Tighter line spacing
        wordWrap='CJK',  # Better word wrapping
        splitLongWords=True,  # Split long words if necessary
        leftIndent=1,
        rightIndent=1
    )
    
    # Add subcomponent data with text wrapping
    for subcomponent in subcomponents_qs.order_by('-date'):
        # Use project and component names with text wrapping
        project_name = subcomponent.projectID.project if subcomponent.projectID else ''
        component_name = subcomponent.compID.project_components if subcomponent.compID else ''
        subcomponent_name = subcomponent.subcomponent or ''
        
        row = [
            str(subcomponent.subcompID),
            Paragraph(project_name, cell_style),  # Project name with text wrapping
            Paragraph(component_name, cell_style),  # Component name with text wrapping
            Paragraph(subcomponent_name, cell_style),  # Subcomponent name with text wrapping
            str(subcomponent.currency.currency if subcomponent.currency else ''),
            f"{subcomponent.allocation:,.2f}" if subcomponent.allocation else '0.00',
            subcomponent.date.strftime('%Y-%m-%d') if subcomponent.date else '',
        ]
        data.append(row)
    
    # Create table with optimized column widths for A4 portrait (170mm available width)
    # Adjusted widths to prevent overflow and accommodate text wrapping
    table = Table(data, colWidths=[
        12*mm,  # Sub ID (reduced)
        40*mm,  # Project Name (adequate for long names)
        30*mm,  # Component (adequate)
        35*mm,  # Subcomponent (adequate)
        15*mm,  # Currency (reduced)
        20*mm,  # Allocation (reduced)
        18*mm,  # Date (reduced)
    ])
    
    # Apply table style
    table.setStyle(TableStyle([
        # Header styling
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        
        # Data rows
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
        
        # Borders
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('LINEBELOW', (0, 0), (-1, 0), 2, colors.HexColor('#366092')),
        
        # Padding
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('LEFTPADDING', (0, 0), (-1, -1), 3),
        ('RIGHTPADDING', (0, 0), (-1, -1), 3),
    ]))
    
    # Build PDF
    elements = [title, Spacer(1, 12), table]
    doc.build(elements)
    
    print(f"PDF export completed: {filename}")
    return response

@login_required
def addactivity(request):
    if request.method == 'POST':
        # Using standard forms
        form = addActivitiesForm(request.POST)
        if form.is_valid():
            activity = form.save(commit=False)
            activity.loginUser = request.user  # Set the logged-in user
            activity.save()
            messages.success(request, 'Activity created successfully!')
            return redirect('PIU_Financial_mgt:activities')  
    else:
        # Using standard forms
        form = addActivitiesForm()
    
    # Get all projects, currencies, and years for the dropdowns
    projects = Project.objects.all()
    currencies = Currency.objects.all()
    years = YEAR.objects.all()
    
    context = {
        'form': form,
        'projects': projects,
        'currencies': currencies,
        'years': years,
    }
    return render(request, 'PIU_Financial_mgt/activity/add-activity.html', context)

@login_required
def load_project_components(request):
    """HTMX view to load components for selected project"""
    project_id = request.GET.get('projectID')
    components = Component.objects.none()
    
    if project_id:
        components = Component.objects.filter(projectID=project_id)
    
    return render(request, 'PIU_Financial_mgt/htmx/components_dropdown.html', {'components': components})

@login_required
def load_project_subcomponents(request):
    """HTMX view to load subcomponents for selected component"""
    component_id = request.GET.get('compID')
    print(f"=== LOAD SUBCOMPONENTS DEBUG ===")
    print(f"Component ID received: {component_id}")
    
    subcomponents = Subcomponent.objects.none()
    
    if component_id:
        try:
            # Convert to int and filter by compID (foreign key)
            comp_id_int = int(component_id)
            
            # Try different ways to query subcomponents
            print(f"Trying compID_id={comp_id_int}")
            subcomponents_by_id = Subcomponent.objects.filter(compID_id=comp_id_int)
            print(f"Found {subcomponents_by_id.count()} subcomponents using compID_id")
            
            print(f"Trying compID={comp_id_int}")
            subcomponents_by_field = Subcomponent.objects.filter(compID=comp_id_int)
            print(f"Found {subcomponents_by_field.count()} subcomponents using compID")
            
            # Check all subcomponents to see what exists
            all_subcomps = Subcomponent.objects.all()
            print(f"Total subcomponents in database: {all_subcomps.count()}")
            for sub in all_subcomps[:5]:  # Show first 5
                print(f"- SubcompID: {sub.subcompID}, Component: {sub.compID}, CompID value: {sub.compID.compID if sub.compID else 'None'}")
            
            # Use the result that has subcomponents
            subcomponents = subcomponents_by_field if subcomponents_by_field.exists() else subcomponents_by_id
            
            print(f"Final result: {subcomponents.count()} subcomponents for component {comp_id_int}")
            for sub in subcomponents:
                print(f"- {sub.subcompID}: {sub.subcomponent}")
                
        except (ValueError, TypeError) as e:
            print(f"Error converting component_id to int: {e}")
    
    return render(request, 'PIU_Financial_mgt/htmx/subcomponents_dropdown.html', {'subcomponents': subcomponents})

@login_required
def activities(request):
    """Enhanced activities list with filtering, pagination, and statistics"""
    from django.db.models import Sum, Count
    from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
    from PIU_Financial_mgt.models import Currency
    
    # Get all activities with proper field references
    activities_qs = Activities.objects.all().select_related('projectID', 'compID', 'subcompID', 'currency')
    
    # Filter parameters
    project_filter = request.GET.get('project', '')
    component_filter = request.GET.get('component', '')
    subcomponent_filter = request.GET.get('subcomponent', '')
    currency_filter = request.GET.get('currency', '')
    year_filter = request.GET.get('year', '')
    search_filter = request.GET.get('search', '')
    
    # Apply filters with correct field names
    if project_filter:
        activities_qs = activities_qs.filter(projectID__projectID__icontains=project_filter)
    
    if component_filter:
        activities_qs = activities_qs.filter(compID__compID=component_filter)
    
    if subcomponent_filter:
        activities_qs = activities_qs.filter(subcompID__subcompID=subcomponent_filter)
    
    if currency_filter:
        activities_qs = activities_qs.filter(currency__id=currency_filter)
    
    if year_filter:
        activities_qs = activities_qs.filter(year=year_filter)
    
    if search_filter:
        activities_qs = activities_qs.filter(
            activity__icontains=search_filter
        )
    
    # Calculate statistics
    total_activities = Activities.objects.count()
    filtered_count = activities_qs.count()
    total_allocation = activities_qs.aggregate(Sum('allocation'))['allocation__sum'] or 0
    unique_projects = activities_qs.values('projectID').distinct().count()
    
    stats = {
        'total_activities': total_activities,
        'filtered_count': filtered_count,
        'total_allocation': total_allocation,
        'unique_projects': unique_projects,
    }
    
    # Order activities by year and date (newest first)
    activities_qs = activities_qs.order_by('-year', '-date')
    
    # Pagination - 5 records per page
    paginator = Paginator(activities_qs, 5)
    page = request.GET.get('page')
    
    try:
        activities = paginator.page(page)
    except PageNotAnInteger:
        activities = paginator.page(1)
    except EmptyPage:
        activities = paginator.page(paginator.num_pages)
    
    # Debug pagination - remove after testing
    # print(f"DEBUG PAGINATION: Page count: {paginator.count}, Num pages: {paginator.num_pages}")
    # print(f"DEBUG CURRENT PAGE: {activities.number}, Has activities: {len(activities.object_list)}")
    
    # Get filter options
    projects = Project.objects.all()
    components = Component.objects.all()
    subcomponents = Subcomponent.objects.all()
    currencies = Currency.objects.all()
    years = YEAR.objects.all().order_by('-profile_year')
    
    is_filtered = bool(project_filter or component_filter or subcomponent_filter or currency_filter or year_filter or search_filter)
    
    context = {
        'activities': activities,
        'stats': stats,
        'is_filtered': is_filtered,
        'projects': projects,
        'components': components,
        'subcomponents': subcomponents,
        'currencies': currencies,
        'years': years,
    }
    
    return render(request, 'PIU_Financial_mgt/activity/activities.html', context)


@login_required
def export_activities_excel(request):
    """Export activities to Excel with professional formatting"""
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse
    from datetime import datetime
    from PIU_Financial_mgt.models import Currency
    
    print("=== ACTIVITIES EXCEL EXPORT DEBUG ===")
    print(f"Excel export accessed by user: {request.user.username}")
    print(f"GET parameters: {request.GET}")
    print("=" * 42)
    
    # Get filtered activities using same logic as list view
    activities_qs = Activities.objects.all().select_related('projectID', 'compID', 'subcompID', 'currency')
    
    # Apply same filters as main view
    project_filter = request.GET.get('project', '')
    component_filter = request.GET.get('component', '')
    subcomponent_filter = request.GET.get('subcomponent', '')
    currency_filter = request.GET.get('currency', '')
    year_filter = request.GET.get('year', '')
    search_filter = request.GET.get('search', '')
    
    print(f"Excel filters applied - project: '{project_filter}', component: '{component_filter}', subcomponent: '{subcomponent_filter}', currency: '{currency_filter}', year: '{year_filter}', search: '{search_filter}'")
    
    if project_filter:
        activities_qs = activities_qs.filter(projectID__projectID__icontains=project_filter)
        print(f"Filtered by project ID: {project_filter}")
    
    if component_filter:
        activities_qs = activities_qs.filter(compID__compID=component_filter)
        print(f"Filtered by component ID: {component_filter}")
    
    if subcomponent_filter:
        activities_qs = activities_qs.filter(subcompID__subcompID=subcomponent_filter)
        print(f"Filtered by subcomponent ID: {subcomponent_filter}")
    
    if currency_filter:
        activities_qs = activities_qs.filter(currency__id=currency_filter)
        print(f"Filtered by currency ID: {currency_filter}")
    
    if year_filter:
        activities_qs = activities_qs.filter(year=year_filter)
        print(f"Filtered by year: {year_filter}")
    
    if search_filter:
        activities_qs = activities_qs.filter(activity__icontains=search_filter)
        print(f"Filtered by search: {search_filter}")
    
    print(f"Final activity count for Excel export: {activities_qs.count()}")
    
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Activities"
    
    # Headers
    headers = ['Activity ID', 'Project Name', 'Component Name', 'Subcomponent Name', 'Activity Name', 
               'Currency', 'Allocation', 'Year', 'Date Created', 'Created By']
    ws.append(headers)
    
    # Set column widths for better readability
    column_widths = {
        'A': 12,  # Activity ID
        'B': 30,  # Project Name
        'C': 25,  # Component Name
        'D': 25,  # Subcomponent Name
        'E': 35,  # Activity Name
        'F': 10,  # Currency
        'G': 15,  # Allocation
        'H': 8,   # Year
        'I': 18,  # Date
        'J': 15,  # Created By
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Add data rows
    for activity in activities_qs.order_by('-year', '-date'):
        row_data = [
            activity.activityID,
            activity.projectID.project if activity.projectID else '',
            activity.compID.project_components if activity.compID else '',
            activity.subcompID.subcomponent if activity.subcompID else '',
            activity.activity or '',
            activity.currency.currency if activity.currency else '',
            activity.allocation or 0,
            activity.year.profile_year if activity.year else '',
            activity.date.strftime('%Y-%m-%d %H:%M') if activity.date else '',
            activity.loginUser.username if activity.loginUser else '',
        ]
        ws.append(row_data)
    
    # Style the header row
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Add borders to all cells
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border
            if cell.row > 1:  # Data rows
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    
    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'activities_export_{timestamp}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    print(f"Excel export completed: {filename}")
    
    return response


@login_required
def export_activities_pdf(request):
    """Export activities to PDF with professional A4 formatting"""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from django.http import HttpResponse
    from datetime import datetime
    from PIU_Financial_mgt.models import Currency
    
    print("=== ACTIVITIES PDF EXPORT DEBUG ===")
    print(f"PDF export accessed by user: {request.user.username}")
    print(f"GET parameters: {request.GET}")
    print("=" * 41)
    
    # Get filtered activities using same logic as list view
    activities_qs = Activities.objects.all().select_related('projectID', 'compID', 'subcompID', 'currency')
    
    # Apply same filters as main view
    project_filter = request.GET.get('project', '')
    component_filter = request.GET.get('component', '')
    subcomponent_filter = request.GET.get('subcomponent', '')
    currency_filter = request.GET.get('currency', '')
    year_filter = request.GET.get('year', '')
    search_filter = request.GET.get('search', '')
    
    print(f"PDF filters applied - project: '{project_filter}', component: '{component_filter}', subcomponent: '{subcomponent_filter}', currency: '{currency_filter}', year: '{year_filter}', search: '{search_filter}'")
    
    if project_filter:
        activities_qs = activities_qs.filter(projectID__projectID__icontains=project_filter)
        print(f"Filtered by project ID: {project_filter}")
    
    if component_filter:
        activities_qs = activities_qs.filter(compID__compID=component_filter)
        print(f"Filtered by component ID: {component_filter}")
    
    if subcomponent_filter:
        activities_qs = activities_qs.filter(subcompID__subcompID=subcomponent_filter)
        print(f"Filtered by subcomponent ID: {subcomponent_filter}")
    
    if currency_filter:
        activities_qs = activities_qs.filter(currency__id=currency_filter)
        print(f"Filtered by currency ID: {currency_filter}")
    
    if year_filter:
        activities_qs = activities_qs.filter(year=year_filter)
        print(f"Filtered by year: {year_filter}")
    
    if search_filter:
        activities_qs = activities_qs.filter(activity__icontains=search_filter)
        print(f"Filtered by search: {search_filter}")
    
    print(f"Final activity count for PDF export: {activities_qs.count()}")
    
    # Create response
    response = HttpResponse(content_type='application/pdf')
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'activities_export_{timestamp}.pdf'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Create PDF document with A4 portrait page size
    doc = SimpleDocTemplate(response, pagesize=A4, rightMargin=20*mm, leftMargin=20*mm, 
                           topMargin=20*mm, bottomMargin=20*mm)
    
    # Get styles
    styles = getSampleStyleSheet()
    
    # Create title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=20,
        alignment=1,  # Center alignment
    )
    
    title = Paragraph("Project Activities Report", title_style)
    
    # Prepare table data
    data = []
    
    # Headers
    headers = ['Activity ID', 'Project', 'Component', 'Subcomponent', 'Activity', 'Currency', 'Allocation', 'Year']
    data.append(headers)
    
    # Cell style for text wrapping - optimized for A4 portrait
    cell_style = ParagraphStyle(
        'CellStyle',
        parent=styles['Normal'],
        fontSize=7,  # Smaller font for better fit
        leading=8,   # Tighter line spacing
        wordWrap='CJK',  # Better word wrapping
        splitLongWords=True,  # Split long words if necessary
        leftIndent=1,
        rightIndent=1
    )
    
    # Add activity data with text wrapping
    for activity in activities_qs.order_by('-year', '-date'):
        # Use project, component, subcomponent and activity names with text wrapping
        project_name = activity.projectID.project if activity.projectID else ''
        component_name = activity.compID.project_components if activity.compID else ''
        subcomponent_name = activity.subcompID.subcomponent if activity.subcompID else ''
        activity_name = activity.activity or ''
        
        row = [
            str(activity.activityID),
            Paragraph(project_name, cell_style),  # Project name with text wrapping
            Paragraph(component_name, cell_style),  # Component name with text wrapping
            Paragraph(subcomponent_name, cell_style),  # Subcomponent name with text wrapping
            Paragraph(activity_name, cell_style),  # Activity name with text wrapping
            str(activity.currency.currency if activity.currency else ''),
            f"{activity.allocation:,.2f}" if activity.allocation else '0.00',
            str(activity.year.profile_year if activity.year else ''),
        ]
        data.append(row)
    
    # Create table with optimized column widths for A4 portrait (170mm available width)
    # Adjusted widths to prevent overflow and accommodate text wrapping
    table = Table(data, colWidths=[
        15*mm,  # Activity ID
        35*mm,  # Project (adequate for long names)
        25*mm,  # Component
        25*mm,  # Subcomponent
        40*mm,  # Activity (increased width for longer activity descriptions)
        12*mm,  # Currency
        18*mm,  # Allocation
        10*mm,  # Year
    ])
    
    # Apply table style
    table.setStyle(TableStyle([
        # Header styling
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        
        # Data rows
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
        
        # Borders
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('LINEBELOW', (0, 0), (-1, 0), 2, colors.HexColor('#366092')),
        
        # Padding
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('LEFTPADDING', (0, 0), (-1, -1), 3),
        ('RIGHTPADDING', (0, 0), (-1, -1), 3),
    ]))
    
    # Build PDF
    elements = [title, Spacer(1, 12), table]
    doc.build(elements)
    
    print(f"PDF export completed: {filename}")
    return response


############################################ Dashboard ############################################

@login_required
def dashboard(request):
    projects = Project.objects.all()
    components = Component.objects.all()
    subcomponents = Subcomponent.objects.all()
    activities = Activities.objects.all()
    
    context = {
        'projects': projects,
        'components': components,
        'subcomponents': subcomponents,
        'activities': activities,
    }
    return render(request, 'PIU_Financial_mgt/dashboard.html', context)

############################################ Budget Summary ############################################

@login_required
def budget_summary(request):
    # Get filter parameters
    year_filter = request.GET.get('year')
    project_filter = request.GET.get('project')
    component_filter = request.GET.get('component')
    subcomponent_filter = request.GET.get('subcomponent')
    
    # Start with all projects
    projects = Project.objects.all()
    
    # Apply filters
    if project_filter:
        try:
            project_id = int(project_filter)
            projects = projects.filter(projectID=project_id)
        except (ValueError, TypeError):
            pass
    
    # Get related components and subcomponents
    components = Component.objects.all()
    if component_filter:
        try:
            component_id = int(component_filter)
            components = components.filter(compID=component_id)
        except (ValueError, TypeError):
            pass
    
    subcomponents = Subcomponent.objects.all()
    if subcomponent_filter:
        try:
            subcomponent_id = int(subcomponent_filter)
            subcomponents = subcomponents.filter(subcompID=subcomponent_id)
        except (ValueError, TypeError):
            pass
    
    # Get all available filters for the form
    all_projects = Project.objects.all()
    all_components = Component.objects.all()
    all_subcomponents = Subcomponent.objects.all()
    
    # Calculate total budget
    total_budget = sum(project.funding for project in projects if project.funding) if projects else 0
    
    context = {
        'projects': projects,
        'components': components,
        'subcomponents': subcomponents,
        'all_projects': all_projects,
        'all_components': all_components,
        'all_subcomponents': all_subcomponents,
        'total_budget': total_budget,
        'year_filter': year_filter,
        'project_filter': project_filter,
        'component_filter': component_filter,
        'subcomponent_filter': subcomponent_filter,
    }
    
    return render(request, 'PIU_Financial_mgt/budget_summary.html', context)

# Removed another duplicate function - using the enhanced version with component filtering

@login_required
def simple_financial_dashboard(request):
    """
    Comprehensive financial dashboard with statistics and recent activities
    """
    from django.db.models import Sum, Count, Avg
    from datetime import datetime, timedelta
    
    # Basic statistics
    total_projects = Project.objects.count()
    total_components = Component.objects.count()
    total_subcomponents = Subcomponent.objects.count()
    total_activities = Activities.objects.count()
    
    # Financial calculations
    total_funding = Project.objects.aggregate(
        total=Sum('funding')
    )['total'] or 0
    
    avg_project_funding = Project.objects.aggregate(
        avg=Avg('funding')
    )['avg'] or 0
    
    # Active projects (projects with recent activity or within date range)
    today = datetime.now().date()
    thirty_days_ago = today - timedelta(days=30)
    active_projects = Project.objects.filter(
        date__gte=thirty_days_ago
    ).count()
    
    # Currency count
    currencies_count = Currency.objects.count()
    
    # Recent projects (last 5)
    recent_projects = list(Project.objects.order_by('-date')[:5])
    
    # Top funded projects (top 5)
    top_funded = list(Project.objects.filter(
        funding__isnull=False
    ).order_by('-funding')[:5])
    
    # Recent components (last 5) - convert to list for template evaluation
    recent_components = list(Component.objects.order_by('-date')[:5])
    
    # Recent subcomponents (last 5) - convert to list for template evaluation
    recent_subcomponents = list(Subcomponent.objects.order_by('-date')[:5])
    
    # Recent activities (last 5) - Use Django ORM for SQLite compatibility - convert to list for template evaluation
    try:
        recent_activities = list(Activities.objects.select_related('projectID', 'compID', 'subcompID').order_by('-date')[:5])
        print(f"Successfully loaded {len(recent_activities)} recent activities using Django ORM")
    except Exception as e:
        print(f"Error loading recent activities: {e}")
        recent_activities = []
    
    context = {
        'stats': {
            'total_projects': total_projects,
            'total_components': total_components,
            'total_subcomponents': total_subcomponents,
            'total_activities': total_activities,
        },
        'total_projects': total_projects,
        'total_components': total_components,
        'total_subcomponents': total_subcomponents,
        'total_activities': total_activities,
        'total_funding': total_funding,
        'avg_project_funding': avg_project_funding,
        'active_projects': active_projects,
        'currencies_count': currencies_count,
        'recent_projects': recent_projects,
        'top_funded': top_funded,
        'recent_components': recent_components,
        'recent_subcomponents': recent_subcomponents,
        'recent_activities': recent_activities,
    }
    
    return render(request, 'PIU_Financial_mgt/simple_dashboard.html', context)

@login_required
def export_projects_excel(request):
    """Export projects to Excel with applied filters"""
    from django.http import HttpResponse
    from openpyxl import Workbook
    from setup.models import Donor
    from PIU_Financial_mgt.models import Currency
    import datetime
    
    print(f"=== EXCEL EXPORT DEBUG ===")
    print(f"Excel export accessed by user: {request.user}")
    print(f"User authenticated: {request.user.is_authenticated}")
    print(f"Request method: {request.method}")
    print(f"GET parameters: {request.GET}")
    print(f"Request path: {request.path}")
    print(f"===========================")
    
    # Apply same filtering logic as projects view
    projects_qs = Project.objects.all().select_related('currency').prefetch_related('donors', 'contributors')
    
    # Filter parameters (same as projects view)
    project_id = request.GET.get('projectID', '')
    project_name = request.GET.get('project', '')
    donor_id = request.GET.get('donor', '')  # Fixed to match template parameter name
    currency_id = request.GET.get('currency', '')
    funding_min = request.GET.get('funding_min', '')
    funding_max = request.GET.get('funding_max', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    # Apply filters
    if project_id:
        projects_qs = projects_qs.filter(projectID__icontains=project_id)
    if project_name:
        projects_qs = projects_qs.filter(project__icontains=project_name)
    if donor_id:
        projects_qs = projects_qs.filter(donors__donorID=donor_id)
    if currency_id:
        projects_qs = projects_qs.filter(currency__id=currency_id)
    if funding_min:
        try:
            projects_qs = projects_qs.filter(funding__gte=float(funding_min))
        except ValueError:
            pass
    if funding_max:
        try:
            projects_qs = projects_qs.filter(funding__lte=float(funding_max))
        except ValueError:
            pass
    if date_from:
        projects_qs = projects_qs.filter(date__gte=date_from)
    if date_to:
        projects_qs = projects_qs.filter(date__lte=date_to)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Projects"
    
    # Headers
    headers = ['Project ID', 'Project Name', 'Donors', 'Contributors', 'Currency', 'Funding', 
               'Effectiveness Date', 'Closure Date', 'Last Disbursement Date', 'Created Date', 'Created By']
    ws.append(headers)
    
    # Set column widths for better readability
    from openpyxl.utils import get_column_letter
    column_widths = {
        'A': 15,  # Project ID
        'B': 50,  # Project Name (lengthier cell)
        'C': 30,  # Donors
        'D': 30,  # Contributors
        'E': 12,  # Currency
        'F': 15,  # Funding
        'G': 18,  # Effectiveness Date
        'H': 15,  # Closure Date
        'I': 20,  # Last Disbursement Date
        'J': 15,  # Created Date
        'K': 15,  # Created By
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Style headers
    from openpyxl.styles import Font, PatternFill, Alignment
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Data rows
    for project in projects_qs.order_by('-date'):
        donors_list = ', '.join([donor.name for donor in project.donors.all()])
        contributors_list = ', '.join([contrib.name for contrib in project.contributors.all()])
        
        row = [
            project.projectID,
            project.project,
            donors_list,
            contributors_list,
            project.currency.currency if project.currency else '',
            float(project.funding) if project.funding else 0,
            project.effectiveness_Date.strftime('%Y-%m-%d') if project.effectiveness_Date else '',
            project.closure_Date.strftime('%Y-%m-%d') if project.closure_Date else '',
            project.last_date_of_Disbursement.strftime('%Y-%m-%d') if project.last_date_of_Disbursement else '',
            project.date.strftime('%Y-%m-%d') if project.date else '',
            project.loginUser.username if project.loginUser else ''
        ]
        ws.append(row)
        
        # Apply text wrapping for long content
        row_num = ws.max_row
        for col_num, cell in enumerate(ws[row_num], 1):
            cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
            # Special formatting for funding column
            if col_num == 6 and cell.value:  # Funding column
                cell.number_format = '#,##0.00'
    
    # Prepare response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=projects_export_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    wb.save(response)
    return response

@login_required
def export_projects_pdf(request):
    """Export projects to PDF with A4 portrait layout"""
    from django.http import HttpResponse
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.units import inch
    from setup.models import Donor
    import datetime
    
    print(f"=== PDF EXPORT DEBUG ===")
    print(f"PDF export accessed by user: {request.user}")
    print(f"User authenticated: {request.user.is_authenticated}")
    print(f"Request method: {request.method}")
    print(f"GET parameters: {request.GET}")
    print(f"Request path: {request.path}")
    print(f"========================")
    
    # Apply same filtering logic as projects view
    projects_qs = Project.objects.all().select_related('currency').prefetch_related('donors', 'contributors')
    
    # Filter parameters (same as projects view)
    project_id = request.GET.get('projectID', '')
    project_name = request.GET.get('project', '')
    donor_id = request.GET.get('donor', '')  # Fixed to match template parameter name
    currency_id = request.GET.get('currency', '')
    funding_min = request.GET.get('funding_min', '')
    funding_max = request.GET.get('funding_max', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    print(f"PDF filters applied - project: '{project_name}', donor: '{donor_id}', currency: '{currency_id}'")
    
    # Apply filters
    if project_id:
        projects_qs = projects_qs.filter(projectID__icontains=project_id)
        print(f"Filtered by project ID: {project_id}")
    if project_name:
        projects_qs = projects_qs.filter(project__icontains=project_name)
        print(f"Filtered by project name: {project_name}")
    if donor_id:
        projects_qs = projects_qs.filter(donors__donorID=donor_id)
        print(f"Filtered by donor ID: {donor_id}")
    if currency_id:
        projects_qs = projects_qs.filter(currency__id=currency_id)
    if funding_min:
        try:
            projects_qs = projects_qs.filter(funding__gte=float(funding_min))
        except ValueError:
            pass
    if funding_max:
        try:
            projects_qs = projects_qs.filter(funding__lte=float(funding_max))
        except ValueError:
            pass
    if date_from:
        projects_qs = projects_qs.filter(date__gte=date_from)
    if date_to:
        projects_qs = projects_qs.filter(date__lte=date_to)
    
    # Create response
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=projects_export_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
    
    # Create PDF document with standard A4 margins (20mm)
    from reportlab.lib.units import mm
    doc = SimpleDocTemplate(
        response, 
        pagesize=A4,
        rightMargin=20*mm,
        leftMargin=20*mm,
        topMargin=20*mm,
        bottomMargin=20*mm
    )
    
    # Build content
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=30,
        textColor=colors.HexColor('#366092'),
        alignment=1  # Center alignment
    )
    title = Paragraph("PIU Financial Management - Projects Report", title_style)
    elements.append(title)
    
    # Export date
    date_style = ParagraphStyle(
        'DateStyle',
        parent=styles['Normal'],
        fontSize=10,
        spaceAfter=20,
        alignment=1
    )
    export_date = Paragraph(f"Generated on: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", date_style)
    elements.append(export_date)
    
    # Table data
    data = []
    
    # Headers with improved column distribution for A4 portrait
    headers = ['Project ID', 'Project Name', 'Donors', 'Currency', 'Funding', 'Status']
    data.append(headers)
    
    # Data rows
    for project in projects_qs.order_by('-date'):
        donors_list = ', '.join([donor.name for donor in project.donors.all()])
        
        # Determine project status
        status = "Active"
        if project.closure_Date:
            if project.closure_Date < datetime.date.today():
                status = "Closed"
            else:
                status = "Closing Soon"
        
        # Create Paragraph objects for long text to enable text wrapping with smaller font
        cell_style = ParagraphStyle(
            'CellStyle',
            parent=styles['Normal'],
            fontSize=8,
            leading=10,
            wordWrap='LTR'
        )
        project_name_para = Paragraph(str(project.project) if project.project else '', cell_style)
        donors_para = Paragraph(donors_list if donors_list else '', cell_style)
        
        row = [
            str(project.projectID) if project.projectID else '',
            project_name_para,  # Use Paragraph for text wrapping
            donors_para,        # Use Paragraph for text wrapping
            str(project.currency.currency) if project.currency else '',
            f"{project.funding:,.2f}" if project.funding else '0.00',
            status
        ]
        data.append(row)
    
    # Create table with optimized column widths for A4 portrait with 20mm margins
    # A4 width = 210mm, minus margins (40mm) = 170mm available
    col_widths = [
        20*mm,   # Project ID
        65*mm,   # Project Name (wide for long names)
        25*mm,   # Donors
        15*mm,   # Currency
        25*mm,   # Funding
        20*mm    # Status
    ]  # Total: 170mm - fits exactly within available space
    
    table = Table(data, colWidths=col_widths)
    
    # Table styling
    table.setStyle(TableStyle([
        # Header styling
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        
        # Data styling
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        
        # Text wrapping and cell height
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        
        # Borders
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        
        # Alignment for specific columns
        ('ALIGN', (4, 1), (4, -1), 'RIGHT'),  # Funding column right-aligned
        ('ALIGN', (5, 1), (5, -1), 'CENTER'), # Status column center-aligned
        
        # Padding - reduced for better fit
        ('TOPPADDING', (0, 1), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
        ('LEFTPADDING', (0, 0), (-1, -1), 3),
        ('RIGHTPADDING', (0, 0), (-1, -1), 3),
    ]))
    
    elements.append(table)
    
    # Add summary footer
    elements.append(Spacer(1, 20))
    summary_text = f"Total Projects: {projects_qs.count()}"
    summary = Paragraph(summary_text, styles['Normal'])
    elements.append(summary)
    
    # Build PDF
    doc.build(elements)
    
    return response

@login_required
def export_components_excel(request):
    """Export components to Excel with applied filters"""
    from django.http import HttpResponse
    from openpyxl import Workbook
    from PIU_Financial_mgt.models import Currency
    import datetime
    
    # Apply same filtering logic as components view
    components_qs = Component.objects.all().select_related('projectID', 'currency', 'loginUser')
    
    # Filter parameters (same as components view)
    project_id = request.GET.get('project', '')
    component_name = request.GET.get('component', '')
    currency_id = request.GET.get('currency', '')
    allocation_min = request.GET.get('allocation_min', '')
    allocation_max = request.GET.get('allocation_max', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    # Apply filters
    if project_id:
        components_qs = components_qs.filter(projectID__projectID__icontains=project_id)
    if component_name:
        components_qs = components_qs.filter(project_components__icontains=component_name)
    if currency_id:
        components_qs = components_qs.filter(currency__currency=currency_id)
    if allocation_min:
        try:
            components_qs = components_qs.filter(allocation__gte=float(allocation_min))
        except ValueError:
            pass
    if allocation_max:
        try:
            components_qs = components_qs.filter(allocation__lte=float(allocation_max))
        except ValueError:
            pass
    if date_from:
        components_qs = components_qs.filter(date__gte=date_from)
    if date_to:
        components_qs = components_qs.filter(date__lte=date_to)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Components"
    
    # Headers
    headers = ['Project ID', 'Project Name', 'Component Name', 'Component Description', 
               'Currency', 'Allocation', 'Created Date', 'Created By']
    ws.append(headers)
    
    # Data rows
    for component in components_qs.order_by('-date'):
        row = [
            component.projectID.projectID if component.projectID else '',
            component.projectID.project if component.projectID else '',
            component.project_components,
            component.component_description,
            component.currency.currency if component.currency else '',
            float(component.allocation) if component.allocation else 0,
            component.date.strftime('%Y-%m-%d') if component.date else '',
            component.loginUser.username if component.loginUser else ''
        ]
        ws.append(row)
        
        # Apply text wrapping for long content
        row_num = ws.max_row
        for col_num, cell in enumerate(ws[row_num], 1):
            cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
            # Special formatting for funding column
            if col_num == 6 and cell.value:  # Funding column
                cell.number_format = '#,##0.00'
    
    # Prepare response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=components_export_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    wb.save(response)
    return response

@login_required
def edit_project(request, project_id):
    """Edit an existing project"""
    try:
        project = Project.objects.get(projectID=project_id)
    except Project.DoesNotExist:
        messages.error(request, 'Project not found.')
        return redirect('PIU_Financial_mgt:projects')
    
    if request.method == 'POST':
        form = addProjectForm(request.POST, instance=project)
        if form.is_valid():
            project = form.save(commit=False)
            project.loginUser = request.user
            project.save()
            form.save_m2m()  # Save many-to-many relationships
            messages.success(request, 'Project updated successfully!')
            return redirect('PIU_Financial_mgt:project_list')
        else:
            # Add form error messages for debugging
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"Error in {field}: {error}")
    else:
        form = addProjectForm(instance=project)
    
    context = {
        'form': form,
        'project': project,
        'is_edit': True,
    }
    return render(request, 'PIU_Financial_mgt/projects/edit-project.html', context)


@login_required
def delete_project(request, project_id):
    """Delete a project"""
    try:
        project = Project.objects.get(projectID=project_id)
    except Project.DoesNotExist:
        messages.error(request, 'Project not found.')
        return redirect('PIU_Financial_mgt:projects')
    
    if request.method == 'POST':
        project_name = project.project
        project.delete()
        messages.success(request, f'Project "{project_name}" deleted successfully!')
        return redirect('PIU_Financial_mgt:projects')
    
    context = {
        'project': project,
    }
    return render(request, 'PIU_Financial_mgt/projects/delete_project.html', context)


@login_required
def activity_detail(request, activity_id):
    """View activity details"""
    activity = get_object_or_404(Activities, activityID=activity_id)
    
    context = {
        'activity': activity,
    }
    return render(request, 'PIU_Financial_mgt/activity/activity_detail.html', context)

@login_required
def subcomponent_detail(request, subcomponent_id):
    """View subcomponent details"""
    subcomponent = get_object_or_404(Subcomponent, subcompID=subcomponent_id)
    
    # Get related activities
    activities = Activities.objects.filter(subcompID=subcomponent)
    
    context = {
        'subcomponent': subcomponent,
        'activities': activities,
    }
    return render(request, 'PIU_Financial_mgt/subcomponents/subcomponent_detail.html', context)

@login_required
def edit_subcomponent(request, subcomponent_id):
    """Edit subcomponent"""
    subcomponent = get_object_or_404(Subcomponent, subcompID=subcomponent_id)
    
    if request.method == 'POST':
        form = updatesubcomponentForm(request.POST, instance=subcomponent)
        if form.is_valid():
            try:
                updated_subcomponent = form.save(commit=False)
                updated_subcomponent.loginUser = request.user
                updated_subcomponent.save()
                messages.success(request, 'Subcomponent updated successfully!')
                return redirect('PIU_Financial_mgt:subcomponents')
            except Exception as e:
                messages.error(request, f'Error updating subcomponent: {str(e)}')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = updatesubcomponentForm(instance=subcomponent)
    
    context = {
        'form': form,
        'subcomponent': subcomponent,
        'is_edit': True,
    }
    return render(request, 'PIU_Financial_mgt/subcomponents/edit_subcomponent.html', context)

@login_required
def delete_subcomponent(request, subcomponent_id):
    """Delete subcomponent"""
    subcomponent = get_object_or_404(Subcomponent, subcompID=subcomponent_id)
    
    if request.method == 'POST':
        try:
            subcomponent.delete()
            messages.success(request, 'Subcomponent deleted successfully!')
            return redirect('PIU_Financial_mgt:subcomponents')
        except Exception as e:
            messages.error(request, f'Error deleting subcomponent: {str(e)}')
            return redirect('PIU_Financial_mgt:subcomponent_detail', subcomponent_id=subcomponent_id)
    
    context = {
        'subcomponent': subcomponent,
    }
    return render(request, 'PIU_Financial_mgt/subcomponents/delete_subcomponent.html', context)

# Financial Validation API Endpoints
@require_http_methods(["POST"])
@login_required
def validate_project_funding(request):
    """API endpoint to validate project funding against components total"""
    try:
        data = json.loads(request.body)
        project_id = data.get('project_id')
        funding = Decimal(str(data.get('funding', 0)))
        currency = data.get('currency')
        
        try:
            project = Project.objects.get(projectID=project_id)
            components_total = project.get_total_components_allocation()
            
            is_valid = funding == components_total
            difference = abs(funding - components_total)
            
            response = {
                'is_valid': is_valid,
                'project_funding': float(funding),
                'components_total': float(components_total),
                'difference': float(difference),
                'currency': currency or 'GMD',
                'severity': 'error' if difference > 0 else 'success'
            }
            
            if is_valid:
                response['message'] = f"Project funding matches components total: {funding} {currency}"
            else:
                if funding > components_total:
                    response['message'] = f"Project funding exceeds components total by {difference} {currency}"
                    response['suggestions'] = "Consider adding more components or reducing project funding"
                else:
                    response['message'] = f"Components total exceeds project funding by {difference} {currency}"
                    response['suggestions'] = "Reduce component allocations or increase project funding"
                    
        except Project.DoesNotExist:
            # For new projects, just validate that funding is set
            response = {
                'is_valid': True,
                'project_funding': float(funding),
                'components_total': 0.0,
                'difference': float(funding),
                'currency': currency or 'GMD',
                'severity': 'info',
                'message': f"New project funding set to {funding} {currency or 'GMD'}. Add components to validate allocation.",
                'suggestions': "This is a new project. Components can be added after project creation."
            }
            
    except Exception as e:
        response = {
            'is_valid': False,
            'message': f'Validation error: {str(e)}',
            'severity': 'error'
        }
    
    return JsonResponse(response)

@require_http_methods(["POST"])
@login_required
def validate_component_allocation(request):
    """API endpoint to validate component allocation against project funding and subcomponents"""
    try:
        data = json.loads(request.body)
        component_id = data.get('component_id')
        project_id = data.get('project_id')
        allocation = Decimal(str(data.get('allocation', 0)))
        
        # First validate against project funding if project_id is provided
        if project_id:
            try:
                project = Project.objects.get(projectID=project_id)
                project_funding = project.funding
                
                # Check if this component allocation exceeds project funding
                if allocation > project_funding:
                    return JsonResponse({
                        'is_valid': False,
                        'message': f"Component allocation ({allocation}) cannot exceed project funding ({project_funding})",
                        'severity': 'error',
                        'component_allocation': float(allocation),
                        'project_funding': float(project_funding)
                    })
                
                # Check if total component allocations would exceed project funding
                other_components_total = Component.objects.filter(
                    projectID=project
                ).exclude(compID=component_id if component_id else 0).aggregate(
                    total=Sum('allocation')
                )['total'] or Decimal('0.00')
                
                total_with_this_component = other_components_total + allocation
                if total_with_this_component > project_funding:
                    return JsonResponse({
                        'is_valid': False,
                        'message': f"Total component allocations ({total_with_this_component}) would exceed project funding ({project_funding}). Current total: {other_components_total}",
                        'severity': 'error',
                        'component_allocation': float(allocation),
                        'project_funding': float(project_funding),
                        'current_total': float(other_components_total)
                    })
                    
            except Project.DoesNotExist:
                return JsonResponse({
                    'is_valid': False,
                    'message': 'Project not found',
                    'severity': 'error'
                })
        
        # For existing components, validate against subcomponents if they exist
        if component_id:
            try:
                component = Component.objects.get(compID=component_id)
                subcomponents_total = component.get_total_subcomponents_allocation()
                
                # Only validate against subcomponents if there are any
                if component.subcomponent_set.exists():
                    is_valid = allocation == subcomponents_total
                    difference = abs(allocation - subcomponents_total)
                    
                    if not is_valid:
                        if allocation > subcomponents_total:
                            return JsonResponse({
                                'is_valid': False,
                                'message': f"Component allocation exceeds subcomponents total by {difference}",
                                'severity': 'warning',
                                'suggestions': "Add more subcomponents or reduce component allocation",
                                'component_allocation': float(allocation),
                                'subcomponents_total': float(subcomponents_total)
                            })
                        else:
                            return JsonResponse({
                                'is_valid': False,
                                'message': f"Subcomponents total exceeds component allocation by {difference}",
                                'severity': 'error',
                                'suggestions': "Reduce subcomponent allocations or increase component allocation",
                                'component_allocation': float(allocation),
                                'subcomponents_total': float(subcomponents_total)
                            })
                    
            except Component.DoesNotExist:
                # For new components, this is expected and valid
                pass
        
        # If we reach here, validation passed
        return JsonResponse({
            'is_valid': True,
            'message': 'Component allocation is valid',
            'severity': 'success',
            'component_allocation': float(allocation)
        })
            
    except Exception as e:
        return JsonResponse({
            'is_valid': False,
            'message': f'Validation error: {str(e)}',
            'severity': 'error'
        })

@require_http_methods(["POST"])
@login_required
def validate_subcomponent_allocation(request):
    """API endpoint to validate subcomponent allocation against activities total"""
    try:
        data = json.loads(request.body)
        subcomponent_id = data.get('subcomponent_id')
        allocation = Decimal(str(data.get('allocation', 0)))
        
        # If no subcomponent ID provided, skip validation
        if not subcomponent_id or subcomponent_id == 'null' or subcomponent_id == '':
            return JsonResponse({
                'is_valid': True,
                'message': 'Please select a subcomponent to validate allocation',
                'severity': 'info'
            })
        
        try:
            subcomponent = Subcomponent.objects.get(subcompID=subcomponent_id)
            activities_total = subcomponent.get_total_activities_allocation()
            
            is_valid = allocation == activities_total
            difference = abs(allocation - activities_total)
            
            response = {
                'is_valid': is_valid,
                'subcomponent_allocation': float(allocation),
                'activities_total': float(activities_total),
                'difference': float(difference),
                'currency': str(subcomponent.currency),
                'severity': 'error' if difference > 0 else 'success'
            }
            
            if is_valid:
                response['message'] = f"Subcomponent allocation matches activities total: {allocation} {subcomponent.currency}"
            else:
                if allocation > activities_total:
                    response['message'] = f"Subcomponent allocation exceeds activities total by {difference} {subcomponent.currency}"
                    response['suggestions'] = "Add more activities or reduce subcomponent allocation"
                else:
                    response['message'] = f"Activities total exceeds subcomponent allocation by {difference} {subcomponent.currency}"
                    response['suggestions'] = "Reduce activity allocations or increase subcomponent allocation"
                    
        except Subcomponent.DoesNotExist:
            response = {
                'is_valid': False,
                'message': 'Subcomponent not found',
                'severity': 'error'
            }
            
    except Exception as e:
        response = {
            'is_valid': False,
            'message': f'Validation error: {str(e)}',
            'severity': 'error'
        }
    
    return JsonResponse(response)

@require_http_methods(["POST"])
@login_required
def validate_activity_allocation(request):
    """API endpoint to validate activity allocation within subcomponent limits"""
    try:
        data = json.loads(request.body)
        subcomponent_id = data.get('subcomponent_id')
        activity_id = data.get('activity_id')
        allocation = Decimal(str(data.get('allocation', 0)))
        
        # If no subcomponent ID provided, skip validation
        if not subcomponent_id or subcomponent_id == 'null' or subcomponent_id == '':
            return JsonResponse({
                'is_valid': True,
                'message': 'Please select a subcomponent to validate activity allocation',
                'severity': 'info'
            })
        
        try:
            subcomponent = Subcomponent.objects.get(subcompID=subcomponent_id)
            
            # Calculate current activities total excluding this activity if editing
            activities_total = Decimal('0.00')
            if activity_id:
                activities_total = subcomponent.activities_set.exclude(activityID=activity_id).aggregate(
                    total=Sum('allocation'))['total'] or Decimal('0.00')
            else:
                activities_total = subcomponent.get_total_activities_allocation()
            
            new_total = activities_total + allocation
            remaining = subcomponent.allocation - new_total
            
            is_valid = new_total <= subcomponent.allocation
            
            response = {
                'is_valid': is_valid,
                'activity_allocation': float(allocation),
                'current_activities_total': float(activities_total),
                'new_total': float(new_total),
                'subcomponent_allocation': float(subcomponent.allocation),
                'remaining': float(remaining),
                'currency': str(subcomponent.currency),
                'severity': 'warning' if remaining < 0 else 'success'
            }
            
            if is_valid:
                response['message'] = f"Activity allocation is valid. Remaining: {remaining} {subcomponent.currency}"
            else:
                response['message'] = f"Activity allocation exceeds subcomponent limit by {abs(remaining)} {subcomponent.currency}"
                response['suggestions'] = f"Maximum allocation available: {subcomponent.allocation - activities_total} {subcomponent.currency}"
                    
        except Subcomponent.DoesNotExist:
            response = {
                'is_valid': False,
                'message': 'Subcomponent not found',
                'severity': 'error'
            }
            
    except Exception as e:
        response = {
            'is_valid': False,
            'message': f'Validation error: {str(e)}',
            'severity': 'error'
        }
    
    return JsonResponse(response)

@login_required
def edit_activity(request, activity_id):
    """Edit an activity"""
    activity = get_object_or_404(Activities, activityID=activity_id)
    
    if request.method == 'POST':
        # Using standard forms
        form = updateActivitiesForm(request.POST, instance=activity)
        if form.is_valid():
            activity = form.save(commit=False)
            activity.loginUser = request.user
            activity.save()
            messages.success(request, 'Activity updated successfully!')
            return redirect('PIU_Financial_mgt:activities')
    else:
        # Using standard forms
        form = updateActivitiesForm(instance=activity)
    
    # Get all projects, currencies, and years for the dropdowns
    projects = Project.objects.all()
    currencies = Currency.objects.all()
    years = YEAR.objects.all()
    
    # For edit mode, get related components and subcomponents
    components = Component.objects.filter(projectID=activity.projectID) if activity.projectID else Component.objects.none()
    subcomponents = Subcomponent.objects.filter(compID=activity.compID) if activity.compID else Subcomponent.objects.none()
    
    context = {
        'form': form,
        'activity': activity,
        'projects': projects,
        'components': components,
        'subcomponents': subcomponents,
        'currencies': currencies,
        'years': years,
        'is_edit': True,
    }
    return render(request, 'PIU_Financial_mgt/activity/add-activity.html', context)

@login_required
def delete_activity(request, activity_id):
    """Delete an activity"""
    activity = get_object_or_404(Activities, activityID=activity_id)
    
    if request.method == 'POST':
        activity.delete()
        messages.success(request, 'Activity deleted successfully!')
        return redirect('PIU_Financial_mgt:activities')
    
    context = {
        'activity': activity,
        'delete_url': 'PIU_Financial_mgt:delete_activity',
        'cancel_url': 'PIU_Financial_mgt:activities',
    }
    return render(request, 'PIU_Financial_mgt/activity/delete_activity.html', context)
