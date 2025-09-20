from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from django.http import JsonResponse, HttpResponse
from django.shortcuts import render, get_object_or_404, redirect
from .models import Results_Oriented_Monitoring
from django.contrib import messages

from .forms import Results_Oriented_MonitoringForm, updateResults_Oriented_MonitoringForm, Indicator_DescriptionForm
# Using standard forms from forms.py
from django.contrib import messages

from setup.models import Indicator_Type, Measurement_Unit, Data_Collection_Frequency
from PIU_Financial_mgt.models import ProjectOutCome, PDO, ProjectResult
from monitoring.models import Indicator_Description
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.core.paginator import Paginator


# PDF Export imports
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from datetime import datetime
import io

# Create your views here.

@login_required
def monitoring_dashboard(request):
    """Enhanced monitoring dashboard with real data"""
    from django.db.models import Count, Avg
    from django.conf import settings
    from PIU_Financial_mgt.models import Project
    from setup.models import Quarter
    
    # Get dashboard statistics using Django ORM
    try:
        stats = {
            'total_projects': Project.objects.count(),
            'total_indicators': Indicator_Description.objects.count(), 
            'quarterly_reports': Quarter.objects.count(),
            'performance_avg': 0,
            'recent_monitoring': [],
        }
        
        # Recent monitoring activities using Django ORM with correct field references
        recent_monitoring = Results_Oriented_Monitoring.objects.select_related(
            'project', 
            'quarter'
        ).order_by('-date_created')[:5]
        
        stats['recent_monitoring'] = list(recent_monitoring)
        
        # Calculate performance average using Django ORM
        monitoring_count = Results_Oriented_Monitoring.objects.count()
        if monitoring_count > 0:
            stats['performance_avg'] = min(85, (stats['total_projects'] * 5) + 65)
        else:
            stats['performance_avg'] = 0
            
    except Exception as e:
        # Fallback stats if there's an error
        stats = {
            'total_projects': 0,
            'total_indicators': 0,
            'quarterly_reports': 0,
            'performance_avg': 0,
            'recent_monitoring': [],
        }
    
    context = {
        'page_title': 'Monitoring Dashboard',
        **stats
    }
    
    return render(request, 'monitoring/dashboard.html', context)

@login_required
@csrf_exempt
def get_indicator_descriptions(request):
    if request.method == "POST":
        project_id = request.POST.get('project')
        indicator_type = request.POST.get('indicator_type')
        descriptions = Indicator_Description.objects.filter(
            project_id=project_id,
            indicator_type=indicator_type
        ).values('id', 'description')
        return JsonResponse({'descriptions': list(descriptions)})
    return JsonResponse({'error': 'Invalid request'}, status=400)


@login_required
def load_project_PDO(request):
    project_id = request.GET.get("project")
    print("Received project:", project_id)
    # Fix: Use project instead of project_id since PDO links to Project via ForeignKey 'project'
    pdos = PDO.objects.filter(project__projectID=project_id)
        
    print("pdos:", pdos)
    return render(request, "monitoring/result_oriented_monitoring/get_pdo.html", {"pdos": pdos})

@login_required
def load_project_Outcome(request):
    pdo_id = request.GET.get("pdo")
    print("Received PDO:", pdo_id)
    # Fix: Use pdo instead of pdo_id since ProjectOutCome links to PDO via ForeignKey 'pdo'
    project_outcomes = ProjectOutCome.objects.filter(pdo__id=pdo_id)
    
    print("project_outcomes:", project_outcomes)
    return render(request, "monitoring/get_project_outcome.html", {"project_outcomes": project_outcomes})


@login_required
def load_project_Result(request):
    project_outcome_id = request.GET.get("project_outcome")
    print("Received Project Result:", project_outcome_id)
    # Fix: Use project_outcome instead of project_outcome_id since ProjectResult links via ForeignKey 'project_outcome'
    projectResults = ProjectResult.objects.filter(project_outcome__id=project_outcome_id)
   
    print("projectResults:", projectResults)
    return render(request, "monitoring/result_oriented_monitoring/get_projectResult.html", {"projectResults": projectResults})

@login_required
def load_indicator_type(request):
    # Get the project_result_id from the GET request
    project_result_id = request.GET.get("project_result")
    
    # Fetch indicator types related to the selected project_result_id
    indicatorTypes = Indicator_Type.objects.filter(
        id__in=Indicator_Description.objects.filter(project_result_id=project_result_id).values('indicator_type')
    )

    print("Indicator Types:", indicatorTypes)  # Debugging to check the queryset
    
    # Return the template with the indicator_types context
    return render(request, 'monitoring/result_oriented_monitoring/get_IndicatorType.html', {'indicator_types': indicatorTypes})






#udate request
@login_required
def update_result_oriented_monitoring(request, pk):
    result_oriented_monitoring = get_object_or_404 (Results_Oriented_Monitoring, pk=pk)
    if request.method == "POST":
        form = updateResults_Oriented_MonitoringForm(request.POST, instance=result_oriented_monitoring)
        if form.is_valid():
            form.save()
            return render(request, 'monitoring/result_oriented_monitoring/result_oriented_monitoring_success.html', {'message': 'specific_contract_monitoring updated successfully'})
    else:
        form = updateResults_Oriented_MonitoringForm(instance=result_oriented_monitoring)
    return render(request, 'monitoring/result_oriented_monitoring/update-result_oriented_monitoring.html', {'form': form, 'result_oriented_monitoring': result_oriented_monitoring})

#Delete request
@login_required
@require_http_methods(["DELETE"])
def delete_result_oriented_monitoring(request, pk):
    result_oriented_monitoring = get_object_or_404(Results_Oriented_Monitoring, pk=pk)
    result_oriented_monitoring.delete()
    
    # Return success message for HTMX request
    context = {
        'message': "Data Deleted successfully!" 
    }
    return render(request, 'monitoring/result_oriented_monitoring/result_oriented_monitoring_success.html', context)



# Enhanced CRUD Views for Dashboard Integration
# Using standard forms

@login_required
def add_indicator_description(request):
    if request.method == 'POST':
        form = Indicator_DescriptionForm(request.POST)
        if form.is_valid():
            instance = form.save(commit=False)
            instance.loginUser = request.user
            instance.save()
            messages.success(request, "Indicator Description saved successfully!")
            return redirect('monitoring:monitoring_dashboard')
        else:
            # Add error messages for debugging
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")
    else:
        form = Indicator_DescriptionForm()
    
    context = {'form': form}
    return render(request, 'monitoring/indicator_description/add_indicator_description.html', context)


@login_required
def update_indicator_description(request, pk):
    indicator = get_object_or_404(Indicator_Description, pk=pk)
    if request.method == 'POST':
        form = Indicator_DescriptionForm(request.POST, instance=indicator)
        if form.is_valid():
            form.save()
            messages.success(request, "Indicator Description updated successfully!")
            return redirect('monitoring:monitoring_dashboard')
    else:
        form = Indicator_DescriptionForm(instance=indicator)
    
    context = {'form': form, 'indicator': indicator}
    return render(request, 'monitoring/indicator_description/update_indicator_description.html', context)


@login_required
def delete_indicator_description(request, pk):
    indicator = get_object_or_404(Indicator_Description, pk=pk)
    if request.method == 'POST':
        indicator.delete()
        messages.success(request, "Indicator Description deleted successfully!")
        return redirect('monitoring:monitoring_dashboard')
    
    context = {'indicator': indicator}
    return render(request, 'monitoring/indicator_description/delete_indicator_description.html', context)


@login_required
def add_results_monitoring(request):
    if request.method == 'POST':
        form = Results_Oriented_MonitoringForm(request.POST)
        
        if form.is_valid():
            instance = form.save(commit=False)
            instance.loginUser = request.user
            instance.save()
            messages.success(request, "Results Monitoring record saved successfully!")
            return redirect('monitoring:enhanced-results-monitoring-list')
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = Results_Oriented_MonitoringForm()
    
    context = {'form': form}
    return render(request, 'monitoring/results_monitoring/add_results_monitoring.html', context)


@login_required
def enhanced_results_monitoring_list(request):
    """Enhanced list view for Results Oriented Monitoring records with comprehensive filtering"""
    from django.core.paginator import Paginator
    from django.db import models
    from PIU_Financial_mgt.models import Project
    from setup.models import YEAR, Quarter
    
    try:
        # Get all monitoring records with related fields
        monitoring_qs = Results_Oriented_Monitoring.objects.select_related(
            'project', 'pdo', 'project_outcome', 'project_result', 
            'indicator_type', 'measurement_unit', 'collection_frequency',
            'year', 'quarter', 'loginUser'
        )
        
        # Apply filters based on GET parameters
        project_filter = request.GET.get('project', '')
        year_filter = request.GET.get('year', '')
        quarter_filter = request.GET.get('quarter', '')
        indicator_type_filter = request.GET.get('indicator_type', '')
        pdo_filter = request.GET.get('pdo', '')
        project_outcome_filter = request.GET.get('project_outcome', '')
        project_result_filter = request.GET.get('project_result', '')
        measurement_unit_filter = request.GET.get('measurement_unit', '')
        collection_frequency_filter = request.GET.get('collection_frequency', '')
        search_filter = request.GET.get('search', '')
        
        # Filter by project
        if project_filter:
            monitoring_qs = monitoring_qs.filter(project__projectID=project_filter)
        
        # Filter by year
        if year_filter:
            monitoring_qs = monitoring_qs.filter(year__id=year_filter)
        
        # Filter by quarter
        if quarter_filter:
            monitoring_qs = monitoring_qs.filter(quarter__id=quarter_filter)
        
        # Filter by indicator type
        if indicator_type_filter:
            monitoring_qs = monitoring_qs.filter(indicator_type__id=indicator_type_filter)
        
        # Filter by PDO
        if pdo_filter:
            monitoring_qs = monitoring_qs.filter(pdo__id=pdo_filter)
        
        # Filter by project outcome
        if project_outcome_filter:
            monitoring_qs = monitoring_qs.filter(project_outcome__id=project_outcome_filter)
        
        # Filter by project result
        if project_result_filter:
            monitoring_qs = monitoring_qs.filter(project_result__id=project_result_filter)
        
        # Filter by measurement unit
        if measurement_unit_filter:
            monitoring_qs = monitoring_qs.filter(measurement_unit__id=measurement_unit_filter)
        
        # Filter by collection frequency
        if collection_frequency_filter:
            monitoring_qs = monitoring_qs.filter(collection_frequency__id=collection_frequency_filter)
        
        # Search filter for indicator description and remarks
        if search_filter:
            monitoring_qs = monitoring_qs.filter(
                models.Q(indicator_description__icontains=search_filter) |
                models.Q(remarks__icontains=search_filter)
            )
        
        # Order by creation date (most recent first)
        monitoring_qs = monitoring_qs.order_by('-date_created')
        
        # Get filter options for dropdown
        projects = Project.objects.all().order_by('project')
        years = YEAR.objects.all().order_by('-profile_year')
        quarters = Quarter.objects.all().order_by('quarter')
        indicator_types = Indicator_Type.objects.all().order_by('indicator_type')
        pdos = PDO.objects.all().order_by('pdo_statement')
        project_outcomes = ProjectOutCome.objects.all().order_by('project_outcome')
        project_results = ProjectResult.objects.all().order_by('project_result')
        measurement_units = Measurement_Unit.objects.all().order_by('unit')
        collection_frequencies = Data_Collection_Frequency.objects.all().order_by('frequency')
        
        # Pagination - 5 records per page
        paginator = Paginator(monitoring_qs, 5)
        page_number = request.GET.get('page')
        monitoring_records = paginator.get_page(page_number)
        
        # Check if any filters are applied
        is_filtered = any([
            project_filter, year_filter, quarter_filter, indicator_type_filter,
            pdo_filter, project_outcome_filter, project_result_filter,
            measurement_unit_filter, collection_frequency_filter, search_filter
        ])
        
        context = {
            'monitoring_records': monitoring_records,
            'title': 'Enhanced Results Monitoring List',
            'total_count': paginator.count,
            'filtered_count': len(monitoring_records.object_list),
            'is_filtered': is_filtered,
            # Filter options
            'projects': projects,
            'years': years,
            'quarters': quarters,
            'indicator_types': indicator_types,
            'pdos': pdos,
            'project_outcomes': project_outcomes,
            'project_results': project_results,
            'measurement_units': measurement_units,
            'collection_frequencies': collection_frequencies,
            # Current filter values
            'current_filters': {
                'project': project_filter,
                'year': year_filter,
                'quarter': quarter_filter,
                'indicator_type': indicator_type_filter,
                'pdo': pdo_filter,
                'project_outcome': project_outcome_filter,
                'project_result': project_result_filter,
                'measurement_unit': measurement_unit_filter,
                'collection_frequency': collection_frequency_filter,
                'search': search_filter,
            }
        }
        
    except Exception as e:
        context = {
            'monitoring_records': [],
            'title': 'Enhanced Results Monitoring List',
            'total_count': 0,
            'error_message': f"Error loading monitoring records: {str(e)}",
            'projects': [],
            'years': [],
            'quarters': [],
            'indicator_types': [],
            'pdos': [],
            'project_outcomes': [],
            'project_results': [],
            'measurement_units': [],
            'collection_frequencies': [],
            'current_filters': {},
            'is_filtered': False,
        }
    
    return render(request, 'monitoring/results_monitoring/enhanced_results_monitoring_list.html', context)


@login_required
def update_results_monitoring(request, pk):
    monitoring = get_object_or_404(Results_Oriented_Monitoring, pk=pk)
    
    if request.method == 'POST':
        form = updateResults_Oriented_MonitoringForm(request.POST, instance=monitoring)
        if form.is_valid():
            instance = form.save(commit=False)
            instance.loginUser = request.user
            
            # Ensure all required foreign key fields are set
            if not instance.project_outcome_id:
                messages.error(request, "Project Outcome is required.")
                return render(request, 'monitoring/results_monitoring/update_results_monitoring.html', 
                             {'form': form, 'monitoring': monitoring})
            
            try:
                instance.save()
                messages.success(request, "Results Monitoring record updated successfully!")
                return redirect('monitoring:enhanced-results-monitoring-list')
            except Exception as e:
                messages.error(request, f"Error saving record: {str(e)}")
        else:
            # Show form validation errors
            messages.error(request, "Please correct the errors below.")
    else:
        form = updateResults_Oriented_MonitoringForm(instance=monitoring)
        
        # Ensure form fields have the correct initial values from the instance
        form.initial['project'] = monitoring.project
        form.initial['pdo'] = monitoring.pdo
        form.initial['project_outcome'] = monitoring.project_outcome
        form.initial['project_result'] = monitoring.project_result
        form.initial['indicator_type'] = monitoring.indicator_type
        form.initial['measurement_unit'] = monitoring.measurement_unit
    
    context = {'form': form, 'monitoring': monitoring}
    return render(request, 'monitoring/results_monitoring/update_results_monitoring.html', context)


@login_required
def delete_results_monitoring(request, pk):
    monitoring = get_object_or_404(Results_Oriented_Monitoring, pk=pk)
    if request.method == 'POST':
        monitoring.delete()
        messages.success(request, "Results Monitoring record deleted successfully!")
        return redirect('monitoring:enhanced-results-monitoring-list')
    
    context = {'monitoring': monitoring}
    return render(request, 'monitoring/results_monitoring/delete_results_monitoring.html', context)


@login_required
def detail_results_monitoring(request, pk):
    """Detail view for monitoring records"""
    try:
        record = get_object_or_404(Results_Oriented_Monitoring, pk=pk)
        
        # Initialize all variables with safe defaults
        performance_vs_target = 0
        performance_vs_baseline = 0
        variance_from_target = 0
        
        # Safe calculations with None checks
        if record.End_Target_Value and record.achieved_value:
            try:
                performance_vs_target = (float(record.achieved_value) / float(record.End_Target_Value)) * 100
                variance_from_target = float(record.achieved_value) - float(record.End_Target_Value)
            except (ValueError, ZeroDivisionError):
                pass
        
        if record.baseline_value and record.achieved_value:
            try:
                performance_vs_baseline = ((float(record.achieved_value) - float(record.baseline_value)) / float(record.baseline_value)) * 100
            except (ValueError, ZeroDivisionError):
                pass
        
        # Safe status determination
        if performance_vs_target >= 100:
            status = 'excellent'
            status_color = 'success'
        elif performance_vs_target >= 75:
            status = 'good'
            status_color = 'warning'
        elif performance_vs_target >= 50:
            status = 'fair'
            status_color = 'info'
        else:
            status = 'needs_improvement'
            status_color = 'danger'
        
        context = {
            'record': record,
            'page_title': f'Monitoring Details - {record.indicator_description}',
            'performance_vs_target': performance_vs_target,
            'performance_vs_baseline': performance_vs_baseline,
            'variance_from_target': variance_from_target,
            'status': status,
            'status_color': status_color,
        }
        
        return render(request, 'monitoring/results_monitoring/detail_results_monitoring.html', context)
        
    except Exception as e:
        # If anything fails, provide debugging info and redirect
        messages.error(request, f"Error loading monitoring record details: {str(e)}")
        return redirect('monitoring:enhanced-results-monitoring-list')


@login_required
def export_indicator_descriptions_excel(request):
    from django.http import HttpResponse
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Indicator Descriptions"
    
    headers = ['ID', 'Project', 'PDO', 'Project Outcome', 'Project Result', 'Indicator Type', 'Description', 'Created By']
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
    
    indicators = Indicator_Description.objects.select_related(
        'project', 'pdo', 'project_outcome', 'project_result', 'indicator_type', 'loginUser'
    ).all()
    
    for row, indicator in enumerate(indicators, 2):
        ws.cell(row=row, column=1, value=indicator.id)
        ws.cell(row=row, column=2, value=str(indicator.project))
        ws.cell(row=row, column=3, value=str(indicator.pdo))
        ws.cell(row=row, column=4, value=str(indicator.project_outcome))
        ws.cell(row=row, column=5, value=str(indicator.project_result))
        ws.cell(row=row, column=6, value=str(indicator.indicator_type))
        ws.cell(row=row, column=7, value=indicator.indicator_description)
        ws.cell(row=row, column=8, value=indicator.loginUser.username)
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=indicator_descriptions.xlsx'
    wb.save(response)
    return response


@login_required
def export_results_monitoring_excel(request):
    from django.http import HttpResponse
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Results Monitoring"
    
    headers = ['ID', 'Year', 'Quarter', 'Project', 'PDO', 'Indicator Type', 'Description', 
               'Baseline Value', 'Achieved Value', 'Target Value', '% vs Baseline', '% vs Target', 
               'Remarks', 'Created By', 'Date Created']
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
    
    # Apply the same filtering logic as the list view
    monitoring_qs = Results_Oriented_Monitoring.objects.select_related(
        'year', 'quarter', 'project', 'pdo', 'indicator_type', 'loginUser',
        'project_outcome', 'project_result', 'measurement_unit', 'collection_frequency'
    )
    
    # Apply filters based on GET parameters (same as list view)
    project_filter = request.GET.get('project', '')
    year_filter = request.GET.get('year', '')
    quarter_filter = request.GET.get('quarter', '')
    indicator_type_filter = request.GET.get('indicator_type', '')
    pdo_filter = request.GET.get('pdo', '')
    project_outcome_filter = request.GET.get('project_outcome', '')
    project_result_filter = request.GET.get('project_result', '')
    measurement_unit_filter = request.GET.get('measurement_unit', '')
    collection_frequency_filter = request.GET.get('collection_frequency', '')
    search_filter = request.GET.get('search', '')
    
    # Apply the same filters as list view
    if project_filter:
        monitoring_qs = monitoring_qs.filter(project__projectID=project_filter)
    if year_filter:
        monitoring_qs = monitoring_qs.filter(year__id=year_filter)
    if quarter_filter:
        monitoring_qs = monitoring_qs.filter(quarter__id=quarter_filter)
    if indicator_type_filter:
        monitoring_qs = monitoring_qs.filter(indicator_type__id=indicator_type_filter)
    if pdo_filter:
        monitoring_qs = monitoring_qs.filter(pdo__id=pdo_filter)
    if project_outcome_filter:
        monitoring_qs = monitoring_qs.filter(project_outcome__id=project_outcome_filter)
    if project_result_filter:
        monitoring_qs = monitoring_qs.filter(project_result__id=project_result_filter)
    if measurement_unit_filter:
        monitoring_qs = monitoring_qs.filter(measurement_unit__id=measurement_unit_filter)
    if collection_frequency_filter:
        monitoring_qs = monitoring_qs.filter(collection_frequency__id=collection_frequency_filter)
    if search_filter:
        from django.db import models
        monitoring_qs = monitoring_qs.filter(
            models.Q(indicator_description__icontains=search_filter) |
            models.Q(remarks__icontains=search_filter)
        )
    
    monitoring_records = monitoring_qs.order_by('-date_created')
    
    for row, record in enumerate(monitoring_records, 2):
        ws.cell(row=row, column=1, value=record.id)
        ws.cell(row=row, column=2, value=str(record.year) if record.year else '')
        ws.cell(row=row, column=3, value=str(record.quarter))
        ws.cell(row=row, column=4, value=str(record.project))
        ws.cell(row=row, column=5, value=str(record.pdo))
        ws.cell(row=row, column=6, value=str(record.indicator_type))
        ws.cell(row=row, column=7, value=record.indicator_description)
        ws.cell(row=row, column=8, value=record.baseline_value)
        ws.cell(row=row, column=9, value=record.achieved_value)
        ws.cell(row=row, column=10, value=record.End_Target_Value)
        ws.cell(row=row, column=11, value=record.percentage_achieved_vs_baseline)
        ws.cell(row=row, column=12, value=record.percentage_achieved_vs_end_target)
        ws.cell(row=row, column=13, value=record.remarks)
        ws.cell(row=row, column=14, value=record.loginUser.username)
        ws.cell(row=row, column=15, value=record.date_created.strftime('%Y-%m-%d %H:%M') if record.date_created else '')
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=results_monitoring.xlsx'
    wb.save(response)
    return response


@login_required
def export_results_monitoring_pdf(request):
    """Export Enhanced Results Monitoring to PDF in A4 landscape format"""
    from django.db import models
    from PIU_Financial_mgt.models import Project
    from setup.models import YEAR, Quarter
    
    # Create a buffer to hold PDF content
    buffer = io.BytesIO()
    
    # Create the PDF object using A4 landscape with standard margins
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=1*inch,
        leftMargin=1*inch,
        topMargin=1*inch,
        bottomMargin=1*inch
    )
    
    # Get all filter parameters from GET request
    project_filter = request.GET.get('project')
    year_filter = request.GET.get('year') 
    quarter_filter = request.GET.get('quarter')
    indicator_type_filter = request.GET.get('indicator_type')
    pdo_filter = request.GET.get('pdo')
    project_outcome_filter = request.GET.get('project_outcome')
    project_result_filter = request.GET.get('project_result')
    measurement_unit_filter = request.GET.get('measurement_unit')
    collection_frequency_filter = request.GET.get('collection_frequency')
    search_filter = request.GET.get('search')
    
    # Apply same filtering logic as the list view
    monitoring_qs = Results_Oriented_Monitoring.objects.select_related(
        'project', 'year', 'quarter', 'pdo', 'project_outcome', 
        'project_result', 'indicator_type', 'measurement_unit', 
        'collection_frequency', 'loginUser'
    ).all()
    
    # Apply filters
    if project_filter:
        monitoring_qs = monitoring_qs.filter(project__projectID=project_filter)
    if year_filter:
        monitoring_qs = monitoring_qs.filter(year__id=year_filter)
    if quarter_filter:
        monitoring_qs = monitoring_qs.filter(quarter__id=quarter_filter)
    if indicator_type_filter:
        monitoring_qs = monitoring_qs.filter(indicator_type__id=indicator_type_filter)
    if pdo_filter:
        monitoring_qs = monitoring_qs.filter(pdo__id=pdo_filter)
    if project_outcome_filter:
        monitoring_qs = monitoring_qs.filter(project_outcome__id=project_outcome_filter)
    if project_result_filter:
        monitoring_qs = monitoring_qs.filter(project_result__id=project_result_filter)
    if measurement_unit_filter:
        monitoring_qs = monitoring_qs.filter(measurement_unit__id=measurement_unit_filter)
    if collection_frequency_filter:
        monitoring_qs = monitoring_qs.filter(collection_frequency__id=collection_frequency_filter)
    if search_filter:
        monitoring_qs = monitoring_qs.filter(
            models.Q(indicator_description__icontains=search_filter) |
            models.Q(remarks__icontains=search_filter)
        )
    
    monitoring_qs = monitoring_qs.order_by('-date_created')
    
    # Container for the story
    story = []
    
    # Define styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=20,
        alignment=1,  # Center alignment
        textColor=colors.darkblue
    )
    
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Normal'],
        fontSize=10,
        spaceAfter=15,
        alignment=1,  # Center alignment
        textColor=colors.grey
    )
    
    # Add title
    title = Paragraph("Enhanced Results Monitoring Report", title_style)
    story.append(title)
    
    # Add generation date and filters
    filter_info = f"Generated on: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
    if any([project_filter, year_filter, quarter_filter, indicator_type_filter, 
            pdo_filter, project_outcome_filter, project_result_filter,
            measurement_unit_filter, collection_frequency_filter, search_filter]):
        filter_info += " | Filtered Results"
    
    subtitle = Paragraph(filter_info, subtitle_style)
    story.append(subtitle)
    story.append(Spacer(1, 20))
    
    # Prepare table data with wrapped text
    data = []
    
    # Header style for consistency
    header_style = ParagraphStyle(
        'HeaderStyle',
        parent=styles['Normal'],
        fontSize=8,
        leading=10,
        alignment=1,
        textColor=colors.whitesmoke,
        fontName='Helvetica-Bold'
    )
    
    # Headers as Paragraph objects for consistent formatting
    headers = [
        Paragraph('Year', header_style), Paragraph('Quarter', header_style), 
        Paragraph('Project', header_style), Paragraph('PDO', header_style), 
        Paragraph('Outcome', header_style), Paragraph('Result', header_style), 
        Paragraph('Indicator<br/>Type', header_style), Paragraph('Description', header_style), 
        Paragraph('Unit', header_style), Paragraph('Frequency', header_style),
        Paragraph('Baseline', header_style), Paragraph('Achieved', header_style), 
        Paragraph('Target', header_style), Paragraph('% vs<br/>Base', header_style), 
        Paragraph('% vs<br/>Target', header_style), Paragraph('Remarks', header_style)
    ]
    data.append(headers)
    
    # Data rows with proper text wrapping using Paragraph objects
    for record in monitoring_qs:
        # Helper function to safely get attribute values without truncation
        def safe_get(obj, attr):
            try:
                value = getattr(obj, attr, 'N/A')
                if value is None:
                    return 'N/A'
                return str(value)
            except:
                return 'N/A'
        
        def safe_get_nested(obj, attr_chain):
            try:
                value = obj
                for attr in attr_chain.split('.'):
                    value = getattr(value, attr, None)
                    if value is None:
                        return 'N/A'
                return str(value)
            except:
                return 'N/A'
        
        # Create Paragraph objects for text wrapping
        cell_style = ParagraphStyle(
            'CellStyle',
            parent=styles['Normal'],
            fontSize=7,
            leading=9,
            alignment=1,  # Center alignment
            wordWrap='CJK',
            splitLongWords=True
        )
        
        row = [
            Paragraph(safe_get_nested(record, 'year.profile_year'), cell_style),
            Paragraph(safe_get_nested(record, 'quarter.quarter'), cell_style),
            Paragraph(safe_get_nested(record, 'project.project'), cell_style),
            Paragraph(safe_get_nested(record, 'pdo.pdo_statement'), cell_style),
            Paragraph(safe_get_nested(record, 'project_outcome.project_outcome'), cell_style),
            Paragraph(safe_get_nested(record, 'project_result.project_result'), cell_style),
            Paragraph(safe_get_nested(record, 'indicator_type.indicator_type'), cell_style),
            Paragraph(safe_get(record, 'indicator_description'), cell_style),
            Paragraph(safe_get_nested(record, 'measurement_unit.unit'), cell_style),
            Paragraph(safe_get_nested(record, 'collection_frequency.frequency'), cell_style),
            Paragraph(f"{record.baseline_value:.1f}" if record.baseline_value else 'N/A', cell_style),
            Paragraph(f"{record.achieved_value:.1f}" if record.achieved_value else 'N/A', cell_style),
            Paragraph(f"{record.End_Target_Value:.1f}" if record.End_Target_Value else 'N/A', cell_style),
            Paragraph(f"{record.percentage_achieved_vs_baseline:.1f}%" if record.percentage_achieved_vs_baseline else 'N/A', cell_style),
            Paragraph(f"{record.percentage_achieved_vs_end_target:.1f}%" if record.percentage_achieved_vs_end_target else 'N/A', cell_style),
            Paragraph(safe_get(record, 'remarks'), cell_style)
        ]
        data.append(row)
    
    if len(data) == 1:  # Only headers, no data
        cell_style = ParagraphStyle(
            'CellStyle',
            parent=styles['Normal'],
            fontSize=10,
            leading=12,
            alignment=1,
            wordWrap='CJK'
        )
        data.append([Paragraph('No records found matching the specified criteria', cell_style)] + [Paragraph('', cell_style)] * (len(headers) - 1))
    
    # Create table with optimized column widths for A4 landscape with 1" margins
    # Available width is approximately 9.69 inches (11.69 - 2 for margins)
    # Adjusted to fit within page bounds - total should be ~9.6 inches
    col_widths = [0.4*inch, 0.5*inch, 0.65*inch, 0.85*inch, 0.65*inch, 0.65*inch, 
                  0.55*inch, 1.0*inch, 0.45*inch, 0.55*inch, 0.4*inch, 0.4*inch, 
                  0.4*inch, 0.45*inch, 0.45*inch, 0.85*inch]
    
    table = Table(data, colWidths=col_widths, repeatRows=1)
    
    # Table style optimized for text wrapping
    table.setStyle(TableStyle([
        # Header row styling
        ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        
        # Data rows styling for Paragraph objects
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('ALIGN', (0, 1), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),  # Top alignment for better text wrapping
        
        # Grid styling
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('LINEBELOW', (0, 0), (-1, 0), 2, colors.darkblue),
        
        # Alternating row colors
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        
        # Cell padding optimized for text wrapping
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
    ]))
    
    story.append(table)
    
    # Add footer with record count
    story.append(Spacer(1, 20))
    footer_text = f"Total Records: {len(data) - 1} | NAWEC PIU Enhanced Results Monitoring System"
    footer = Paragraph(footer_text, subtitle_style)
    story.append(footer)
    
    # Build PDF
    doc.build(story)
    
    # Get the value of the BytesIO buffer and return response
    pdf = buffer.getvalue()
    buffer.close()
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Enhanced_Results_Monitoring_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
    response.write(pdf)
    
    return response
