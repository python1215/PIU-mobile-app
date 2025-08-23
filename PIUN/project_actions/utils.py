import os
from datetime import datetime, timedelta
from django.http import HttpResponse
from django.utils import timezone
from django.db.models import Sum, Count, Avg, Q
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas
from io import BytesIO

from .models import Contract_Profiling_works, Contract_Profiling_goods_services, Specific_Contract_Monitoring


def export_works_contracts_to_excel(queryset=None):
    """Export Contract Profiling Works to Excel with advanced formatting"""
    
    if queryset is None:
        queryset = Contract_Profiling_works.objects.all().select_related(
            'projectID', 'compID', 'subcompID', 'activityID', 'project_Category',
            'funding_source', 'currency', 'loginUser'
        )
    
    # Create workbook and worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Works Contracts"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    data_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    number_alignment = Alignment(horizontal="right", vertical="center")
    date_alignment = Alignment(horizontal="center", vertical="center")
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = [
        'Contract Ref No', 'Project', 'Component', 'Subcomponent', 'Activity',
        'Category', 'Funding Source', 'Currency', 'Contract Value',
        'Contractor', 'Consultant', 'Start Date', 'End Date', 'Duration',
        'Location', 'Latitude', 'Longitude', 'Floor Area (m²)',
        'Beneficiaries', 'Intervention Focus', 'Amendments', 'Remarks',
        'Created By', 'Created Date'
    ]
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Write data
    for row_num, contract in enumerate(queryset, 2):
        data = [
            contract.contract_refNo,
            str(contract.projectID) if contract.projectID else '',
            str(contract.compID) if contract.compID else '',
            str(contract.subcompID) if contract.subcompID else '',
            str(contract.activityID) if contract.activityID else '',
            str(contract.project_Category) if contract.project_Category else '',
            str(contract.funding_source) if contract.funding_source else '',
            str(contract.currency) if contract.currency else '',
            float(contract.contract_value) if contract.contract_value else 0,
            contract.name_of_contractor or '',
            contract.name_of_consultant or '',
            contract.contract_start_date,
            contract.contract_end_date,
            contract.duration or '',
            contract.location_of_investment or '',
            float(contract.Latitude) if contract.Latitude else '',
            float(contract.Longitude) if contract.Longitude else '',
            contract.gross_floor_area_m2 or '',
            contract.target_number_of_beneficiary_settlements or '',
            contract.main_intervention_focus_result or '',
            'Yes' if contract.amendments else 'No',
            contract.remarks or '',
            str(contract.loginUser) if contract.loginUser else '',
            contract.date.strftime('%Y-%m-%d %H:%M') if contract.date else ''
        ]
        
        for col_num, value in enumerate(data, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = border
            
            # Apply specific formatting based on data type
            if col_num == 9:  # Contract Value
                cell.alignment = number_alignment
                cell.number_format = '#,##0.00'
            elif col_num in [12, 13]:  # Dates
                cell.alignment = date_alignment
                if value:
                    cell.number_format = 'YYYY-MM-DD'
            elif col_num in [16, 17]:  # Coordinates
                cell.alignment = number_alignment
                cell.number_format = '0.000000'
            else:
                cell.alignment = data_alignment
    
    # Auto-adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # Freeze header row
    worksheet.freeze_panes = 'A2'
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="works_contracts_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    workbook.save(response)
    return response


def export_goods_services_contracts_to_excel(queryset=None):
    """Export Contract Profiling Goods & Services to Excel"""
    
    if queryset is None:
        queryset = Contract_Profiling_goods_services.objects.all().select_related(
            'projectID', 'compID', 'subcompID', 'activityID', 'project_Category',
            'funding_source', 'currency', 'loginUser'
        )
    
    # Create workbook and worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Goods & Services Contracts"
    
    # Define styles (same as works contracts)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="28a745", end_color="28a745", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    data_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    number_alignment = Alignment(horizontal="right", vertical="center")
    date_alignment = Alignment(horizontal="center", vertical="center")
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = [
        'Contract Ref No', 'Project', 'Component', 'Subcomponent', 'Activity',
        'Category', 'Funding Source', 'Currency', 'Contract Value',
        'Supplier', 'Consultant', 'Start Date', 'End Date', 'Duration',
        'Amendments', 'Remarks', 'Created By', 'Created Date'
    ]
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Write data
    for row_num, contract in enumerate(queryset, 2):
        data = [
            contract.contract_refNo,
            str(contract.projectID) if contract.projectID else '',
            str(contract.compID) if contract.compID else '',
            str(contract.subcompID) if contract.subcompID else '',
            str(contract.activityID) if contract.activityID else '',
            str(contract.project_Category) if contract.project_Category else '',
            str(contract.funding_source) if contract.funding_source else '',
            str(contract.currency) if contract.currency else '',
            float(contract.contract_value) if contract.contract_value else 0,
            contract.name_of_Supplier or '',
            contract.name_of_consultant or '',
            contract.contract_start_date,
            contract.contract_end_date,
            contract.duration or '',
            'Yes' if contract.amendments else 'No',
            contract.remarks or '',
            str(contract.loginUser) if contract.loginUser else '',
            contract.date.strftime('%Y-%m-%d %H:%M') if contract.date else ''
        ]
        
        for col_num, value in enumerate(data, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = border
            
            # Apply specific formatting based on data type
            if col_num == 9:  # Contract Value
                cell.alignment = number_alignment
                cell.number_format = '#,##0.00'
            elif col_num in [12, 13]:  # Dates
                cell.alignment = date_alignment
                if value:
                    cell.number_format = 'YYYY-MM-DD'
            else:
                cell.alignment = data_alignment
    
    # Auto-adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # Freeze header row
    worksheet.freeze_panes = 'A2'
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="goods_services_contracts_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    workbook.save(response)
    return response


def export_monitoring_records_to_excel(queryset=None):
    """Export Contract Monitoring Records to Excel"""
    
    if queryset is None:
        queryset = Specific_Contract_Monitoring.objects.all().select_related(
            'project', 'quarter', 'type_of_monitoring', 'Type_of_Investment',
            'Kpi_description', 'Contract_implementation_Status', 'loginUser'
        )
    
    # Create workbook and worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Monitoring Records"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="17a2b8", end_color="17a2b8", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    data_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    date_alignment = Alignment(horizontal="center", vertical="center")
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = [
        'Contract Ref No', 'Project', 'Monitoring Date', 'Quarter',
        'Monitoring Type', 'Investment Type', 'KPI Description',
        'Milestone Start', 'Milestone End', 'Target', 'Achieved Status',
        'Implementation Status', 'Remarks', 'Created By', 'Created Date'
    ]
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Write data
    for row_num, record in enumerate(queryset, 2):
        data = [
            record.contract_refNo,
            str(record.project) if record.project else '',
            record.monitoring_date,
            str(record.quarter) if record.quarter else '',
            str(record.type_of_monitoring) if record.type_of_monitoring else '',
            str(record.Type_of_Investment) if record.Type_of_Investment else '',
            str(record.Kpi_description) if record.Kpi_description else '',
            record.milestone_start_date,
            record.milestone_end_date,
            record.Target or '',
            record.Achieved_status or '',
            str(record.Contract_implementation_Status) if record.Contract_implementation_Status else '',
            record.remarks or '',
            str(record.loginUser) if record.loginUser else '',
            record.date.strftime('%Y-%m-%d %H:%M') if record.date else ''
        ]
        
        for col_num, value in enumerate(data, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = border
            
            # Apply specific formatting based on data type
            if col_num in [3, 8, 9]:  # Dates
                cell.alignment = date_alignment
                if value:
                    cell.number_format = 'YYYY-MM-DD'
            else:
                cell.alignment = data_alignment
    
    # Auto-adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # Freeze header row
    worksheet.freeze_panes = 'A2'
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="monitoring_records_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    workbook.save(response)
    return response


def get_dashboard_analytics():
    """Get comprehensive analytics for dashboard"""
    
    today = timezone.now().date()
    
    analytics = {
        'works_contracts': {
            'total': 0,
            'active': 0,
            'completed': 0,
            'upcoming': 0,
            'total_value': 0,
            'with_amendments': 0,
        },
        'goods_services_contracts': {
            'total': 0,
            'active': 0,
            'completed': 0,
            'upcoming': 0,
            'total_value': 0,
            'with_amendments': 0,
        },
        'monitoring': {
            'total_records': 0,
            'unique_contracts': 0,
            'overdue_milestones': 0,
            'active_milestones': 0,
        },
        'recent_activity': {
            'works_contracts': [],
            'goods_services_contracts': [],
            'monitoring_records': [],
        }
    }
    
    try:
        # Works contracts analytics
        works_contracts = Contract_Profiling_works.objects.all()
        analytics['works_contracts'].update({
            'total': works_contracts.count(),
            'active': works_contracts.filter(
                contract_start_date__lte=today,
                contract_end_date__gte=today
            ).count(),
            'completed': works_contracts.filter(contract_end_date__lt=today).count(),
            'upcoming': works_contracts.filter(contract_start_date__gt=today).count(),
            'total_value': works_contracts.aggregate(
                total=Sum('contract_value')
            )['total'] or 0,
            'with_amendments': works_contracts.filter(amendments=True).count(),
        })
        
        # Goods & Services contracts analytics
        gs_contracts = Contract_Profiling_goods_services.objects.all()
        analytics['goods_services_contracts'].update({
            'total': gs_contracts.count(),
            'active': gs_contracts.filter(
                contract_start_date__lte=today,
                contract_end_date__gte=today
            ).count(),
            'completed': gs_contracts.filter(contract_end_date__lt=today).count(),
            'upcoming': gs_contracts.filter(contract_start_date__gt=today).count(),
            'total_value': gs_contracts.aggregate(
                total=Sum('contract_value')
            )['total'] or 0,
            'with_amendments': gs_contracts.filter(amendments=True).count(),
        })
        
        # Monitoring analytics
        monitoring_records = Specific_Contract_Monitoring.objects.all()
        analytics['monitoring'].update({
            'total_records': monitoring_records.count(),
            'unique_contracts': monitoring_records.values('contract_refNo').distinct().count(),
            'overdue_milestones': monitoring_records.filter(
                milestone_end_date__lt=today
            ).count(),
            'active_milestones': monitoring_records.filter(
                milestone_start_date__lte=today,
                milestone_end_date__gte=today
            ).count(),
        })
        
        # Recent activity
        analytics['recent_activity'].update({
            'works_contracts': works_contracts.order_by('-date')[:5],
            'goods_services_contracts': gs_contracts.order_by('-date')[:5],
            'monitoring_records': monitoring_records.order_by('-date')[:5],
        })
        
    except Exception as e:
        print(f"Error in dashboard analytics: {str(e)}")
    
    return analytics


def export_monitoring_records_to_pdf(queryset=None):
    """Export Contract Monitoring Records to PDF with A4 portrait formatting"""
    
    if queryset is None:
        queryset = Specific_Contract_Monitoring.objects.all().select_related(
            'project', 'quarter', 'type_of_monitoring', 'Type_of_Investment',
            'Kpi_description', 'Contract_implementation_Status', 'loginUser'
        )
    
    # Create PDF buffer
    buffer = BytesIO()
    
    # Create PDF document with A4 portrait
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=0.5*inch,
        leftMargin=0.5*inch,
        topMargin=1*inch,
        bottomMargin=0.75*inch
    )
    
    # Get styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=30,
        textColor=colors.black,
        alignment=1  # Center alignment
    )
    
    # Build PDF content
    story = []
    
    # Title
    title = Paragraph("Contract Monitoring Records Report", title_style)
    story.append(title)
    
    # Date and summary
    date_text = f"Generated on: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
    summary_text = f"Total Records: {queryset.count()}"
    
    story.append(Paragraph(date_text, styles['Normal']))
    story.append(Paragraph(summary_text, styles['Normal']))
    story.append(Spacer(1, 20))
    
    # Prepare table data
    data = []
    headers = [
        'Contract Ref', 'Project', 'Monitoring Type', 'Monitoring Date', 
        'Quarter', 'Investment Type', 'KPI Description', 'Target', 
        'Achievement', 'Status', 'Remarks'
    ]
    data.append(headers)
    
    # Add data rows with Paragraph objects for text wrapping
    for record in queryset:
        row = [
            Paragraph(record.contract_refNo or '', styles['Normal']),
            Paragraph(str(record.project) if record.project else '', styles['Normal']),
            Paragraph(str(record.type_of_monitoring) if record.type_of_monitoring else '', styles['Normal']),
            Paragraph(record.monitoring_date.strftime('%Y-%m-%d') if record.monitoring_date else '', styles['Normal']),
            Paragraph(str(record.quarter) if record.quarter else '', styles['Normal']),
            Paragraph(str(record.Type_of_Investment) if record.Type_of_Investment else '', styles['Normal']),
            Paragraph(str(record.Kpi_description) if record.Kpi_description else '', styles['Normal']),
            Paragraph(str(record.Target) if record.Target else '', styles['Normal']),
            Paragraph(str(record.Achieved_status) if record.Achieved_status else '', styles['Normal']),
            Paragraph(str(record.Contract_implementation_Status) if record.Contract_implementation_Status else '', styles['Normal']),
            Paragraph(str(record.remarks) if record.remarks else '', styles['Normal'])
        ]
        data.append(row)
    
    # Create table with proper column widths for A4
    col_widths = [0.8*inch, 1.2*inch, 0.7*inch, 0.8*inch, 0.6*inch, 1.0*inch, 1.2*inch, 0.7*inch, 0.7*inch, 0.8*inch, 1.2*inch]
    
    table = Table(data, colWidths=col_widths, repeatRows=1)
    
    # Style the table
    table.setStyle(TableStyle([
        # Header style
        ('BACKGROUND', (0, 0), (-1, 0), colors.navy),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        
        # Data style
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('LEFTPADDING', (0, 0), (-1, -1), 3),
        ('RIGHTPADDING', (0, 0), (-1, -1), 3),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ('WORDWRAP', (0, 0), (-1, -1), 'CJK'),
    ]))
    
    story.append(table)
    
    # Build PDF
    doc.build(story)
    
    # Get PDF data
    pdf_data = buffer.getvalue()
    buffer.close()
    
    # Create response
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="monitoring_records_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
    response.write(pdf_data)
    
    return response


def export_works_contracts_to_pdf(queryset=None):
    """Export Works Contracts to PDF with A4 portrait formatting"""
    
    if queryset is None:
        queryset = Contract_Profiling_works.objects.all().select_related(
            'projectID', 'compID', 'subcompID', 'activityID', 'project_Category',
            'funding_source', 'currency', 'loginUser'
        )
    
    # Create PDF buffer
    buffer = BytesIO()
    
    # Create PDF document with A4 portrait
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=0.5*inch,
        leftMargin=0.5*inch,
        topMargin=1*inch,
        bottomMargin=0.75*inch
    )
    
    # Get styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=30,
        textColor=colors.black,
        alignment=1  # Center alignment
    )
    
    # Build PDF content
    story = []
    
    # Title
    title = Paragraph("Works Contracts Report", title_style)
    story.append(title)
    
    # Date and summary
    date_text = f"Generated on: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
    summary_text = f"Total Contracts: {queryset.count()}"
    
    story.append(Paragraph(date_text, styles['Normal']))
    story.append(Paragraph(summary_text, styles['Normal']))
    story.append(Spacer(1, 20))
    
    # Prepare table data
    data = []
    headers = [
        'Contract Ref', 'Project', 'Component', 'Contractor', 
        'Contract Value', 'Currency', 'Start Date', 'End Date', 
        'Duration (Days)', 'Location', 'Status'
    ]
    data.append(headers)
    
    # Add data rows with Paragraph objects for text wrapping
    for contract in queryset:
        # Calculate duration
        duration = ''
        if contract.contract_start_date and contract.contract_end_date:
            duration = str((contract.contract_end_date - contract.contract_start_date).days)
        
        # Status determination
        status = 'Inactive'
        if contract.contract_start_date and contract.contract_end_date:
            if contract.contract_start_date <= datetime.now().date() <= contract.contract_end_date:
                status = 'Active'
        
        row = [
            Paragraph(contract.contract_refNo or '', styles['Normal']),
            Paragraph(str(contract.projectID) if contract.projectID else '', styles['Normal']),
            Paragraph(str(contract.compID) if contract.compID else '', styles['Normal']),
            Paragraph(contract.name_of_contractor if contract.name_of_contractor else '', styles['Normal']),
            Paragraph(f"{contract.contract_value:,.2f}" if contract.contract_value else '', styles['Normal']),
            Paragraph(str(contract.currency) if contract.currency else '', styles['Normal']),
            Paragraph(contract.contract_start_date.strftime('%Y-%m-%d') if contract.contract_start_date else '', styles['Normal']),
            Paragraph(contract.contract_end_date.strftime('%Y-%m-%d') if contract.contract_end_date else '', styles['Normal']),
            Paragraph(duration, styles['Normal']),
            Paragraph(str(contract.location_of_investment) if hasattr(contract, 'location_of_investment') and contract.location_of_investment else '', styles['Normal']),
            Paragraph(status, styles['Normal'])
        ]
        data.append(row)
    
    # Create table with proper column widths for A4
    col_widths = [1.0*inch, 1.3*inch, 1.0*inch, 1.2*inch, 0.9*inch, 0.6*inch, 0.8*inch, 0.8*inch, 0.7*inch, 1.0*inch, 0.7*inch]
    
    table = Table(data, colWidths=col_widths, repeatRows=1)
    
    # Style the table
    table.setStyle(TableStyle([
        # Header style
        ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        
        # Data style
        ('BACKGROUND', (0, 1), (-1, -1), colors.lightgrey),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('LEFTPADDING', (0, 0), (-1, -1), 3),
        ('RIGHTPADDING', (0, 0), (-1, -1), 3),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ('WORDWRAP', (0, 0), (-1, -1), 'CJK'),
    ]))
    
    story.append(table)
    
    # Build PDF
    doc.build(story)
    
    # Get PDF data
    pdf_data = buffer.getvalue()
    buffer.close()
    
    # Create response
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="works_contracts_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
    response.write(pdf_data)
    
    return response


def export_goods_services_contracts_to_pdf(queryset=None):
    """Export Goods & Services Contracts to PDF with A4 portrait formatting"""
    
    if queryset is None:
        queryset = Contract_Profiling_goods_services.objects.all().select_related(
            'projectID', 'compID', 'subcompID', 'activityID', 'project_Category',
            'funding_source', 'currency', 'loginUser'
        )
    
    # Create PDF buffer
    buffer = BytesIO()
    
    # Create PDF document with A4 portrait
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=0.5*inch,
        leftMargin=0.5*inch,
        topMargin=1*inch,
        bottomMargin=0.75*inch
    )
    
    # Get styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=30,
        textColor=colors.black,
        alignment=1  # Center alignment
    )
    
    # Build PDF content
    story = []
    
    # Title
    title = Paragraph("Goods & Services Contracts Report", title_style)
    story.append(title)
    
    # Date and summary
    date_text = f"Generated on: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
    summary_text = f"Total Contracts: {queryset.count()}"
    
    story.append(Paragraph(date_text, styles['Normal']))
    story.append(Paragraph(summary_text, styles['Normal']))
    story.append(Spacer(1, 20))
    
    # Prepare table data
    data = []
    headers = [
        'Contract Ref', 'Project', 'Component', 'Supplier', 
        'Contract Value', 'Currency', 'Start Date', 'End Date', 
        'Duration (Days)', 'Description', 'Status'
    ]
    data.append(headers)
    
    # Add data rows with Paragraph objects for text wrapping
    for contract in queryset:
        # Calculate duration
        duration = ''
        if contract.contract_start_date and contract.contract_end_date:
            duration = str((contract.contract_end_date - contract.contract_start_date).days)
        
        # Status determination
        status = 'Inactive'
        if contract.contract_start_date and contract.contract_end_date:
            if contract.contract_start_date <= datetime.now().date() <= contract.contract_end_date:
                status = 'Active'
        
        row = [
            Paragraph(contract.contract_refNo or '', styles['Normal']),
            Paragraph(str(contract.projectID) if contract.projectID else '', styles['Normal']),
            Paragraph(str(contract.compID) if contract.compID else '', styles['Normal']),
            Paragraph(contract.name_of_Supplier if contract.name_of_Supplier else '', styles['Normal']),
            Paragraph(f"{contract.contract_value:,.2f}" if contract.contract_value else '', styles['Normal']),
            Paragraph(str(contract.currency) if contract.currency else '', styles['Normal']),
            Paragraph(contract.contract_start_date.strftime('%Y-%m-%d') if contract.contract_start_date else '', styles['Normal']),
            Paragraph(contract.contract_end_date.strftime('%Y-%m-%d') if contract.contract_end_date else '', styles['Normal']),
            Paragraph(duration, styles['Normal']),
            Paragraph(str(contract.remarks) if hasattr(contract, 'remarks') and contract.remarks else '', styles['Normal']),
            Paragraph(status, styles['Normal'])
        ]
        data.append(row)
    
    # Create table with proper column widths for A4
    col_widths = [1.0*inch, 1.3*inch, 1.0*inch, 1.2*inch, 0.9*inch, 0.6*inch, 0.8*inch, 0.8*inch, 0.7*inch, 1.2*inch, 0.7*inch]
    
    table = Table(data, colWidths=col_widths, repeatRows=1)
    
    # Style the table
    table.setStyle(TableStyle([
        # Header style
        ('BACKGROUND', (0, 0), (-1, 0), colors.darkgreen),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        
        # Data style
        ('BACKGROUND', (0, 1), (-1, -1), colors.lightgreen),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('LEFTPADDING', (0, 0), (-1, -1), 3),
        ('RIGHTPADDING', (0, 0), (-1, -1), 3),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ('WORDWRAP', (0, 0), (-1, -1), 'CJK'),
    ]))
    
    story.append(table)
    
    # Build PDF
    doc.build(story)
    
    # Get PDF data
    pdf_data = buffer.getvalue()
    buffer.close()
    
    # Create response
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="goods_services_contracts_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
    response.write(pdf_data)
    
    return response


def calculate_contract_duration_days(start_date, end_date):
    """Calculate contract duration in days"""
    if start_date and end_date:
        return (end_date - start_date).days
    return 0


def get_contract_status(start_date, end_date):
    """Get contract status based on dates"""
    today = timezone.now().date()
    
    if not start_date or not end_date:
        return 'unknown'
    
    if start_date > today:
        return 'upcoming'
    elif end_date < today:
        return 'completed'
    else:
        return 'active'


def get_milestone_status(start_date, end_date, monitoring_date=None):
    """Get milestone status"""
    today = timezone.now().date()
    
    if not start_date or not end_date:
        return 'unknown'
    
    if start_date > today:
        return 'upcoming'
    elif end_date < today:
        if monitoring_date and monitoring_date >= end_date:
            return 'completed'
        else:
            return 'overdue'
    else:
        return 'active'
