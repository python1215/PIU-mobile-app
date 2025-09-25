from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse, HttpResponse, Http404
from django.contrib import messages
from django.conf import settings
from django.core.paginator import Paginator
from django.db import models
from django.db.models import Q, Count, Sum, Avg
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.utils.decorators import method_decorator
from django.views.decorators.cache import never_cache
from django.utils import timezone
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from .models import ESIA, PAP, GrievianceMonitoringLog, OHS_Monitoring, CommunityConsult_Engagement
from .forms import (ESIAForm, ESIAUpdateForm, PAPForm, PAPUpdateForm,
                    GrievianceMonitoringLogForm, GrievianceUpdateForm,
                    OHSMonitoringForm, OHSUpdateForm, CommunityEngagementForm)
from .filters import (ESIAFilter, PAPFilter, GrievianceMonitoringLogFilter,
                      OHSMonitoringFilter, CommunityEngagementFilter)
from setup.models import Districts, Settlement, Regions, Quarter, YEAR
from PIU_Financial_mgt.models import KPI_For_Contract, Project
from PIU_Financial_mgt.models import ProjectOutCome, PDO, ProjectResult
from monitoring.models import Indicator_Description
# Using Django ORM exclusively for all database operations


# ======================== ESIA Views ========================
@login_required
def esia_list(request):
    """Enhanced ESIA/ESMP list view with filtering and pagination"""
    # Get all ESIA records with related data
    esia_list = ESIA.objects.select_related('project_name',
                                            'type_of_investment',
                                            'loginUser').all()

    # Apply Django Filter
    esia_filter = ESIAFilter(request.GET, queryset=esia_list)
    filtered_esia = esia_filter.qs

    # Pagination with page size support
    page_size = request.GET.get('page_size', 10)
    try:
        page_size = int(page_size)
        if page_size not in [10, 25, 50, 100]:
            page_size = 10
    except (ValueError, TypeError):
        page_size = 10

    paginator = Paginator(filtered_esia, page_size)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Statistics
    stats = {
        'total_esia': esia_list.count(),
        'filtered_count': filtered_esia.count(),
        'avg_duration': esia_list.aggregate(Avg('project_duration'))['project_duration__avg'] or 0,
        'total_communities': esia_list.aggregate(Sum('number_of_communities'))['number_of_communities__sum'] or 0,
    }

    context = {
        'page_obj': page_obj,
        'filter': esia_filter,
        'stats': stats,
        'is_filtered': bool(request.GET),
        'title': 'ESIA/ESMP Records'
    }

    return render(request, 'social_and_env/esia/esia_list.html', context)


@login_required
def esia_detail(request, pk):
    """ESIA detail view"""
    esia = get_object_or_404(ESIA.objects.select_related(
        'project_name', 'type_of_investment', 'loginUser'),
                             pk=pk)

    context = {'esia': esia}
    return render(request, 'social_and_env/esia/esia_detail.html', context)


@login_required
def esia_add(request):
    """Add new ESIA record"""
    if request.method == 'POST':
        form = ESIAForm(request.POST)
        if form.is_valid():
            try:
                esia = form.save(commit=False)
                esia.loginUser = request.user
                esia.save()
                messages.success(request,
                                 'ESIA/ESMP record created successfully!')
                return redirect('esia_list')
            except Exception as e:
                messages.error(request,
                               f'Error saving ESIA/ESMP record: {str(e)}')
        else:
            # Print form errors for debugging
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        form = ESIAForm()

    context = {'form': form, 'title': 'Add ESIA/ESMP Record'}
    return render(request, 'social_and_env/esia/esia_form.html', context)


@login_required
def esia_edit(request, pk):
    """Edit ESIA/ESMP record"""
    esia = get_object_or_404(ESIA, pk=pk)

    if request.method == 'POST':
        form = ESIAForm(request.POST, instance=esia)
        if form.is_valid():
            try:
                esia = form.save(commit=False)
                esia.loginUser = request.user
                esia.save()
                messages.success(request,
                                 'ESIA/ESMP record updated successfully!')
                return redirect('esia_list')
            except Exception as e:
                messages.error(request,
                               f'Error updating ESIA/ESMP record: {str(e)}')
        else:
            # Print form errors for debugging
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        form = ESIAForm(instance=esia)
        # Ensure the investment type is properly populated for editing
        if esia.type_of_investment:
            investment_code = esia.type_of_investment.monitoring_Type_Code
            form.fields['type_of_investment'].initial = investment_code
            
            # Also set the widget value to ensure it displays correctly
            from django.forms.widgets import Select
            if hasattr(form.fields['type_of_investment'].widget, 'choices'):
                current_choices = list(form.fields['type_of_investment'].widget.choices)
                # Ensure current value is in choices
                current_in_choices = any(choice[0] == investment_code for choice in current_choices)
                if not current_in_choices:
                    current_name = str(esia.type_of_investment.type_of_investment) if hasattr(esia.type_of_investment, 'type_of_investment') else str(esia.type_of_investment)
                    current_choices.append((investment_code, current_name))
                    form.fields['type_of_investment'].widget.choices = current_choices

    context = {'form': form, 'esia': esia, 'title': 'Edit ESIA/ESMP Record'}
    return render(request, 'social_and_env/esia/esia_edit.html', context)


@login_required
@require_http_methods(["DELETE"])
def esia_delete(request, pk):
    """Delete ESIA/ESMP record"""
    esia = get_object_or_404(ESIA, pk=pk)
    esia.delete()
    messages.success(request, 'ESIA/ESMP record deleted successfully!')
    return JsonResponse({'success': True})


def esia_export_excel(request):
    """Export ESIA data to Excel"""
    from django.http import HttpResponse
    from django.shortcuts import redirect
    from django.contrib import messages
    from datetime import datetime
    import logging
    
    logger = logging.getLogger(__name__)
    logger.info("ESIA Excel export initiated")
    
    try:
        import openpyxl
        logger.info("openpyxl imported successfully")
    except ImportError as e:
        logger.error(f"openpyxl import failed: {e}")
        messages.error(request, 'Excel export functionality is not available. openpyxl is missing.')
        return redirect('esia_list')
    
    try:
        # Get ESIA data
        esia_data = ESIA.objects.select_related('project_name', 'type_of_investment', 'loginUser').all()
        logger.info(f"Retrieved {esia_data.count()} ESIA records")
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "ESIA Records"
        logger.info("Workbook created successfully")
        
        # Headers
        headers = ['ESIA ID', 'Project', 'Investment Type', 'Duration', 'Phase', 'Locations', 'Communities', 'Findings', 'Date', 'User']
        
        # Write headers
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Write data
        for row, esia in enumerate(esia_data, 2):
            try:
                ws.cell(row=row, column=1, value=esia.esiaID)
                ws.cell(row=row, column=2, value=str(esia.project_name.project) if esia.project_name else 'N/A')
                ws.cell(row=row, column=3, value=str(esia.type_of_investment.type_of_investment) if esia.type_of_investment else 'N/A')
                ws.cell(row=row, column=4, value=esia.project_duration)
                ws.cell(row=row, column=5, value=esia.project_phase)
                ws.cell(row=row, column=6, value=esia.project_locations)
                ws.cell(row=row, column=7, value=esia.number_of_communities)
                ws.cell(row=row, column=8, value=str(esia.esia_findings)[:500] if esia.esia_findings else '')
                ws.cell(row=row, column=9, value=esia.date_created.strftime('%Y-%m-%d') if esia.date_created else '')
                ws.cell(row=row, column=10, value=str(esia.loginUser.username) if esia.loginUser else '')
            except Exception as row_error:
                logger.error(f"Error processing row {row}: {row_error}")
                continue
        
        logger.info("Data written to workbook successfully")
        
        # Prepare response
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f'ESIA_Records_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Save to response using BytesIO to ensure proper content
        from io import BytesIO
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        response.write(buffer.getvalue())
        buffer.close()
        
        logger.info("Excel file saved to response successfully")
        return response
        
    except Exception as e:
        logger.error(f"ESIA Excel export failed: {str(e)}", exc_info=True)
        messages.error(request, f'Export failed: {str(e)}')
        return redirect('esia_list')


# ======================== PAP Views ========================


@login_required
def pap_list(request):
    """Enhanced PAP list view with filtering and pagination"""

    # Check total records in database
    total_db_records = PAP.objects.count()
    print(f"🔍 Total PAP records in database: {total_db_records}")

    # Initial queryset
    pap_list = PAP.objects.select_related(
        'project', 'type_of_investment', 'region', 'district',
        'pap_Current_Address', 'type_of_pap', 'pap_category',
        'vulnerability_category', 'type_of_impact', 'nature_of_compensation',
        'loginUser').all()

    print(f"🔍 Records after select_related: {pap_list.count()}")

    # Check for any filters being applied
    print(f"🔍 GET parameters: {request.GET}")

    # Apply filters
    pap_filter = PAPFilter(request.GET, queryset=pap_list)
    filtered_pap = pap_filter.qs

    print(f"🔍 Records after filtering: {filtered_pap.count()}")

    # Check if any records have broken relationships
    broken_records = []
    for pap in PAP.objects.all()[:5]:  # Check first 5 records
        try:
            # Try to access all foreign keys
            _ = pap.project.project
            _ = pap.type_of_investment.type_of_investment
            _ = pap.region.region_name
            _ = pap.district.district_name
            _ = pap.pap_Current_Address.settlement_name
            _ = pap.type_of_pap.type_of_pap
            _ = pap.pap_category.pap_category
            _ = pap.vulnerability_category.vulnerability
            _ = pap.type_of_impact.impact
            _ = pap.nature_of_compensation.nature_of_settlement
            _ = pap.loginUser.username
        except Exception as e:
            broken_records.append(f"PAP {pap.pap_identification_number}: {e}")

    if broken_records:
        print(f"🚨 Records with broken relationships: {broken_records}")

    # Pagination
    paginator = Paginator(filtered_pap, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    print(f"🔍 Records on current page: {len(page_obj)}")
    print(f"🔍 Total pages: {paginator.num_pages}")
    print(f"🔍 Current page number: {page_obj.number}")

    # Statistics
    stats = {
        'total_pap':
        pap_list.count(),
        'filtered_count':
        filtered_pap.count(),
        'compensated':
        filtered_pap.filter(pap_compensated='Y').count(),
        'not_compensated':
        filtered_pap.filter(pap_compensated='N').count(),
        'total_compensation':
        filtered_pap.aggregate(Sum('amount'))['amount__sum'] or 0,
        'male_count':
        filtered_pap.filter(sex='M').count(),
        'female_count':
        filtered_pap.filter(sex='F').count(),
    }

    context = {
        'page_obj': page_obj,
        'filter': pap_filter,
        'stats': stats,
        'is_filtered': bool(request.GET),
        'debug_info': {  # Add debug info to template
            'total_db_records': total_db_records,
            'after_select_related': pap_list.count(),
            'after_filtering': filtered_pap.count(),
            'current_page_count': len(page_obj),
            'total_pages': paginator.num_pages,
        }
    }

    return render(request, 'social_and_env/pap/pap_list.html', context)


@login_required
def pap_detail(request, pk):
    """PAP detail view"""
    import logging
    logger = logging.getLogger(__name__)
    logger.info(f"🔍 PAP Detail view called with pk: {pk}")
    
    try:
        pap = get_object_or_404(PAP.objects.select_related(
            'project', 'type_of_investment', 'region', 'district',
            'pap_Current_Address', 'type_of_pap', 'pap_category',
            'vulnerability_category', 'type_of_impact', 'nature_of_compensation', 'loginUser'),
                                pk=pk)

        logger.info(f"🔍 Found PAP: {pap.pap_name} (ID: {pap.pk})")
        
        # Create a documents context - check if document_upload field exists
        documents = []
        if hasattr(pap, 'document_upload') and pap.document_upload:
            documents = [{
                'id': 1,
                'document_name': pap.document_upload.name.split('/')[-1],
                'upload_date': pap.date_created,
                'uploaded_by': pap.loginUser,
                'file_url': pap.document_upload.url
            }]
        
        logger.info(f"🔍 Found {len(documents)} documents for PAP")
        
        context = {
            'pap': pap,
            'title': f'PAP Details - {pap.pap_name}',
            'documents': documents,
        }

        logger.info(f"🔍 Rendering pap_detail.html with context")
        return render(request, 'social_and_env/pap/pap_detail.html', context)

    except Exception as e:
        logger.error(f"🔍 Error in pap_detail view: {str(e)}")
        messages.error(request, f'Error loading PAP details: {str(e)}')
        return redirect('pap_list')


@login_required
def export_pap_excel(request):
    """Export filtered PAP data to Excel format"""
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from datetime import datetime
    import logging

    logger = logging.getLogger(__name__)
    logger.info("🔍 Starting PAP Excel export with filters")

    try:
        # Get PAP data using Django ORM with same filtering logic as pap_list view
        initial_queryset = PAP.objects.select_related(
            'project', 'type_of_investment', 'region', 'district',
            'pap_Current_Address', 'type_of_pap', 'pap_category',
            'vulnerability_category', 'type_of_impact', 'nature_of_compensation',
            'loginUser').all()

        # Apply filters using PAPFilter - same as in pap_list view
        pap_filter = PAPFilter(request.GET, queryset=initial_queryset)
        pap_list = pap_filter.qs
        
        logger.info(f"🔍 Found {pap_list.count()} filtered PAP records for export")
        
        # Log applied filters for debugging
        applied_filters = {k: v for k, v in request.GET.items() if v}
        if applied_filters:
            logger.info(f"🔍 Applied filters: {applied_filters}")
        else:
            logger.info("🔍 No filters applied - exporting all records")

        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = 'PAP Records'

        # Define headers
        headers = [
            'PAP ID', 'PAP Name', 'Gender', 'Project', 'Region', 'District',
            'Settlement', 'Location of Impact', 'Amount', 'Area', 'Compensated',
            'Compensation Date', 'Compensation Ref No', 'Type of PAP',
            'PAP Category', 'Vulnerability Category', 'Type of Impact',
            'Type of Investment', 'Nature of Compensation', 'Pre-Project Situation', 'Remarks',
            'Date Created', 'User'
        ]

        # Write headers
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        # Write data
        row_num = 2
        for pap in pap_list:
            try:
                ws.cell(row=row_num, column=1, value=pap.pap_identification_number or '')
                ws.cell(row=row_num, column=2, value=pap.pap_name or '')
                ws.cell(row=row_num, column=3, value=pap.sex or '')
                ws.cell(row=row_num, column=4, value=pap.project.project if pap.project else '')
                ws.cell(row=row_num, column=5, value=pap.region.region_name if pap.region else '')
                ws.cell(row=row_num, column=6, value=pap.district.district_name if pap.district else '')
                ws.cell(row=row_num, column=7, value=pap.pap_Current_Address.settlement_name if pap.pap_Current_Address else '')
                ws.cell(row=row_num, column=8, value=pap.location_of_impact or '')
                ws.cell(row=row_num, column=9, value=pap.amount or 0)
                ws.cell(row=row_num, column=10, value=pap.area or '')
                ws.cell(row=row_num, column=11, value=pap.pap_compensated or '')
                ws.cell(row=row_num, column=12, value=pap.compensation_date.strftime('%Y-%m-%d') if pap.compensation_date else '')
                ws.cell(row=row_num, column=13, value=pap.compensation_RefNo or '')
                ws.cell(row=row_num, column=14, value=pap.type_of_pap.type_of_pap if pap.type_of_pap else '')
                ws.cell(row=row_num, column=15, value=pap.pap_category.pap_category if pap.pap_category else '')
                ws.cell(row=row_num, column=16, value=pap.vulnerability_category.vulnerability if pap.vulnerability_category else '')
                ws.cell(row=row_num, column=17, value=pap.type_of_impact.impact if pap.type_of_impact else '')
                ws.cell(row=row_num, column=18, value=pap.type_of_investment.type_of_investment if pap.type_of_investment else '')
                ws.cell(row=row_num, column=19, value=pap.nature_of_compensation.nature_of_settlement if pap.nature_of_compensation else '')
                ws.cell(row=row_num, column=20, value=pap.pre_project_situation or '')
                ws.cell(row=row_num, column=21, value=pap.remarks or '')
                ws.cell(row=row_num, column=22, value=pap.date_created.strftime('%Y-%m-%d %H:%M') if pap.date_created else '')
                ws.cell(row=row_num, column=23, value=pap.loginUser.username if pap.loginUser else '')
                row_num += 1
            except Exception as field_error:
                logger.error(f"🔍 Error processing PAP {pap.pap_identification_number} at row {row_num}: {field_error}")
                # Continue with next record instead of failing completely
                row_num += 1
                continue

        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15

        # Create filename with filter indication
        timestamp = datetime.now().strftime("%Y%m%d_%H%M")
        if applied_filters:
            filename = f'pap_records_filtered_{timestamp}.xlsx'
        else:
            filename = f'pap_records_all_{timestamp}.xlsx'

        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename={filename}'

        wb.save(response)
        logger.info(f"🔍 PAP Excel export completed successfully: {filename}")
        return response

    except Exception as e:
        logger.error(f"🔍 Error exporting PAP data: {str(e)}")
        messages.error(request, f'Error exporting PAP data: {str(e)}')
        return redirect('pap_list')


@login_required
def pap_add(request):
    """Add new PAP record"""
    if request.method == 'POST':
        form = PAPForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                pap = form.save(commit=False)
                pap.loginUser = request.user

                # Handle the excluded CharField fields manually by getting raw POST data
                type_of_investment_value = request.POST.get(
                    'type_of_investment')
                if type_of_investment_value:
                    try:
                        kpi = KPI_For_Contract.objects.get(
                            monitoring_Type_Code=type_of_investment_value)
                        pap.type_of_investment = kpi
                    except KPI_For_Contract.DoesNotExist:
                        # Use first available for this project as fallback
                        pap.type_of_investment = KPI_For_Contract.objects.filter(
                            project=pap.project).first()

                district_value = request.POST.get('district')
                if district_value:
                    try:
                        district = Districts.objects.get(
                            district_code=district_value)
                        pap.district = district
                    except Districts.DoesNotExist:
                        # Use first district in region as fallback
                        if pap.region:
                            pap.district = Districts.objects.filter(
                                region_code=pap.region).first()

                settlement_value = request.POST.get('pap_Current_Address')
                if settlement_value:
                    try:
                        settlement = Settlement.objects.get(
                            settlement_code=settlement_value)
                        pap.pap_Current_Address = settlement
                    except Settlement.DoesNotExist:
                        # Use first settlement in district as fallback
                        if pap.district:
                            pap.pap_Current_Address = Settlement.objects.filter(
                                district_code=pap.district).first()

                # Ensure pap_Current_Address is not None (required field)
                if not pap.pap_Current_Address:
                    if pap.district:
                        pap.pap_Current_Address = Settlement.objects.filter(
                            district_code=pap.district).first()
                    elif pap.region:
                        # Find any settlement in the region
                        first_district = Districts.objects.filter(
                            region_code=pap.region).first()
                        if first_district:
                            pap.pap_Current_Address = Settlement.objects.filter(
                                district_code=first_district).first()

                    # Last resort - use any available settlement
                    if not pap.pap_Current_Address:
                        pap.pap_Current_Address = Settlement.objects.first()

                # Ensure required lookup fields have values
                pap.type_of_pap = pap.type_of_pap or (
                    TypeOfPAP.objects.first()
                    if TypeOfPAP.objects.exists() else None)
                pap.pap_category = pap.pap_category or (
                    PAPCategory.objects.first()
                    if PAPCategory.objects.exists() else None)
                pap.vulnerability_category = pap.vulnerability_category or (
                    VulnerabilityCategory.objects.first()
                    if VulnerabilityCategory.objects.exists() else None)
                pap.type_of_impact = pap.type_of_impact or (
                    TypeOfImpact.objects.first()
                    if TypeOfImpact.objects.exists() else None)
                pap.nature_of_compensation = pap.nature_of_compensation or (
                    NatureOfSettlement.objects.first()
                    if NatureOfSettlement.objects.exists() else None)

                # Default text/numeric fields
                pap.area = pap.area or '0'
                pap.pap_compensated = pap.pap_compensated or 'N'
                pap.pre_project_situation = pap.pre_project_situation or 'Information not provided'

                pap.save()
                
                # Document upload functionality removed
                
                messages.success(
                    request,
                    f'PAP record {pap.pap_identification_number} added successfully.'
                )
                return redirect('pap_list')

            except Exception as e:
                import traceback
                messages.error(request, f'Error saving PAP record: {str(e)}')
                print(traceback.format_exc())
        else:
            print("Form errors:", form.errors)
            messages.error(request, f'Please correct the errors below.')
    else:
        form = PAPForm()

    return render(request, 'social_and_env/pap/pap_form.html', {
        'form': form,
        'title': 'Add PAP Record'
    })


@login_required
def pap_edit(request, pk):
    """Edit PAP record with dedicated edit template"""
    try:
        pap = get_object_or_404(PAP.objects.select_related(
            'project', 'type_of_investment', 'region', 'district',
            'pap_Current_Address', 'type_of_pap', 'pap_category',
            'vulnerability_category', 'type_of_impact',
            'nature_of_compensation', 'loginUser'),
                                pk=pk)

        if request.method == 'POST':
            form = PAPUpdateForm(request.POST, request.FILES, instance=pap)
            if form.is_valid():
                try:
                    updated_pap = form.save(commit=False)
                    updated_pap.loginUser = request.user
                    # Note: date_modified field doesn't exist in PAP model, removed to avoid error
                    updated_pap.save()
                    messages.success(
                        request,
                        f'PAP record "{pap.pap_name}" updated successfully.')
                    return redirect('pap_list')
                except Exception as e:
                    messages.error(request,
                                   f'Error updating PAP record: {str(e)}')
            else:
                # Display form validation errors
                for field, errors in form.errors.items():
                    for error in errors:
                        messages.error(request, f'{field}: {error}')
        else:
            form = PAPUpdateForm(instance=pap)

        context = {
            'form': form,
            'pap': pap,
            'title': f'Edit PAP Record - {pap.pap_name}',
        }

        return render(request, 'social_and_env/pap/pap_edit.html', context)

    except Exception as e:
        messages.error(request,
                       f'Error loading PAP record for editing: {str(e)}')
        return redirect('pap_list')


@login_required
def pap_delete(request, pk):
    """Delete PAP record"""
    pap = get_object_or_404(PAP, pk=pk)

    if request.method == 'POST':
        try:
            pap.delete()
            messages.success(request, 'PAP record deleted successfully.')
            return redirect('pap_list')
        except Exception as e:
            messages.error(request, f'Error deleting PAP record: {str(e)}')

    return render(request, 'social_and_env/pap/pap_confirm_delete.html', {
        'pap': pap,
        'title': 'Delete PAP Record'
    })


@login_required
def social_env_dashboard(request):
    """Dashboard with overview statistics"""
    from django.db.models import Count, Sum

    try:
        # Get statistics using Django ORM
        stats = {
            'total_pap':
            PAP.objects.count(),
            'total_grievances':
            GrievianceMonitoringLog.objects.count(),
            'total_esia':
            ESIA.objects.count(),
            'total_ohs':
            OHS_Monitoring.objects.count(),
            'total_community':
            CommunityConsult_Engagement.objects.count(),
            'compensated_pap':
            PAP.objects.filter(pap_compensated='Y').count(),
            'male_pap':
            PAP.objects.filter(sex='M').count(),
            'female_pap':
            PAP.objects.filter(sex='F').count(),
            'open_grievances':
            GrievianceMonitoringLog.objects.filter(
                was_complainant_satisfied_with_decision='N').count(),
            'closed_grievances':
            GrievianceMonitoringLog.objects.filter(
                was_complainant_satisfied_with_decision='Y').count(),
        }

        # Get recent activities - convert to lists for template evaluation
        recent_pap = list(
            PAP.objects.select_related(
                'project', 'loginUser').order_by('-date_created')[:5])
        recent_grievances = list(
            GrievianceMonitoringLog.objects.select_related(
                'project', 'loginUser').order_by('-date_created')[:5])
        recent_esia = list(
            ESIA.objects.select_related(
                'project', 'loginUser').order_by('-date_created')[:5])

        context = {
            'stats': stats,
            'recent_pap': recent_pap,
            'recent_grievances': recent_grievances,
            'recent_esia': recent_esia,
            'title': 'Social & Environmental Dashboard'
        }

        return render(request, 'social_and_env/dashboard.html', context)

    except Exception as e:
        messages.error(request, f'Error loading dashboard: {str(e)}')
        return render(
            request, 'social_and_env/dashboard.html', {
                'stats': {},
                'recent_pap': [],
                'recent_grievances': [],
                'recent_esia': [],
                'title': 'Social & Environmental Dashboard'
            })


# ======================== Grievance Views ========================
@login_required
def grievance_list(request):
    """Enhanced Grievance list view with filtering and pagination"""
    from django.core.paginator import Paginator

    try:
        # Initialize the filter
        grievance_filter = GrievianceMonitoringLogFilter(
            request.GET,
            queryset=GrievianceMonitoringLog.objects.select_related(
                'project', 'type_of_investment', 'decision_outcome',
                'loginUser').order_by('-date_claim_recieved'))

        # Get filtered queryset
        grievances = grievance_filter.qs

        # Search functionality (additional to filters)
        search_query = request.GET.get('search', '')
        if search_query:
            grievances = grievances.filter(
                Q(case_no__icontains=search_query)
                | Q(name_of_complainant__icontains=search_query)
                | Q(complaint_content__icontains=search_query)
                | Q(project__project__icontains=search_query))

        # Pagination
        page_size = request.GET.get('page_size', 15)
        try:
            page_size = int(page_size)
            if page_size not in [10, 15, 25, 50, 100]:
                page_size = 15
        except (ValueError, TypeError):
            page_size = 15

        paginator = Paginator(grievances, page_size)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        # Get statistics from filtered queryset
        total_cases = grievances.count()
        satisfied = grievances.filter(
            was_complainant_satisfied_with_decision='Y').count()
        not_satisfied = grievances.filter(
            was_complainant_satisfied_with_decision='N').count()
        pending = grievances.filter(
            was_complainant_satisfied_with_decision__isnull=True).count()

        stats = {
            'total_cases': total_cases,
            'satisfied': satisfied,
            'not_satisfied': not_satisfied,
            'pending': pending,
            'male_complainants': grievances.filter(sex='M').count(),
            'female_complainants': grievances.filter(sex='F').count(),
        }

        context = {
            'page_obj': page_obj,
            'filter': grievance_filter,
            'stats': stats,
            'search_query': search_query,
            'is_filtered': bool(request.GET),
            'title': 'Grievance Management'
        }

        return render(request, 'social_and_env/grievance/grievance_list.html',
                      context)

    except Exception as e:
        messages.error(request, f'Error loading grievances: {str(e)}')
        return render(
            request, 'social_and_env/grievance/grievance_list.html', {
                'page_obj': None,
                'filter': None,
                'stats': {},
                'search_query': '',
                'is_filtered': False,
                'title': 'Grievance Management'
            })


@login_required
def grievance_detail(request, pk):
    """Grievance detail view"""
    try:
        grievance = get_object_or_404(
            GrievianceMonitoringLog.objects.select_related(
                'project', 'type_of_investment', 'decision_outcome',
                'loginUser'),
            pk=pk)

        context = {
            'grievance': grievance,
            'title': f'Grievance Details - {grievance.case_no}',
        }

        return render(request,
                      'social_and_env/grievance/grievance_detail.html',
                      context)

    except Exception as e:
        messages.error(request, f'Error loading grievance details: {str(e)}')
        return redirect('grievance_list')


@login_required
def grievance_add(request):
    """Add new Grievance record"""
    if request.method == 'POST':
        print("=== GRIEVANCE ADD DEBUG START ===")
        print(f"POST data: {dict(request.POST)}")

        form = GrievianceMonitoringLogForm(request.POST)
        print(f"Form created, is_valid: {form.is_valid()}")

        if not form.is_valid():
            print(f"Form errors: {form.errors}")
            print(f"Non-field errors: {form.non_field_errors()}")

            # Form is not valid, show errors
            error_messages = []
            for field, errors in form.errors.items():
                for error in errors:
                    error_messages.append(f"{field}: {error}")
                    print(f"Field error - {field}: {error}")

            if form.non_field_errors():
                for error in form.non_field_errors():
                    error_messages.append(f"General: {error}")
                    print(f"Non-field error: {error}")

            if error_messages:
                messages.error(
                    request,
                    f'Form validation failed: {"; ".join(error_messages)}')
            print("=== FORM VALIDATION FAILED ===")
        else:
            print("Form is valid, attempting to save...")
            try:
                grievance = form.save(commit=False)
                grievance.loginUser = request.user

                # Debug: Print submitted data
                project_value = request.POST.get('project')
                investment_value = request.POST.get('type_of_investment')
                print(
                    f"Form data: project={project_value}, type_of_investment={investment_value}"
                )
                print(
                    f"Grievance object before save: project={grievance.project}, investment={grievance.type_of_investment}"
                )

                # Validate that type_of_investment is properly set
                if not grievance.type_of_investment and investment_value:
                    # Try to find the investment type by monitoring_Type_Code
                    try:
                        investment_obj = KPI_For_Contract.objects.get(
                            monitoring_Type_Code=investment_value)
                        grievance.type_of_investment = investment_obj
                        print(
                            f"Found investment type by code: {investment_obj}")
                    except KPI_For_Contract.DoesNotExist:
                        print(
                            f"No investment type found for code: {investment_value}"
                        )

                # Final check - ensure investment type is set
                if not grievance.type_of_investment:
                    # Try to get a default investment type for the project
                    if grievance.project:
                        default_investment = KPI_For_Contract.objects.filter(
                            project=grievance.project,
                            monitoring_type_id='ESS').first()
                        if default_investment:
                            grievance.type_of_investment = default_investment
                            print(
                                f"Auto-assigned investment type: {default_investment}"
                            )
                        else:
                            messages.error(
                                request,
                                'No investment type found for the selected project.'
                            )
                            print("=== NO INVESTMENT TYPE FOUND ===")
                            return render(
                                request,
                                'social_and_env/grievance/grievance_add_form.html',
                                {
                                    'form': form,
                                    'title': 'Add New Grievance Record'
                                })

                print(f"Final grievance object: {grievance.__dict__}")
                grievance.save()
                print("Grievance saved successfully!")
                messages.success(request,
                                 'Grievance record added successfully.')
                return redirect('grievance_list')
            except Exception as e:
                import traceback
                error_details = traceback.format_exc()
                messages.error(request,
                               f'Error saving grievance record: {str(e)}')
                print(f"Grievance save error: {error_details}")
                print("=== SAVE ERROR ===")
    else:
        form = GrievianceMonitoringLogForm()

    return render(request, 'social_and_env/grievance/grievance_add_form.html',
                  {
                      'form': form,
                      'title': 'Add New Grievance Record'
                  })


@login_required
def grievance_edit(request, pk):
    """Edit Grievance record"""
    grievance = get_object_or_404(GrievianceMonitoringLog, pk=pk)

    if request.method == 'POST':
        form = GrievianceUpdateForm(request.POST, instance=grievance)
        if form.is_valid():
            try:
                grievance = form.save(commit=False)
                grievance.loginUser = request.user
                grievance.save()
                messages.success(request,
                                 'Grievance record updated successfully.')
                return redirect('grievance_list')
            except Exception as e:
                messages.error(request,
                               f'Error updating grievance record: {str(e)}')
    else:
        form = GrievianceUpdateForm(instance=grievance)

    return render(request, 'social_and_env/grievance/grievance_edit_form.html',
                  {
                      'form': form,
                      'grievance': grievance,
                      'title': 'Edit Grievance Record'
                  })


@login_required
def grievance_delete(request, pk):
    """Delete Grievance record"""
    grievance = get_object_or_404(GrievianceMonitoringLog, pk=pk)

    if request.method == 'POST':
        try:
            grievance.delete()
            messages.success(request, 'Grievance record deleted successfully.')
            return redirect('grievance_list')
        except Exception as e:
            messages.error(request,
                           f'Error deleting grievance record: {str(e)}')

    return render(request,
                  'social_and_env/grievance/grievance_confirm_delete.html', {
                      'grievance': grievance,
                      'title': 'Delete Grievance Record'
                  })


# ======================== OHS Views ========================
@login_required
def ohs_list(request):
    """OHS list view with Django filtering - following PAP Management pattern"""
    from django.core.paginator import Paginator
    from django.db.models import Sum
    from .filters import OHSMonitoringFilter

    try:
        print("=== OHS LIST DEBUG START ===")

        # Get total record count for debugging
        total_db_records = OHS_Monitoring.objects.count()
        print(f"Total OHS records in database: {total_db_records}")

        # Get OHS records with relationships - following PAP pattern
        ohs_list = OHS_Monitoring.objects.select_related(
            'project', 'Type_of_Investment', 'region', 'district', 'settlement',
            'year_of_report', 'quarter', 'loginUser'
        ).all()

        print(f"OHS queryset count: {ohs_list.count()}")

        # Apply Django filters
        ohs_filter = OHSMonitoringFilter(request.GET, queryset=ohs_list)
        filtered_ohs = ohs_filter.qs

        print(f"Final OHS data count: {filtered_ohs.count()}")
        print("=== OHS LIST DEBUG END ===")

        # Pagination - following PAP pattern
        page_size = request.GET.get('page_size', 10)
        try:
            page_size = int(page_size)
            if page_size not in [10, 25, 50, 100]:
                page_size = 10
        except (ValueError, TypeError):
            page_size = 10

        paginator = Paginator(filtered_ohs, page_size)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        # Statistics - following PAP pattern
        stats = {
            'total_ohs': ohs_list.count(),
            'filtered_count': filtered_ohs.count(),
            'total_workers': (filtered_ohs.aggregate(Sum('male'))['male__sum'] or 0) + 
                           (filtered_ohs.aggregate(Sum('female'))['female__sum'] or 0),
            'total_male': filtered_ohs.aggregate(Sum('male'))['male__sum'] or 0,
            'total_female': filtered_ohs.aggregate(Sum('female'))['female__sum'] or 0,
            'total_youth': (filtered_ohs.aggregate(Sum('youth_male'))['youth_male__sum'] or 0) + 
                          (filtered_ohs.aggregate(Sum('youth_female'))['youth_female__sum'] or 0),
        }

        context = {
            'page_obj': page_obj,
            'filter': ohs_filter,
            'stats': stats,
            'is_filtered': bool(request.GET),
            'title': 'OHS Monitoring'
        }

        return render(request, 'social_and_env/ohs/ohs_list.html', context)

    except Exception as e:
        messages.error(request, f'Error loading OHS records: {str(e)}')
        return render(request, 'social_and_env/ohs/ohs_list.html', {
            'page_obj': None,
            'filter': OHSMonitoringFilter(),
            'stats': {},
            'is_filtered': False,
            'title': 'OHS Monitoring'
        })





# AJAX endpoints for cascading dropdowns
def load_districts(request):
    """Load districts based on selected region - HTMX compatible"""
    from django.http import JsonResponse, HttpResponse
    from setup.models import Districts, Regions

    region_id = request.GET.get('region_id') or request.GET.get(
        'region') or request.GET.get('id_region')

    # Detect HTMX request
    is_htmx = request.headers.get('HX-Request') == 'true'

    if not region_id:
        if is_htmx:
            return HttpResponse('<option value="">Select District</option>')
        return JsonResponse({'districts': []})

    try:
        # Get districts by region
        try:
            region = Regions.objects.get(pk=region_id)
            districts = Districts.objects.filter(
                region_code_id=region_id).order_by('district_name')
        except Regions.DoesNotExist:
            districts = Districts.objects.none()

        if is_htmx:
            # Return HTML options for HTMX
            options = '<option value="">Select District</option>'
            for district in districts:
                options += f'<option value="{district.pk}">{district.district_name}</option>'
            return HttpResponse(options)
        else:
            # Return JSON for regular AJAX
            district_list = [{
                'id': district.pk,
                'name': district.district_name
            } for district in districts]
            return JsonResponse({'districts': district_list})

    except Exception as e:
        if is_htmx:
            return HttpResponse(f'<option value="">Error: {str(e)}</option>')
        return JsonResponse({'districts': [], 'error': str(e)})


def load_settlements(request):
    """Load settlements based on selected district - HTMX compatible"""
    from django.http import JsonResponse, HttpResponse
    from setup.models import Settlement

    district_id = request.GET.get('district_id') or request.GET.get(
        'district') or request.GET.get('id_district')

    # Detect HTMX request
    is_htmx = request.headers.get('HX-Request') == 'true'

    if not district_id:
        if is_htmx:
            return HttpResponse('<option value="">Select Settlement</option>')
        return JsonResponse({'settlements': []})

    try:
        settlements = Settlement.objects.filter(
            district_code_id=district_id).order_by('settlement_name')

        if is_htmx:
            # Return HTML options for HTMX
            options = '<option value="">Select Settlement</option>'
            for settlement in settlements:
                options += f'<option value="{settlement.pk}">{settlement.settlement_name}</option>'
            return HttpResponse(options)
        else:
            # Return JSON for regular AJAX
            settlement_list = [{
                'id': settlement.pk,
                'name': settlement.settlement_name
            } for settlement in settlements]
            return JsonResponse({'settlements': settlement_list})

    except Exception as e:
        if is_htmx:
            return HttpResponse(f'<option value="">Error: {str(e)}</option>')
        return JsonResponse({'settlements': [], 'error': str(e)})


def load_kpi_descriptions(request):
    """Load KPI descriptions based on selected project for AJAX requests"""
    from django.http import JsonResponse
    from PIU_Financial_mgt.models import KPI_For_Contract

    project_id = request.GET.get('project_id')
    # Remove debug output for production

    if project_id:
        # Use project_id directly as foreign key lookup since project field expects the primary key
        kpi_contracts = KPI_For_Contract.objects.filter(
            project_id=project_id).distinct()

        kpi_list = []
        for kpi in kpi_contracts:
            if kpi.Kpi_description and kpi.Kpi_description.strip():
                kpi_list.append({
                    'id': kpi.monitoring_Type_Code,
                    'description': kpi.Kpi_description.strip()
                })

        return JsonResponse({'kpi_descriptions': kpi_list})
    return JsonResponse({'kpi_descriptions': []})


@login_required
def load_investment_types_esia(request):
    """Load investment types for ESIA form based on selected project - HTMX compatible"""
    from django.http import JsonResponse, HttpResponse
    from PIU_Financial_mgt.models import KPI_For_Contract

    project_id = request.GET.get('project_name') or request.GET.get('project_id')
    
    # Detect HTMX request
    is_htmx = request.headers.get('HX-Request') == 'true'

    try:
        if project_id:
            # Filter investment types by project for ESS (Environmental and Social Safeguards)
            investment_types = KPI_For_Contract.objects.filter(
                project__projectID=project_id,
                monitoring_type_id='ESS'
            ).values(
                'monitoring_Type_Code', 'type_of_investment').distinct()
        else:
            # Get all distinct investment types for ESS if no project selected
            investment_types = KPI_For_Contract.objects.filter(
                monitoring_type_id='ESS'
            ).values(
                'monitoring_Type_Code', 'type_of_investment').distinct()

        investment_list = [{
            'id': inv['monitoring_Type_Code'] or f'inv_{idx}',
            'name': inv['type_of_investment']
        } for idx, inv in enumerate(investment_types, 1)
                           if inv['type_of_investment']]

        if is_htmx:
            # Return HTML options for HTMX
            options = '<option value="">All Investment Types</option>'
            for inv in investment_list:
                options += f'<option value="{inv["id"]}">{inv["name"]}</option>'
            return HttpResponse(options)
        else:
            # Return JSON for regular AJAX
            return JsonResponse({'investment_types': investment_list})
            
    except Exception as e:
        if is_htmx:
            return HttpResponse(f'<option value="">Error: {str(e)}</option>')
        return JsonResponse({'investment_types': [], 'error': str(e)})


@login_required
def load_investment_types_grievance(request):
    """Load investment types for Grievance form based on selected project"""
    from django.http import JsonResponse
    from PIU_Financial_mgt.models import KPI_For_Contract

    project_id = request.GET.get('project_id')

    if project_id:
        try:
            # Filter KPI_For_Contract by project for ESS (Environmental and Social Safeguards)
            investment_types = KPI_For_Contract.objects.filter(
                project__projectID=project_id,
                monitoring_type_id='ESS').values(
                    'monitoring_Type_Code', 'type_of_investment').distinct()

            investment_list = [{
                'id': inv['monitoring_Type_Code'],
                'name': inv['type_of_investment']
            } for inv in investment_types if inv['type_of_investment']]

            return JsonResponse({'investment_types': investment_list})
        except Exception as e:
            return JsonResponse({'investment_types': [], 'error': str(e)})

    return JsonResponse({'investment_types': []})


@login_required
def load_investment_types_ohs(request):
    """Load investment types and KPI descriptions for OHS form based on selected project - HTMX compatible"""
    from django.http import JsonResponse, HttpResponse
    from PIU_Financial_mgt.models import KPI_For_Contract
    import logging

    logger = logging.getLogger(__name__)
    project_id = request.GET.get('project_id') or request.GET.get('project')
    logger.debug(f"OHS AJAX called with project_id: {project_id}")
    
    # Detect HTMX request
    is_htmx = request.headers.get('HX-Request') == 'true'

    try:
        if project_id:
            # Get investment types for specific project (using projectID field)
            investment_types = KPI_For_Contract.objects.filter(
                project__projectID=project_id).values(
                    'monitoring_Type_Code', 'type_of_investment').distinct()

            # Get KPI descriptions for the project (using correct field name: Kpi_description)
            kpi_descriptions = KPI_For_Contract.objects.filter(
                project__projectID=project_id).values(
                    'monitoring_Type_Code', 'Kpi_description').distinct()

            investment_list = [{
                'id': inv['monitoring_Type_Code'],
                'name': inv['type_of_investment']
            } for inv in investment_types if inv['type_of_investment']]

            kpi_list = [{
                'id': kpi['monitoring_Type_Code'],
                'name': kpi['Kpi_description']
            } for kpi in kpi_descriptions if kpi['Kpi_description']]

            logger.debug(
                f"Found {len(investment_list)} investments, {len(kpi_list)} KPIs"
            )

            if is_htmx:
                # Return HTML options for HTMX
                options = '<option value="">All Investment Types</option>'
                for inv in investment_list:
                    options += f'<option value="{inv["id"]}">{inv["name"]}</option>'
                return HttpResponse(options)
            else:
                # Return JSON for regular AJAX
                response_data = {
                    'investment_types': investment_list,
                    'kpi_descriptions': kpi_list
                }
                logger.debug(f"Returning JSON response: {response_data}")
                return JsonResponse(response_data)
        else:
            # No project selected - return empty lists
            logger.debug("No project_id provided, returning empty lists")
            if is_htmx:
                return HttpResponse('<option value="">All Investment Types</option>')
            else:
                return JsonResponse({
                    'investment_types': [],
                    'kpi_descriptions': []
                })

    except Exception as e:
        logger.error(f"Error in load_investment_types_ohs: {str(e)}")
        if is_htmx:
            return HttpResponse(f'<option value="">Error: {str(e)}</option>')
        return JsonResponse({
            'investment_types': [],
            'kpi_descriptions': [],
            'error': str(e)
        })


@login_required
def ohs_add(request):
    """Add new OHS record with improved form handling"""
    if request.method == 'POST':
        print("=== OHS ADD DEBUG START ===")
        print(f"POST data: {dict(request.POST)}")
        print(f"FILES data: {dict(request.FILES)}")

        form = OHSMonitoringForm(request.POST, request.FILES)
        print(f"Form created, is_valid: {form.is_valid()}")

        if not form.is_valid():
            print(f"Form errors: {form.errors}")
            print(f"Non-field errors: {form.non_field_errors()}")

            # Show form validation errors with better debugging
            messages.error(request, 'Please correct the following errors:')
            error_messages = []
            for field, errors in form.errors.items():
                for error in errors:
                    error_messages.append(f"{field}: {error}")
                    print(f"Field error - {field}: {error}")

            if form.non_field_errors():
                for error in form.non_field_errors():
                    error_messages.append(f"General: {error}")
                    print(f"Non-field error: {error}")

            if error_messages:
                messages.error(
                    request,
                    f'Form validation failed: {"; ".join(error_messages)}')
            print("=== OHS FORM VALIDATION FAILED ===")
        else:
            print("Form is valid, attempting to save...")
            try:
                ohs = form.save(commit=False)
                ohs.loginUser = request.user

                print(f"OHS object before save: {ohs.__dict__}")

                # Set default values for optional fields
                if not ohs.male:
                    ohs.male = 0
                if not ohs.female:
                    ohs.female = 0
                if not ohs.youth_male:
                    ohs.youth_male = 0
                if not ohs.youth_female:
                    ohs.youth_female = 0

                # Debug investment type assignment
                if hasattr(ohs, 'Type_of_Investment'):
                    print(
                        f"Investment type assigned: {ohs.Type_of_Investment}")
                else:
                    print("No investment type assigned")

                ohs.save()
                print("OHS record saved successfully!")
                messages.success(request, 'OHS record added successfully.')

                # Redirect to list with success message
                return redirect('ohs_list')
            except Exception as e:
                import traceback
                error_details = traceback.format_exc()
                messages.error(request, f'Error saving OHS record: {str(e)}')
                print(f"OHS save error: {error_details}")
                print("=== OHS SAVE ERROR ===")
    else:
        form = OHSMonitoringForm()

    return render(request, 'social_and_env/ohs/ohs_form.html', {
        'form': form,
        'title': 'Add OHS Record'
    })


@login_required
def load_kpi_descriptions_ohs(request):
    """Load KPI descriptions for OHS form based on selected project"""
    from django.http import JsonResponse
    from PIU_Financial_mgt.models import KPI_For_Contract

    project_id = request.GET.get('project_id')

    try:
        if project_id:
            # Get KPI descriptions for the project (using correct field name: Kpi_description)
            kpi_descriptions = KPI_For_Contract.objects.filter(
                project__projectID=project_id).values(
                    'monitoring_Type_Code', 'Kpi_description').distinct()

            kpi_list = [{
                'id': kpi['monitoring_Type_Code'],
                'name': kpi['Kpi_description']
            } for kpi in kpi_descriptions if kpi['Kpi_description']]

            return JsonResponse({'kpi_descriptions': kpi_list})
        else:
            return JsonResponse({'kpi_descriptions': []})

    except Exception as e:
        return JsonResponse({'kpi_descriptions': [], 'error': str(e)})


@login_required
def ohs_detail(request, pk):
    """Detail view for OHS monitoring record"""
    try:
        ohs = get_object_or_404(OHS_Monitoring.objects.select_related(
            'project', 'Type_of_Investment', 'year_of_report', 'quarter',
            'region', 'district', 'settlement', 'loginUser'),
                                pk=pk)

        context = {
            'ohs':
            ohs,
            'title':
            f'OHS Monitoring - {ohs.project.project if ohs.project else "Unknown"}',
        }

        return render(request, 'social_and_env/ohs/ohs_detail.html', context)

    except Exception as e:
        messages.error(request, f'Error loading OHS details: {str(e)}')
        return redirect('ohs_list')


@login_required
def ohs_edit(request, pk):
    """Edit OHS monitoring record with improved handling"""
    try:
        ohs = get_object_or_404(OHS_Monitoring.objects.select_related(
            'project', 'Type_of_Investment', 'year_of_report', 'quarter',
            'region', 'district', 'settlement'),
                                pk=pk)
    except OHS_Monitoring.DoesNotExist:
        messages.error(request, 'OHS record not found.')
        return redirect('ohs_list')

    if request.method == 'POST':
        form = OHSMonitoringForm(request.POST, request.FILES, instance=ohs)
        if form.is_valid():
            try:
                ohs = form.save(commit=False)
                ohs.loginUser = request.user

                # Ensure numeric fields have valid values
                if not ohs.male:
                    ohs.male = 0
                if not ohs.female:
                    ohs.female = 0
                if not ohs.youth_male:
                    ohs.youth_male = 0
                if not ohs.youth_female:
                    ohs.youth_female = 0

                ohs.save()
                messages.success(request, 'OHS record updated successfully.')
                return redirect('ohs_list')
            except ValidationError as e:
                messages.error(request, f'Validation error: {str(e)}')
            except Exception as e:
                messages.error(request, f'Error updating OHS record: {str(e)}')
        else:
            # Show form validation errors
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        form = OHSMonitoringForm(instance=ohs)

    return render(request, 'social_and_env/ohs/ohs_form.html', {
        'form': form,
        'ohs': ohs,
        'title': 'Edit OHS Record'
    })


@login_required
def ohs_delete(request, pk):
    """Delete OHS monitoring record with improved error handling"""
    try:
        ohs = get_object_or_404(OHS_Monitoring.objects.select_related(
            'project', 'Type_of_Investment', 'region', 'district',
            'settlement'),
                                pk=pk)
    except OHS_Monitoring.DoesNotExist:
        messages.error(request, 'OHS record not found.')
        return redirect('ohs_list')

    if request.method == 'POST':
        try:
            ohs_project = ohs.project.project if ohs.project else "Unknown"
            ohs_id = ohs.ohs_Id
            ohs.delete()
            messages.success(
                request,
                f'OHS record OHS-{ohs_id} for project {ohs_project} deleted successfully.'
            )
            return redirect('ohs_list')
        except Exception as e:
            messages.error(request, f'Error deleting OHS record: {str(e)}')
            return redirect('ohs_detail', pk=pk)

    return render(request, 'social_and_env/ohs/ohs_confirm_delete.html', {
        'ohs': ohs,
        'title': 'Delete OHS Record'
    })


@login_required
def export_ohs_excel(request):
    """Export OHS monitoring data to Excel"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from django.http import HttpResponse
    import io

    try:
        # Create workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "OHS Monitoring Data"

        # Headers
        headers = [
            'OHS ID', 'Project', 'Investment Type', 'Year', 'Quarter', 'Date',
            'Region', 'District', 'Settlement', 'Male Workers',
            'Female Workers', 'Youth Male', 'Youth Female', 'Total Workers',
            'Quality Requirements', 'Working Environment', 'Remarks',
            'Created By', 'Created Date'
        ]

        # Style headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092",
                                  end_color="366092",
                                  fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Get OHS data with related fields
        ohs_records = OHS_Monitoring.objects.select_related(
            'project', 'Type_of_Investment', 'year_of_report', 'quarter',
            'region', 'district', 'settlement', 'loginUser').order_by('-date')

        # Add data rows
        for row, ohs in enumerate(ohs_records, 2):
            ws.cell(row=row, column=1, value=f"OHS-{ohs.ohs_Id}")
            ws.cell(row=row,
                    column=2,
                    value=ohs.project.project if ohs.project else "N/A")
            ws.cell(row=row,
                    column=3,
                    value=ohs.Type_of_Investment.monitoring_Type_Code
                    if ohs.Type_of_Investment else "N/A")
            ws.cell(row=row,
                    column=4,
                    value=ohs.year_of_report.year_name
                    if ohs.year_of_report else "N/A")
            ws.cell(row=row,
                    column=5,
                    value=ohs.quarter.quarter_name if ohs.quarter else "N/A")
            ws.cell(row=row,
                    column=6,
                    value=ohs.date.strftime('%Y-%m-%d') if ohs.date else "N/A")
            ws.cell(row=row,
                    column=7,
                    value=ohs.region.region_name if ohs.region else "N/A")
            ws.cell(
                row=row,
                column=8,
                value=ohs.district.district_name if ohs.district else "N/A")
            ws.cell(row=row,
                    column=9,
                    value=ohs.settlement.settlement_name
                    if ohs.settlement else "N/A")
            ws.cell(row=row, column=10, value=ohs.male or 0)
            ws.cell(row=row, column=11, value=ohs.female or 0)
            ws.cell(row=row, column=12, value=ohs.youth_male or 0)
            ws.cell(row=row, column=13, value=ohs.youth_female or 0)
            ws.cell(row=row, column=14, value=ohs.total_workers)
            ws.cell(row=row,
                    column=15,
                    value=ohs.quality_at_entry_requirement or "N/A")
            ws.cell(row=row, column=16, value=ohs.working_environment or "N/A")
            ws.cell(row=row, column=17, value=ohs.remarks or "N/A")
            ws.cell(row=row,
                    column=18,
                    value=ohs.loginUser.username if ohs.loginUser else "N/A")
            ws.cell(row=row,
                    column=19,
                    value=ohs.date_created.strftime('%Y-%m-%d %H:%M')
                    if ohs.date_created else "N/A")

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save to response
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type=
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response[
            'Content-Disposition'] = 'attachment; filename="OHS_Monitoring_Data.xlsx"'

        return response

    except Exception as e:
        messages.error(request, f'Error exporting OHS data: {str(e)}')
        return redirect('ohs_list')


# ======================== Community Engagement Views ========================
@login_required
def community_list(request):
    """Enhanced Community Engagement list view with filtering and pagination"""
    from django.core.paginator import Paginator
    from django.db.models import Sum

    try:
        # Use Django filter for consistency
        engagement_list = CommunityConsult_Engagement.objects.select_related(
            'project_name', 'stake_holder_engagement_Types', 'year',
            'loginUser').all()

        # Apply Django Filter
        engagement_filter = CommunityEngagementFilter(request.GET, queryset=engagement_list)
        engagements = engagement_filter.qs

        # Pagination
        page_size = request.GET.get('page_size', 10)
        try:
            page_size = int(page_size)
            if page_size not in [10, 15, 25, 50, 100]:
                page_size = 10
        except (ValueError, TypeError):
            page_size = 10

        paginator = Paginator(engagements, page_size)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        # Get statistics with correct field names
        stats = {
            'total_engagements':
            CommunityConsult_Engagement.objects.count(),
            'filtered_count':
            engagements.count(),
            'total_male_participants':
            engagements.aggregate(total=Sum('male'))['total'] or 0,
            'total_female_participants':
            engagements.aggregate(total=Sum('female'))['total'] or 0,
            'total_participants':
            engagements.aggregate(total=Sum('total_participants'))['total']
            or 0,
        }

        context = {
            'page_obj': page_obj,
            'filter': engagement_filter,
            'stats': stats,
            'is_filtered': bool(request.GET),
            'title': 'Community Engagement'
        }

        return render(request, 'social_and_env/community/community_list.html',
                      context)

    except Exception as e:
        messages.error(request,
                       f'Error loading community engagements: {str(e)}')
        # Get filter data for error case
        try:
            from PIU_Financial_mgt.models import Project
            from setup.models import YEAR
            projects = Project.objects.filter(
                project='National Electricity and Water Company')
            years = YEAR.objects.all()
            engagement_types = TypeOfStakeholderEngagement.objects.all()
        except:
            projects = []
            years = []
            engagement_types = []

        return render(
            request, 'social_and_env/community/community_list.html', {
                'page_obj': None,
                'stats': {},
                'projects': projects,
                'years': years,
                'engagement_types': engagement_types,
                'title': 'Community Engagement'
            })


@login_required
def community_add(request):
    """Add new Community Engagement record"""
    if request.method == 'POST':
        form = CommunityEngagementForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                engagement = form.save(commit=False)
                engagement.loginUser = request.user
                engagement.save()
                messages.success(
                    request, 'Community engagement record added successfully.')
                return redirect('community_list')
            except Exception as e:
                messages.error(
                    request,
                    f'Error saving community engagement record: {str(e)}')
    else:
        form = CommunityEngagementForm()

    return render(request, 'social_and_env/community/community_form.html', {
        'form': form,
        'title': 'Add Community Engagement Record'
    })


@login_required
def community_detail(request, pk):
    """Community Engagement detail view"""
    try:
        engagement = get_object_or_404(
            CommunityConsult_Engagement.objects.select_related(
                'project_name', 'stake_holder_engagement_Types', 'year',
                'loginUser'),
            pk=pk)

        context = {
            'engagement':
            engagement,
            'title':
            f'Community Engagement - {engagement.project_name.project if engagement.project_name else "Unknown"}',
        }

        return render(request,
                      'social_and_env/community/community_detail.html',
                      context)

    except Exception as e:
        messages.error(
            request, f'Error loading community engagement details: {str(e)}')
        return redirect('community_list')


@login_required
def community_edit(request, pk):
    """Edit Community Engagement record"""
    engagement = get_object_or_404(CommunityConsult_Engagement, pk=pk)

    if request.method == 'POST':
        form = CommunityEngagementForm(request.POST,
                                       request.FILES,
                                       instance=engagement)
        if form.is_valid():
            try:
                engagement = form.save(commit=False)
                engagement.loginUser = request.user
                engagement.save()
                messages.success(
                    request,
                    'Community engagement record updated successfully.')
                return redirect('community_list')
            except Exception as e:
                messages.error(
                    request,
                    f'Error updating community engagement record: {str(e)}')
    else:
        form = CommunityEngagementForm(instance=engagement)

    return render(request, 'social_and_env/community/community_form.html', {
        'form': form,
        'title': 'Edit Community Engagement Record'
    })


@login_required
def community_delete(request, pk):
    """Delete Community Engagement record"""
    engagement = get_object_or_404(CommunityConsult_Engagement, pk=pk)

    if request.method == 'POST':
        try:
            engagement.delete()
            messages.success(
                request, 'Community engagement record deleted successfully.')
            return redirect('community_list')
        except Exception as e:
            messages.error(
                request,
                f'Error deleting community engagement record: {str(e)}')

    return render(request,
                  'social_and_env/community/community_confirm_delete.html', {
                      'engagement': engagement,
                      'title': 'Delete Community Engagement Record'
                  })


# ======================== ESIA Views ========================
@login_required
def esia_list(request):
    """Enhanced ESIA list view with filtering and pagination"""
    from django.core.paginator import Paginator

    try:
        esias = ESIA.objects.select_related('project_name',
                                            'type_of_investment',
                                            'loginUser').all()

        # Apply filters
        if request.GET.get('project'):
            esias = esias.filter(project_name=request.GET.get('project'))

        if request.GET.get('region'):
            # Since ESIA doesn't have region field, we'll skip this filter
            pass

        # Pagination
        page_size = request.GET.get('page_size', 10)
        try:
            page_size = int(page_size)
            if page_size not in [10, 15, 25, 50, 100]:
                page_size = 10
        except (ValueError, TypeError):
            page_size = 10

        paginator = Paginator(esias, page_size)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        # Get statistics
        stats = {
            'total_esia':
            ESIA.objects.count(),
            'filtered_count':
            esias.count(),
            'total_communities':
            esias.aggregate(total=Sum('number_of_communities'))['total'] or 0,
        }

        # Get filter data
        from PIU_Financial_mgt.models import Project
        projects = Project.objects.all()
        regions = Regions.objects.all()
        years = [record.date_created.year for record in ESIA.objects.all()]
        years = sorted(list(set(years)))

        context = {
            'page_obj': page_obj,
            'stats': stats,
            'projects': projects,
            'regions': regions,
            'years': years,
            'title': 'ESIA/ESMP Management'
        }

        return render(request, 'social_and_env/esia/esia_list.html', context)

    except Exception as e:
        messages.error(request, f'Error loading ESIA/ESMP records: {str(e)}')
        # Get filter data for error case
        from PIU_Financial_mgt.models import Project
        projects = Project.objects.all()
        regions = Regions.objects.all()
        years = []

        return render(
            request, 'social_and_env/esia/esia_list.html', {
                'page_obj': None,
                'stats': {},
                'projects': projects,
                'regions': regions,
                'years': years,
                'title': 'ESIA/ESMP Management'
            })


@login_required
def esia_add(request):
    """Add new ESIA/ESMP record"""
    if request.method == 'POST':
        form = ESIAForm(request.POST)
        if form.is_valid():
            try:
                esia = form.save(commit=False)
                esia.loginUser = request.user
                esia.save()
                messages.success(request,
                                 'ESIA/ESMP record added successfully.')
                return redirect('esia_list')
            except Exception as e:
                messages.error(request,
                               f'Error saving ESIA/ESMP record: {str(e)}')
        else:
            # Debug form errors for validation issues
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        form = ESIAForm()

    return render(request, 'social_and_env/esia/esia_form.html', {
        'form': form,
        'title': 'Add ESIA/ESMP Record'
    })


@login_required
def esia_detail(request, pk):
    """ESIA/ESMP detail view"""
    try:
        esia = get_object_or_404(ESIA.objects.select_related(
            'project_name', 'type_of_investment', 'loginUser'),
                                 pk=pk)

        context = {
            'esia':
            esia,
            'title':
            f'ESIA/ESMP Details - {esia.project_name.project if esia.project_name else "Unknown"}',
        }

        return render(request, 'social_and_env/esia/esia_detail.html', context)

    except Exception as e:
        messages.error(request, f'Error loading ESIA/ESMP details: {str(e)}')
        return redirect('esia_list')


@login_required
def esia_delete(request, pk):
    """Delete ESIA record"""
    esia = get_object_or_404(ESIA, pk=pk)

    if request.method == 'POST':
        try:
            esia.delete()
            messages.success(request, 'ESIA record deleted successfully.')
            return redirect('esia_list')
        except Exception as e:
            messages.error(request, f'Error deleting ESIA record: {str(e)}')

    return render(request, 'social_and_env/esia/esia_confirm_delete.html', {
        'esia': esia,
        'title': 'Delete ESIA Record'
    })


# ======================== Additional AJAX Views ========================
@login_required
def load_investment_types(request):
    """Load investment types based on project selection"""
    from django.http import HttpResponse

    project_id = request.GET.get('project')
    if not project_id:
        return HttpResponse('<option value="">Select Investment Type</option>')

    try:
        investment_types = KPI_For_Contract.objects.filter(
            project=project_id).values_list(
                'type_of_investment',
                flat=True).distinct().order_by('type_of_investment')

        options = '<option value="">Select Investment Type</option>'
        for investment_type in investment_types:
            options += f'<option value="{investment_type}">{investment_type}</option>'

        return HttpResponse(options)

    except Exception as e:
        return HttpResponse(f'<option value="">Error: {str(e)}</option>')


@login_required
def load_investment_types_grievance(request):
    """Load investment types for Grievance based on project selection"""
    from django.http import HttpResponse

    project_id = request.GET.get('project')
    if not project_id:
        return HttpResponse('<option value="">Select Investment Type</option>')

    try:
        # Filter for ESS monitoring types specifically for grievance management
        investment_types = KPI_For_Contract.objects.filter(
            project=project_id,
            monitoring_type_id='ESS').order_by('type_of_investment')

        options = '<option value="">Select Investment Type</option>'
        for investment in investment_types:
            # Use monitoring_Type_Code as value since form expects it
            options += f'<option value="{investment.monitoring_Type_Code}">{investment.type_of_investment}</option>'

        return HttpResponse(options)

    except Exception as e:
        return HttpResponse(f'<option value="">Error: {str(e)}</option>')


@login_required
def load_settlements_ohs(request):
    """Load settlements for OHS based on selected district"""
    from django.http import HttpResponse

    district_id = request.GET.get('district')
    if not district_id:
        return HttpResponse('<option value="">Select Settlement</option>')

    try:
        settlements = Settlement.objects.filter(
            district_code=district_id).values_list(
                'settlement_code',
                'settlement_name').order_by('settlement_name')

        options = '<option value="">Select Settlement</option>'
        for settlement_code, settlement_name in settlements:
            options += f'<option value="{settlement_code}">{settlement_name}</option>'

        return HttpResponse(options)

    except Exception as e:
        return HttpResponse(f'<option value="">Error: {str(e)}</option>')


@login_required
def test_cascading_dropdown(request):
    """Test endpoint to validate cascading dropdown functionality"""
    from django.http import JsonResponse

    try:
        # Test data for cascading dropdowns
        test_data = {
            'status':
            'success',
            'message':
            'Cascading dropdown test endpoint working',
            'regions':
            list(Regions.objects.values('region_code', 'region_name')),
            'districts':
            list(
                Districts.objects.values('district_code', 'district_name',
                                         'region_code')),
            'settlements':
            list(
                Settlement.objects.values('settlement_code', 'settlement_name',
                                          'district_code')),
            'projects':
            list(Project.objects.values('project', 'project_name')),
        }

        return JsonResponse(test_data)

    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})


@login_required
def load_investment_types_pap(request):
    """Load investment types for PAP based on project selection - Enhanced for filter forms"""
    from django.http import HttpResponse
    from django.template.loader import render_to_string
    from urllib.parse import unquote

    # Get project_id and handle URL encoding (especially for & characters)
    project_id = request.GET.get('project') or request.GET.get(
        'project_id') or request.GET.get('id_project')
    
    if project_id:
        # Decode URL-encoded characters (e.g., %26 becomes &)
        project_id = unquote(project_id)

    # Get the selected value to maintain state
    selected_value = request.GET.get('type_of_investment', '')

    try:
        # Debug: Log the project_id being searched
        print(f"🔍 PAP Investment Types Filter - Searching for project: '{project_id}'")
        
        if project_id:
            # Get investment types from database only
            # Handle project lookup by exact projectID match
            investment_types = KPI_For_Contract.objects.filter(
                project__projectID=project_id,
                monitoring_type_id='ESS').distinct().order_by('type_of_investment')

            print(f"🔍 Found {investment_types.count()} investment types")
        else:
            # No project selected, show empty list
            investment_types = KPI_For_Contract.objects.none()
            print(f"🔍 No project selected, showing empty investment types")
        
        # Render using template for consistency
        context = {
            'investment_types': investment_types,
            'selected_value': selected_value,
        }
        
        html = render_to_string('social_and_env/htmx/investment_type_filter.html', context)
        return HttpResponse(html)

    except Exception as e:
        error_msg = f'Database error: {str(e)}'
        print(f"🔍 {error_msg}")
        # Return basic HTML on error
        html = f'''
        <label class="form-label">Type of Investment</label>
        <select name="type_of_investment" class="form-select" id="id_type_of_investment">
            <option value="">Error loading investment types</option>
        </select>
        '''
        return HttpResponse(html)


@login_required
def load_investment_types_esia(request):
    """Load investment types for ESIA based on project selection"""
    from django.http import HttpResponse

    project_id = request.GET.get('project')
    if not project_id:
        return HttpResponse('<option value="">Select Investment Type</option>')

    try:
        # Get full objects to access monitoring_Type_Code for form validation
        investment_objects = KPI_For_Contract.objects.filter(
            project=project_id,
            monitoring_type_id='ESS'
        ).distinct().order_by('type_of_investment')

        options = '<option value="">Select Investment Type</option>'
        for investment in investment_objects:
            options += f'<option value="{investment.monitoring_Type_Code}">{investment.type_of_investment}</option>'

        # If no options found, provide offline fallback
        if not investment_objects.exists():
            fallback_options = [
                ('ESS001', 'Environmental Impact Assessment'),
                ('ESS002', 'Social Impact Assessment'),
                ('ESS003', 'Biodiversity Assessment'),
                ('ESS004', 'Water Quality Assessment'),
                ('ESS005', 'Air Quality Assessment'),
                ('ESS006', 'Soil Assessment'),
                ('ESS007', 'Noise Assessment'),
                ('ESS008', 'Cultural Heritage Assessment'),
                ('ESS009', 'Community Health Assessment'),
                ('ESS010', 'Stakeholder Engagement'),
            ]
            for code, name in fallback_options:
                options += f'<option value="{code}">{name}</option>'

        return HttpResponse(options)

    except Exception as e:
        # Provide fallback options even on error
        fallback_options = '<option value="">Select Investment Type</option>'
        fallback_data = [
            ('ESS001', 'Environmental Impact Assessment'),
            ('ESS002', 'Social Impact Assessment'),
            ('ESS003', 'Biodiversity Assessment'),
            ('ESS004', 'Water Quality Assessment'),
            ('ESS005', 'Air Quality Assessment'),
        ]
        for code, name in fallback_data:
            fallback_options += f'<option value="{code}">{name}</option>'
        return HttpResponse(fallback_options)


@login_required
def esia_export_excel(request):
    """Export ESIA data to Excel"""
    import openpyxl
    from django.http import HttpResponse

    try:
        # Get all ESIA records
        esias = ESIA.objects.select_related('project', 'Type_of_Investment',
                                            'year_of_report', 'quarter',
                                            'loginUser').all()

        # Create workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'ESIA Data'

        # Add headers
        headers = [
            'Project', 'Investment Type', 'Year', 'Quarter', 'Region',
            'Number of Communities', 'ESIA Findings', 'Created By',
            'Date Created'
        ]

        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        # Add data rows
        for row, esia in enumerate(esias, 2):
            ws.cell(row=row,
                    column=1,
                    value=esia.project.project if esia.project else '')
            ws.cell(row=row,
                    column=2,
                    value=esia.Type_of_Investment.type_of_investment
                    if esia.Type_of_Investment else '')
            ws.cell(
                row=row,
                column=3,
                value=esia.year_of_report.year if esia.year_of_report else '')
            ws.cell(row=row,
                    column=4,
                    value=esia.quarter.quarter if esia.quarter else '')
            ws.cell(row=row, column=5, value=esia.region)
            ws.cell(row=row, column=6, value=esia.number_of_communities)
            ws.cell(row=row, column=7, value=esia.esia_findings)
            ws.cell(row=row,
                    column=8,
                    value=esia.loginUser.username if esia.loginUser else '')
            ws.cell(row=row,
                    column=9,
                    value=esia.date_created.strftime('%Y-%m-%d')
                    if esia.date_created else '')

        # Set up response
        response = HttpResponse(
            content_type=
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response[
            'Content-Disposition'] = 'attachment; filename="esia_data.xlsx"'

        wb.save(response)
        return response

    except Exception as e:
        messages.error(request, f'Error exporting ESIA data: {str(e)}')
        return redirect('esia_list')








@login_required
def export_grievance_excel(request):
    """Export Grievance data to Excel"""
    import openpyxl
    from django.http import HttpResponse
    from datetime import datetime

    try:
        # Get all grievance records
        grievances = GrievianceMonitoringLog.objects.select_related(
            'project', 'type_of_investment', 'decision_outcome', 'loginUser'
        ).all()

        # Create workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Grievance Data'

        # Add headers
        headers = [
            'Case No', 'Project', 'Investment Type', 'Complainant Name', 'Gender',
            'Phone Number', 'Date Received', 'Received By', 'How Received',
            'Complaint Content', 'Acknowledgment Sent', 'Expected Decision Date',
            'Decision Outcome', 'Decision Communicated', 'Communication Method',
            'Complainant Satisfied', 'Follow-up Action', 'Created By', 'Date Created'
        ]

        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        # Add data rows
        for row, grievance in enumerate(grievances, 2):
            ws.cell(row=row, column=1, value=grievance.case_no)
            ws.cell(row=row, column=2, value=grievance.project.project if grievance.project else '')
            ws.cell(row=row, column=3, value=grievance.type_of_investment.type_of_investment if grievance.type_of_investment else '')
            ws.cell(row=row, column=4, value=grievance.name_of_complainant)
            ws.cell(row=row, column=5, value=grievance.sex)
            ws.cell(row=row, column=6, value=grievance.tell_no)
            ws.cell(row=row, column=7, value=grievance.date_claim_recieved.strftime('%Y-%m-%d') if grievance.date_claim_recieved else '')
            ws.cell(row=row, column=8, value=grievance.name_of_person_receiving_complaint)
            ws.cell(row=row, column=9, value=grievance.how_complaint_was_received)
            ws.cell(row=row, column=10, value=grievance.complaint_content)
            ws.cell(row=row, column=11, value=grievance.was_recieved_of_complaint_ack)
            ws.cell(row=row, column=12, value=grievance.expected_decision_date.strftime('%Y-%m-%d') if grievance.expected_decision_date else '')
            ws.cell(row=row, column=13, value=str(grievance.decision_outcome) if grievance.decision_outcome else '')
            ws.cell(row=row, column=14, value=grievance.was_decison_communicated_to_complainant)
            ws.cell(row=row, column=15, value=grievance.communication_method)
            ws.cell(row=row, column=16, value=grievance.was_complainant_satisfied_with_decision)
            ws.cell(row=row, column=17, value=grievance.any_follow_up_action)
            ws.cell(row=row, column=18, value=grievance.loginUser.username if grievance.loginUser else '')
            ws.cell(row=row, column=19, value=grievance.date_created.strftime('%Y-%m-%d %H:%M') if grievance.date_created else '')

        # Create HTTP response with Excel file
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="grievance_data_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        
        wb.save(response)
        return response

    except Exception as e:
        messages.error(request, f'Error exporting grievance data: {str(e)}')
        return redirect('grievance_list')
